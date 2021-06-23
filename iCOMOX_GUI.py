from PIL import ImageTk, Image
import sys, os
import openpyxl.workbook
import openpyxl
import common_symbols
import messages_utils
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
from tkinter import font
# from tkinter import tix
import DialogBands
import serial.tools.list_ports
import tk_tools as tk_tools
import datetime
import time
import BinFileConversion
import numpy as np
import MemMap
import ipaddress
# import urllib.parse
if sys.platform.startswith("win"):
	import win32con
	import win32api
	# import win32gui

from widget_speedometer import Speedometer
from widget_thermometer import Thermometer
from widget_PlotFFT import PlotFFT

import statistics_data_logger
import iCOMOX_communication
import iCOMOX_messages
import iCOMOX_datahandling
import helpers
import iCOMOX_list
import common
import iCOMOX_over_Dongle_Communication
import single_app_instance
import TCP_connectivity
import iCOMOX_ClientsTreeView

class mainApp(tk.Frame):
	def __init__(self, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		tk.Frame.tk_setPalette(self, background='white')

		# Variables definition
		self.terminated = False
		self.SerialCommPorts = []
		self.CommPortsUSB = []
		self.CommPortsSMIP = []

		# Style definitions
		self.style = ttk.Style()
		# helpers.OUT(str(self.style.theme_names()))
		self.style.theme_use('default')
		# self.style.configure("TNotebook", background="white", font=('verdana', 8, "bold"))
		# self.style.configure("TNotebook.Tab", padding=8, font=('verdana', 8))
		self.style.configure("EEPROM.Horizontal.TProgressbar", foreground='black', background='green')
		self.style.configure("Infer.Horizontal.TProgressbar", foreground='grey', throughcolor="grey", background='green')
		self.style.configure("Train.Horizontal.TProgressbar", foreground='grey', throughcolor="grey", background='yellow')
		self.style.configure("Anomaly.Green.Horizontal.TProgressbar", foreground='grey', throughcolor="grey", background='green')
		self.style.configure("Anomaly.Yellow.Horizontal.TProgressbar", foreground='grey', throughcolor="grey", background='yellow')
		self.style.configure("Anomaly.Red.Horizontal.TProgressbar", foreground='grey', throughcolor="grey", background='red')

		# Define frames for GUI widgets
		# Top frame
		top_pane = tk.Frame(bd=0, relief=tk.RIDGE)
		top_pane.grid(row=0, sticky="nsew")
		# Tabs frame
		tab_parent = ttk.Notebook()
		tab1 = tk.Frame(tab_parent)
		tab2 = tk.Frame(tab_parent)
		tab3 = tk.Frame(tab_parent)
		tab4 = tk.Frame(tab_parent)
		tab5 = tk.Frame(tab_parent)
		tab6 = tk.Frame(tab_parent)
		tab7 = tk.Frame(tab_parent)
		tab_parent.add(tab1, text="Live Data")
		tab_parent.add(tab2, text="Statistics")
		tab_parent.add(tab3, text="Configuration")
		tab_parent.add(tab4, text="Information")
		tab_parent.add(tab5, text="E²PROM")
		tab_parent.add(tab6, text="Clients")
		tab_parent.add(tab7, text="Diagnostic mode")

		tab_parent.grid(row=1, sticky="nsew")
		#tab_parent.select(tab_id=tab1)

		tab_parent.select(tab_id=tab3)   # Open Configuration tab

		# Bottom frame
		bottom_pane = tk.Frame(bd=0, relief=tk.RIDGE)
		bottom_pane.grid(row=2, pady=3, padx=5, sticky="nsew")

		# Define GUI objects
		self.Notebook = tab_parent
		self.TopPane = TopPane(top_pane)
		self.StatusBar = StatusBar(bottom_pane)
		self.LiveData = LiveData(tab1)
		self.Statistics = Statistics(tab2)
		self.Configuration = Configuration(tab3)
		self.Information = Information(tab4)
		self.EEPROM = EEPROM(tab5)
		self.Clients = Clients(tab6)
		self.DiagnosticMode = DiagnosticMode(tab7)

		self.TopPane.buttonConnect.configure(command=self.buttonConnectCommand)  # Command configuration
		self.TopPane.ComPortsListRefreshButton.configure(command=self.refresh_CommPorts)

		self.refresh_CommPorts()
		self.statisticsLogger = statistics_data_logger.ClassStatisticalDataLogger()
		self.iCOMOX_Data = iCOMOX_datahandling.class_DataHandling()
		self.idle_update_counter = 3

	def EnableButtons(self, liveDataTab, diagnosticTab, configurationTab, eepromTab, clientsTab):
		self.LiveData.EnableButtons(enable=liveDataTab)
		self.DiagnosticMode.Enable(enable=diagnosticTab)
		self.Clients.EnableButtons(enable=clientsTab)   # buttons in Clients tabsheet
		self.EEPROM.EnableButtons(enable=eepromTab)     # buttons in EEPROM tabsheet
		self.Configuration.EnableButtons(enable=configurationTab)

	def clear_LiveData_Statistics_and_Information(self, clear_iCOMOX_only):
		if self.terminated:
			return
		self.statisticsLogger.clear()
		self.Statistics.update_all()
		self.LiveData.clear_all()
		self.Information.clear(clear_iCOMOX_only=clear_iCOMOX_only)

	def refresh_CommPorts(self):
		def sort_commPorts_key(s):  # returns the numerical part of strings like COM12
			i = 0
			while (i < len(s)) and not s[i].isdigit():
				i = i+1
			s = s[i:]
			if len(s) > 0:
				return int(s)
			else:
				return -1

		self.SerialCommPorts = serial.tools.list_ports.comports()
		self.CommPortsUSB = [*map(lambda commPort : commPort.device, [*filter(lambda commPort : (commPort.vid == 0x0403) and (commPort.pid == 0x6015), self.SerialCommPorts)])]
		self.CommPortsUSB.sort(key=sort_commPorts_key)
		self.CommPortsSMIP = [*map(lambda commPort: commPort.device, [*filter(lambda commPort: (commPort.vid == 0x0403) and (commPort.pid == 0x6011), self.SerialCommPorts)])]
		self.CommPortsSMIP.sort(key=sort_commPorts_key)
		self.CommPortsSMIP = self.CommPortsSMIP[3::4]
		self.update_CommPorts()

	def update_CommPorts(self):
		if self.TopPane.rg_usb_smip_tcpip.get() == iCOMOX_list.cCLIENT_TYPE_USB:
			CommPorts = self.CommPortsUSB
		elif common.app.TopPane.rg_usb_smip_tcpip.get() == iCOMOX_list.cCLIENT_TYPE_SMIP:
			CommPorts = self.CommPortsSMIP
		else:
			CommPorts = []

		PrevCommPort = self.TopPane.cb_usb.get()
		# i = 0
		# while i < len(self.CommPorts):
		#     commPort = self.CommPorts[0]
		#
		#     self.CommPortsSmip =
		self.TopPane.cb_usb.configure(values=CommPorts)
		if (PrevCommPort != "") and (PrevCommPort in CommPorts):
			index = CommPorts.index(PrevCommPort)
			self.TopPane.cb_usb.current(index)
		elif len(CommPorts) > 0:
			self.TopPane.cb_usb.current(0)
		else:
			self.TopPane.cb_usb.set("")

	def on_idle_task(self, root):
		if self.terminated:     # if application was terminated, then stop process messages
			return
		# Here we retrieve the messages from the ping pong buffer, as the GUI is now ready to process them
		try:
			if self.iCOMOX_Data is not None:
				self.iCOMOX_Data.process_report_messages_of_current_iCOMOX()
		# except Exception as ex:
		#     if hasattr(ex, "message"):
		#         helpers.OUT(ex.message)
		#     elif hasattr(ex, "args") and (len(ex.args) > 0):
		#         helpers.OUT(ex.args[0])
		#     else:
		#         helpers.OUT("Exception in on_idle_task()")
		finally:
			# Tk does not update the display when we use after_idle(), so we must call update(),
			#   in order to do it explicitly
			self.update()
			root.after_idle(self.on_idle_task, root)  # make sure the on_idle_task will be called again

	def get_current_report_channel(self):
		iCOMOX_connected = (common.app.iCOMOX_Data is not None) and (common.app.iCOMOX_Data.iCOMOX_Comm is not None) and (common.app.iCOMOX_Data.iCOMOX_Comm.is_open())
		Dongle_connected = (common.app.iCOMOX_Data is not None) and (common.app.iCOMOX_Data.Dongle_Comm is not None) and (common.app.iCOMOX_Data.Dongle_Comm.is_open())
		TcpIp_connected = (common.app.iCOMOX_Data is not None) and (common.app.iCOMOX_Data.TcpIp_Comm is not None)# and (len(common.app.iCOMOX_Data.TcpIp_Comm.clients) > 0)
		if iCOMOX_connected:
			return 0
		elif Dongle_connected:
			return 1
		elif TcpIp_connected:
			return 2
		else:
			return -1

	def drawConnectionState(self, connected, listen=False):
		if self is None:
			return
		if not connected:
			common.app.EnableButtons(liveDataTab=False, diagnosticTab=False, configurationTab=False, eepromTab=False, clientsTab=False)
			# common.app.iCOMOX_Data.schedule_stop()
			common.app.StatusBar.progressbar_hide()
			if listen:
				self.TopPane.buttonConnect.config(text="Disconnect")
			else:
				self.TopPane.buttonConnect.config(text="Connect")
			# state = "normal"
			# reset_state = "disabled"
		else:
			iCOMOX_connected = (common.iCOMOXs.current is not None) and (common.iCOMOXs.current.UniqueID() is not None)
			common.app.EnableButtons(
				liveDataTab=iCOMOX_connected,
				configurationTab=iCOMOX_connected,
				diagnosticTab=iCOMOX_connected and (common.iCOMOXs.current.board_type() == iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT),
				eepromTab=iCOMOX_connected,
				clientsTab=iCOMOX_connected and (common.iCOMOXs.current.Type == iCOMOX_list.cCLIENT_TYPE_TCPIP))
			self.LiveData.temp_widget.mintemp = int(self.LiveData.minTempEntry.get())
			self.LiveData.temp_widget.maxtemp = int(self.LiveData.maxTempEntry.get())
			self.LiveData.speed_widget.min_value = int(self.LiveData.minSpeedEntry.get())
			self.LiveData.speed_widget.max_value = int(self.LiveData.maxSpeedEntry.get())
			self.LiveData.speed_widget.draw()
			self.TopPane.buttonConnect.config(text="Disconnect")
			# state = "disabled"
			# reset_state = "normal"

		self.TopPane.State_USB_SMIP_TCPIP(connected=connected)
		common.app.TopPane.EnableRadioButtons(enable=not (connected or listen))
		# self.Configuration.button_reset.config(state=reset_state)
		# Schedule
		# self.Configuration.scheduleEntry.config(state=state)
		# self.Configuration.timeEntry.config(state=state)
		# self.Configuration.radio_stream.config(state=state)
		# self.Configuration.radio_schedule.config(state=state)

	def drawLedState(self, connection_state):
		#enable_TestMode = False
		if connection_state == iCOMOX_datahandling.CONNECTION_STATE_iCOMOX_Disconnect:
			self.StatusBar.set(text="Disconnected from iCOMOX")
			self.TopPane.set_led_status_color(1)  # RED LED
			# common.app.TopPane.cb_usb.configure(state="readonly")
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_iCOMOX_Connecting:
			self.TopPane.set_led_status_color(2)  # YELLOW LED
			self.StatusBar.set(text="Connecting to iCOMOX...")
			# common.app.TopPane.cb_usb.configure(state="disabled")
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_iCOMOX_Connected:
			self.StatusBar.set(text="Connected to iCOMOX")
			self.TopPane.set_led_status_color(3)  # GREEN LED
			#enable_TestMode = True
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_Dongle_Disconnected:
			self.StatusBar.set(text="Disconnected from dongle")
			self.TopPane.set_led_status_color(1)  # RED LED
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_Dongle_Connecting:
			self.StatusBar.set(text="Connecting to dongle...")
			self.TopPane.set_led_status_color(1)  # RED LED
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_Dongle_Connected:
			self.StatusBar.set(text="Connected to dongle")
			self.TopPane.set_led_status_color(2)  # Yellow LED
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_iCOMOX_Connected_via_Dongle:
			self.StatusBar.set(text="iCOMOX connected via the dongle")
			self.TopPane.set_led_status_color(3)    # Green LED
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_TcpIp_ServerDisconnected:
			self.StatusBar.set(text="TCP/IP server is not connected")
			self.TopPane.set_led_status_color(1)    # Red LED
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_TcpIp_ServerConnected:
			self.StatusBar.set(text="TCP/IP server is listening")
			self.TopPane.set_led_status_color(2)    # Yellow LED
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_TcpIp_ClientConnected:
			self.StatusBar.set(text="TCP/IP iComox is connected")
			self.TopPane.set_led_status_color(2)    # Yellow LED
		elif connection_state == iCOMOX_datahandling.CONNECTION_STATE_iCOMOX_Connected_via_TcpIp:
			self.StatusBar.set(text="iCOMOX connected via the TCP/IP socket")
			self.TopPane.set_led_status_color(3)    # Green LED

		#self.DiagnosticMode.enable_widgets(enable=enable_TestMode)

	def buttonConnectCommand(self):
		if self.iCOMOX_Data.iCOMOX_Comm.is_open() or self.iCOMOX_Data.Dongle_Comm.is_open() or self.iCOMOX_Data.TcpIp_Comm.is_listening():
			self.drawConnectionState(connected=True)
			# self.iCOMOX_Data.icomox_set_configuration(RawDataSensors=0)  # disable RawDataSensors from the iCOMOX
			self.iCOMOX_Data.iCOMOX_Comm.close()
			self.iCOMOX_Data.Dongle_Comm.close()
			self.iCOMOX_Data.TcpIp_Comm.close()
			common.app.EnableButtons(liveDataTab=False, configurationTab=False, clientsTab=False, diagnosticTab=False, eepromTab=False)
			# common.app.TopPane.EnableRadioButtons(enable=True)
			common.iCOMOXs.clear()
			self.iCOMOX_Data.on_dongle_updated_iCOMOX_list()
		else:
			if self.TopPane.rg_usb_smip_tcpip.get() == iCOMOX_list.cCLIENT_TYPE_USB:
				if (self.TopPane.cb_usb.current() < 0) or (self.TopPane.cb_usb.current() >= len(self.CommPortsUSB)):
					messagebox.showerror("Error", "Please select a valid USB port for iCOMOX")
					return
				commPort = self.CommPortsUSB[self.TopPane.cb_usb.current()]
			elif self.TopPane.rg_usb_smip_tcpip.get() == iCOMOX_list.cCLIENT_TYPE_SMIP:
				if (self.TopPane.cb_usb.current() < 0) or (self.TopPane.cb_usb.current() >= len(self.CommPortsSMIP)):
					messagebox.showerror("Error", "Please select a valid USB port for SMIP dongle")
					return
				commPort = self.CommPortsSMIP[self.TopPane.cb_usb.current()]
			else:
				commPort = None
			self.drawConnectionState(connected=False)
			self.clear_LiveData_Statistics_and_Information(clear_iCOMOX_only=False)
			# self.iCOMOX_Data.MAG_ENERGY = 0
			if self.TopPane.rg_usb_smip_tcpip.get() == iCOMOX_list.cCLIENT_TYPE_USB:
				self.iCOMOX_Data.iCOMOX_Comm.open(commPort=commPort)
			elif self.TopPane.rg_usb_smip_tcpip.get() == iCOMOX_list.cCLIENT_TYPE_SMIP:
				# self.TopPane.cb_smip.configure(values=[])
				# self.TopPane.cb_smip.set("")
				self.iCOMOX_Data.Dongle_Comm.open(commPort=commPort)
			else:
				# self.TopPane.cb_smip.configure(values=[])
				# self.TopPane.cb_smip.set("")
				self.iCOMOX_Data.TcpIp_Comm.start(host=self.TopPane.cb_server_ip.get(), port=int(self.TopPane.entry_server_port.get()))
			# common.app.TopPane.EnableRadioButtons(enable=False)
			common.app.LiveData.EnableButtons(enable=True)

	#def EnableSensorsToSaveToExcel(self, param):
		#pass


class TopPane(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent

		# Logo image
		image_logo = Image.open(helpers.resource_path("logo.png"))
		image_logo = image_logo.resize((78, 40), Image.ANTIALIAS)
		load_logo = ImageTk.PhotoImage(image_logo)
		image_logo_label = tk.Label(self.parent, image=load_logo)
		image_logo_label.image = load_logo
		image_logo_label.grid(row=0, column=0, padx=(20, 5), pady=15)

		# iCOMOX image
		image_icomox = Image.open(helpers.resource_path("icomox.png"))
		image_icomox = image_icomox.resize((78, 40), Image.ANTIALIAS)
		load_icomox = ImageTk.PhotoImage(image_icomox)
		image_icomox_label = tk.Label(self.parent, image=load_icomox)
		image_icomox_label.image = load_icomox
		image_icomox_label.grid(row=0, column=1, padx=(20, 50))

		frame_current_iCOMOX = tk.Frame(self.parent, bd=0)
		frame_current_iCOMOX.grid(row=1, column=0, columnspan=2, sticky="w")
		if True:
			self.label_iCOMOX_name = tk.Label(frame_current_iCOMOX, font=("Verdana", 8), text="Name:", background="white")
			self.label_iCOMOX_name.grid(row=0, column=0, columnspan=2, sticky="w")

			self.label_iCOMOX_UniqueID = tk.Label(frame_current_iCOMOX, font=("Verdana", 8), text="Unique ID:", background="white")
			self.label_iCOMOX_UniqueID.grid(row=1, column=0, columnspan=2, sticky="w")

		frame_Communication = tk.Frame(self.parent, bd=0)
		frame_Communication.grid(row=0, column=2, rowspan=2, sticky="w")
		if True:
			# Combobox USB
			def on_cb_width_configure(event):
				helpers.cb_adjust_dropbox_width(combo=event.widget)

			self.load_image_refresh = ImageTk.PhotoImage(file=helpers.resource_path("icon-refresh.png"))

			tk.Label(frame_Communication, text="COM port:").grid(row=0, column=0, sticky="w")
			frame_Communication_USB = tk.Frame(frame_Communication, bd=0)
			frame_Communication_USB.grid(row=0, column=1, columnspan=2, sticky="w")
			if True:
				self.cb_usb = ttk.Combobox(frame_Communication_USB, values=[], width=10, font=("Verdana", 8), state="readonly")
				self.cb_usb.grid(row=0, column=1, padx=0, pady=2, sticky="w")
				self.cb_usb.bind("<Configure>", on_cb_width_configure)
				# COM ports list refresh button
				self.ComPortsListRefreshButton = tk.Button(frame_Communication_USB, image=self.load_image_refresh, font=("Verdana", 7), bg="white", fg="black")
				self.ComPortsListRefreshButton.grid(padx=0, row=0, column=2, sticky="w")

			# self.cb_smip = ttk.Combobox(self.parent, width=45, font=("Verdana", 8), state="readonly")
			# self.cb_smip.grid(row=1, column=3, columnspan=2, padx=0, pady=2, sticky="w")
			# self.cb_smip.bind("<<ComboboxSelected>>", on_cb_smip_select)
			# self.cb_smip.bind("<Configure>", on_cb_width_configure)

			# self.frame_tcpip = tk.Frame(self.parent)
			# self.frame_tcpip.grid(row=2, column=4, columnspan=4, sticky="w"

			ttk.Label(frame_Communication, font=("Verdana", 8), text="IP address:", background="white").grid(row=1, column=0, sticky="w")
			self.cb_server_ip = ttk.Combobox(frame_Communication, width=26, font=("Verdana", 8), state="readonly", values=common.IPs)
			self.cb_server_ip.grid(row=1, column=1, sticky="we")
			self.button_refreshIpList = tk.Button(frame_Communication, image=self.load_image_refresh, command=self.RefreshIpList, font=("Verdana, 7"), bg="white", fg="black")
			self.button_refreshIpList.grid(padx=0, row=1, column=2, sticky="w")
			self.RefreshIpList()

			ttk.Label(frame_Communication, font=("Verdana", 8), text="Port:", background="white").grid(row=2, column=0, sticky="w")#padx=10, sticky="e")
			self.entry_server_port = ttk.Entry(frame_Communication, width=8, font=("Verdana", 8), state="enabled")
			self.entry_server_port.grid(row=2, column=1, pady=(0,5), sticky="w")
			self.entry_server_port.insert(0, "1201")

		frame_Communication_Channel = tk.Frame(self.parent, bd=0)
		frame_Communication_Channel.grid(row=0, column=3, rowspan=2, sticky="w")
		if True:
			# Radio Buttons USB/SMIP
			self.rg_usb_smip_tcpip = tk.IntVar()
			self.rg_usb_smip_tcpip.set(iCOMOX_list.cCLIENT_TYPE_USB)

			self.radio_usb = tk.Radiobutton(frame_Communication_Channel, text="USB", command=self.State_USB_SMIP_TCPIP, variable=self.rg_usb_smip_tcpip, value=iCOMOX_list.cCLIENT_TYPE_USB, font=("Verdana", 8))
			self.radio_usb.grid(row=0, column=0, padx=10, sticky="sw")

			self.radio_smip = tk.Radiobutton(frame_Communication_Channel, text="SmartMesh", command=self.State_USB_SMIP_TCPIP, variable=self.rg_usb_smip_tcpip, value=iCOMOX_list.cCLIENT_TYPE_SMIP, font=("Verdana", 8), state="normal") # FIX IT BACK to "normal"
			self.radio_smip.grid(row=1, column=0, padx=10, sticky="w")

			self.radio_tcpip = tk.Radiobutton(frame_Communication_Channel, text="TCP/IP", command=self.State_USB_SMIP_TCPIP, variable=self.rg_usb_smip_tcpip, value=iCOMOX_list.cCLIENT_TYPE_TCPIP, font=("Verdana", 8))
			self.radio_tcpip.grid(row=2, column=0, padx=10, sticky="nw")

		# Connect Button
		self.buttonConnect = tk.Button(self.parent, text="Connect", font=("Verdana", 10))   #, bg="SteelBlue2", fg="white")
		self.buttonConnect.config(height=1, width=15)
		self.buttonConnect.grid(row=0, rowspan=1, column=5, padx=5, pady=1, sticky="")

		# Status LED
		self.led_status = tk_tools.Led(self.parent, size=30, border=-3)
		self.led_status.grid(row=0, rowspan=1, column=6, padx=10)
		self.led_status.to_red(on=True)

	def RefreshIpList(self):
		#self.cb_server_ip.set("")
		common.IPs = TCP_connectivity.GetIpAddresses()
		self.cb_server_ip.configure(values=common.IPs)
		self.cb_server_ip.current(0)
		# try:
		#     self.cb_server_ip.current(newindex=common.IPs.index("192.168.15.101"))
		# except:
		#     pass
		# finally:
		#     pass

	def update_iCOMOX_name(self, hello):
		if hello is None:
			name = bytearray()
			UniqueID = bytearray()
			common.app.LiveData.PlotWidebandAccFFT.hide()

		else:
			name = hello[-37:-5]
			UniqueID = hello[4:20]
			board_type = hello[1]
			if (board_type == iCOMOX_messages.cCOMOX_BOARD_TYPE_POE) and common_symbols.__ADXL1002_GRAPH_SUPPORT__:
				common.app.LiveData.PlotWidebandAccFFT.show()
			else:
				common.app.LiveData.PlotWidebandAccFFT.hide()

		self.label_iCOMOX_name.configure(text="Name: " + helpers.bytearrayToString(bytes=name))
		self.label_iCOMOX_UniqueID.configure(text="Unique ID: " + UniqueID.hex())

	def update_iComox(self):
		iCOMOX = common.iCOMOXs.current
		if iCOMOX is not None:
			UniqueID_text = self.label_iCOMOX_UniqueID["text"][11:]
			if iCOMOX.UniqueID().hex() == UniqueID_text:
				return
			# common.app.EnableButtons(configurationTab=True, eepromTab=True, clientsTab=True)
			# clear statistics, graphs & ping-pong buffers array
			common.app.clear_LiveData_Statistics_and_Information(clear_iCOMOX_only=True)
			common.app.iCOMOX_Data.PingPongArr.clear_all()

			# update information tab
			self.update_iCOMOX_name(hello=iCOMOX.Hello)
			if iCOMOX.Hello is not None:
				common.app.iCOMOX_Data.process_incoming_msg(msg=iCOMOX.Hello, iComox=iCOMOX)

			# update the graphs
			common.app.iCOMOX_Data.process_report_messages_of_current_iCOMOX()
		else:
			self.update_iCOMOX_name(hello=None)
			common.app.EnableButtons(liveDataTab=False, configurationTab=False, diagnosticTab=False, eepromTab=False, clientsTab=False)

	def set_led_status_color(self, color):
		if color == 1:
			self.led_status.to_red(on=True)
		elif color == 2:
			self.led_status.to_yellow(on=True)
		elif color == 3:
			self.led_status.to_green(on=True)
		else:
			self.led_status.to_grey(on=True)

	# disable/enable comboboxes depending on the selected mode (USB/SMIP/TcpIP)
	def State_USB_SMIP_TCPIP(self, connected=False):
		current_report_channel = self.rg_usb_smip_tcpip.get()   # common.app.get_current_report_channel()
		# if current_report_channel != iCOMOX_list.cCLIENT_TYPE_SMIP:
		#     if connected:
		#         self.cb_smip.configure(state="disabled")
		#     else:
		#         self.cb_smip.configure(state="readonly")
		# else:
		#     if connected:
		#         self.cb_smip.configure(state="readonly")
		#     else:
		#         self.cb_smip.configure(state="disabled")
		if connected:
			self.cb_usb.configure(state="disabled")
		else:
			self.cb_usb.configure(state="readonly")
		common.app.update_CommPorts()

	def EnableRadioButtons(self, enable):
		if enable:
			enable_str = "normal"
		else:
			enable_str = "disabled"
		self.radio_usb.configure(state=enable_str)
		self.radio_smip.configure(state=enable_str) # FIX IT BACK to enable_str
		self.radio_tcpip.configure(state=enable_str)


class LiveData(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent
		self.maxTemp = tk.StringVar()
		self.minTemp = tk.StringVar()
		self.maxSpeed = tk.StringVar()
		self.minSpeed = tk.StringVar()

		# Plots
		self.freq_LowPower_Acc = [1E-1]
		self.Pxxf_LowPower_Acc_X_dBg = [-90]
		self.Pxxf_LowPower_Acc_Y_dBg = [-90]
		self.Pxxf_LowPower_Acc_Z_dBg = [-90]
		self.PlotLowPowerAccFFT = PlotFFT(self.parent, title="Low power vibrations Spectrum", xtitle="Frequency [Hz]", ytitle="Amplitude [dB g/$\sqrt{Hz}}$]", xmin=0, xmax=200, ymin=-120, ymax=20, row=3, column=3)

		self.freq_Acc = [1E-1]
		self.Pxxf_Acc_X_dBg = [-90]
		self.Pxxf_Acc_Y_dBg = [-90]
		self.Pxxf_Acc_Z_dBg = [-90]
		self.PlotAccFFT = PlotFFT(self.parent, title="Vibrations Spectrum", xtitle="Frequency [Hz]", ytitle="Amplitude [dB g/$\sqrt{Hz}}$]", xmin=0, xmax=1500, ymin=-100, ymax=0, row=1, column=0)

		self.freq_Acc = [1E-1]
		self.Pxxf_Acc_dBg = [-90]
		self.freq_Mag = [1E-1]
		self.Pxxf_Mag_X_dB = [-60]
		self.Pxxf_Mag_Y_dB = [-60]
		self.Pxxf_Mag_Z_dB = [-60]
		self.PlotMagFFT = PlotFFT(self.parent, title="Magnetic Field Spectrum", xtitle="Frequency [Hz]", ytitle="Amplitude [dB μT/$\sqrt{Hz}$]", xmin=0, xmax=150, ymin=-60, ymax=60, row=3, column=0)

		self.freq_Wideband_Acc = [1E-1]
		self.Pxxf_Wideband_Acc_X_dBg = [-90]
		self.PlotWidebandAccFFT = PlotFFT(self.parent, title="Wideband Vibration Spectrum", xtitle="Frequency [Hz]", ytitle="Amplitude [dB g/$\sqrt{Hz}$]", xmin=0, xmax=11000, ymin=-120, ymax=80, row=3, column=4)    # xmax=2048*4E-5
		if not common_symbols.__ADXL1002_GRAPH_SUPPORT__:
			self.PlotWidebandAccFFT.hide()
		self.PlotMicFFT = PlotFFT(self.parent, title="Acoustic Spectrum", xtitle="Frequency [Hz]", ytitle="Amplitude [dB SPL/$\sqrt{Hz}$]", xmin=0, xmax=10300, ymin=-20, ymax=80, row=3, column=1)
		#self.PlotMicFFT = PlotFFT(self.parent, title="Acoustic Spectrum", xtitle="Time [msec]", ytitle="Amplitude [dB SPL/$\sqrt{Hz}$]", xmin=0, xmax=64, ymin=-10**(131/20)/2, ymax=10**(131/20)/2, row=3, column=1)
		#self.PlotMicFFT = PlotFFT(self.parent, title="Acoustic Spectrum", xtitle="Time [msec]", ytitle="Amplitude [dB SPL/$\sqrt{Hz}$]", xmin=0, xmax=64, ymin=-20, ymax=131, row=3, column=1)

		# Select axis to plot - Low power accelerometer
		radioFrameAcc = tk.Frame(self.parent)
		radioFrameAcc.grid(row=2, column=3, sticky="w", pady=(0, 0), padx=(32, 0))
		labelAxisAcc = tk.Label(radioFrameAcc, text="Axis: ", font=("Verdana", 8))
		labelAxisAcc.pack(side="left")
		self.axis_LowPowerAccPlot = tk.IntVar()  # radio button selection to plot
		self.axis_LowPowerAccPlot.set(iCOMOX_messages.cAXIS_X)
		radioButtonLowPowerAccX = tk.Radiobutton(radioFrameAcc, text="X", padx=1, variable=self.axis_LowPowerAccPlot, value=iCOMOX_messages.cAXIS_X, font=("Verdana", 8), command=self.update_lowpower_acc_plot)
		radioButtonLowPowerAccX.pack(side="left")
		radioButtonLowPowerAccY = tk.Radiobutton(radioFrameAcc, text="Y", padx=1, variable=self.axis_LowPowerAccPlot, value=iCOMOX_messages.cAXIS_Y, font=("Verdana", 8), command=self.update_lowpower_acc_plot)
		radioButtonLowPowerAccY.pack(side="left")
		radioButtonLowPowerAccZ = tk.Radiobutton(radioFrameAcc, text="Z", padx=1, variable=self.axis_LowPowerAccPlot, value=iCOMOX_messages.cAXIS_Z, font=("Verdana", 8), command=self.update_lowpower_acc_plot)
		radioButtonLowPowerAccZ.pack(side="left")

		# Select axis to plot - Accelerometer
		radioFrameAcc = tk.Frame(self.parent)
		radioFrameAcc.grid(row=0, sticky="w", pady=(0, 0), padx=(32, 0))
		labelAxisAcc = tk.Label(radioFrameAcc, text="Axis: ", font=("Verdana", 8))
		labelAxisAcc.pack(side="left")
		self.axis_accPlot = tk.IntVar()  # radio button selection to plot
		self.axis_accPlot.set(iCOMOX_messages.cAXIS_X)
		radioButtonAccX = tk.Radiobutton(radioFrameAcc, text="X", padx=1, variable=self.axis_accPlot, value=iCOMOX_messages.cAXIS_X, font=("Verdana", 8), command=self.update_acc_plot)
		radioButtonAccX.pack(side="left")
		radioButtonAccY = tk.Radiobutton(radioFrameAcc, text="Y", padx=1, variable=self.axis_accPlot, value=iCOMOX_messages.cAXIS_Y, font=("Verdana", 8), command=self.update_acc_plot)
		radioButtonAccY.pack(side="left")
		radioButtonAccZ = tk.Radiobutton(radioFrameAcc, text="Z", padx=1, variable=self.axis_accPlot, value=iCOMOX_messages.cAXIS_Z, font=("Verdana", 8), command=self.update_acc_plot)
		radioButtonAccZ.pack(side="left")

		# Select axis to plot - Magnetometer
		radioFrameMag = tk.Frame(self.parent)
		radioFrameMag.grid(row=2, sticky="w", pady=(0, 0), padx=(32, 0))
		labelAxisMag = tk.Label(radioFrameMag, text="Axis: ", font=("Verdana", 8))
		labelAxisMag.pack(side="left")
		self.axis_magPlot = tk.IntVar()  # radio button selection to plot
		self.axis_magPlot.set(iCOMOX_messages.cAXIS_X)
		radioButtonAccX = tk.Radiobutton(radioFrameMag, text="X", padx=1, variable=self.axis_magPlot, value=iCOMOX_messages.cAXIS_X, font=("Verdana", 8), command=self.update_mag_plot)
		radioButtonAccX.pack(side="left")
		radioButtonAccY = tk.Radiobutton(radioFrameMag, text="Y", padx=1, variable=self.axis_magPlot, value=iCOMOX_messages.cAXIS_Y, font=("Verdana", 8), command=self.update_mag_plot)
		radioButtonAccY.pack(side="left")
		radioButtonAccZ = tk.Radiobutton(radioFrameMag, text="Z", padx=1, variable=self.axis_magPlot, value=iCOMOX_messages.cAXIS_Z, font=("Verdana", 8), command=self.update_mag_plot)
		radioButtonAccZ.pack(side="left")

		self.temp_widget = Thermometer(self.parent, 380, 48, 0, 125)
		self.minTempEntry = tk.Entry(self.parent, textvariable=self.minTemp, width=5)
		self.minTempEntry.place(x=430, y=210, anchor="w")
		tk.Label(self.parent, text=" °C", font=("Verdana", 8)).place(x=460, y=200)
		self.maxTempEntry = tk.Entry(self.parent, textvariable=self.maxTemp, width=5)
		self.maxTempEntry.place(x=430, y=60, anchor="w")
		tk.Label(self.parent, text=" °C", font=("Verdana", 8)).place(x=460, y=50)
		self.minTempEntry.insert("end", "0")
		self.maxTempEntry.insert("end", "125")
		self.maxTempEntry.configure(state="readonly")
		self.minTempEntry.configure(state="readonly")

		self.speed_widget = Speedometer(self.parent, 500, 55, max_value=2000.0, min_value=0.0, size=220, bg_col='white', unit="Speed, RPM")
		self.speed_widget.draw()
		self.minSpeedEntry = tk.Entry(self.parent, textvariable=self.minSpeed, width=5)
		self.minSpeedEntry.place(x=550, y=250, anchor="w")
		self.maxSpeedEntry = tk.Entry(self.parent, textvariable=self.maxSpeed, width=5)
		self.maxSpeedEntry.place(x=640, y=250, anchor="w")
		self.minSpeedEntry.insert("end", "0")
		self.maxSpeedEntry.insert("end", "2000")
		self.maxSpeedEntry.configure(state="readonly")
		self.minSpeedEntry.configure(state="readonly")

		# ML1 support
		self.Frame_ML1 = tk.Frame(self.parent)
		self.Frame_ML1.grid(row=0, rowspan=2, column=3, columnspan=1, sticky="we", pady=(10, 0), padx=(64, 0))
		tk.Label(self.Frame_ML1, text="Anomaly detection", font=("Verdana", 8)).grid(row=0, column=0, columnspan=2, sticky="we")

		labelAnomaly = tk.Label(self.Frame_ML1, text="Anomaly:", font=("Verdana", 8))
		labelAnomaly.grid(row=1, column=0, sticky="w")
		self.progressbar_AnomalyDetection_Anomaly = ttk.Progressbar(self.Frame_ML1, style="Anomaly.Green.Horizontal.TProgressbar", orient="horizontal", mode="determinate", value=0, maximum=100)
		self.progressbar_AnomalyDetection_Anomaly.grid(row=1, column=1, padx=5, pady=10, sticky="we")

		self.progressbar_AnomalyDetection_states = list(range(4))
		self.button_AnomalyDetection_train = list(range(4))
		self.AnomalyDetection_which_model_is_currently_trained = None
		self.AnomalyDetection_train_count = 0
		for i in range(0, len(self.button_AnomalyDetection_train)):
			self.button_AnomalyDetection_train[i] = tk.Button(self.Frame_ML1, text="Preset {}".format(i + 1), font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			self.button_AnomalyDetection_train[i].grid(row=2 + i, column=0, sticky="w")
			self.progressbar_AnomalyDetection_states[i] = ttk.Progressbar(self.Frame_ML1, style="Infer.Horizontal.TProgressbar", orient="horizontal", mode="determinate", value=0, maximum=100)
			self.progressbar_AnomalyDetection_states[i].grid(row=2 + i, column=1, padx=5, pady=10, sticky="we")

		self.button_AnomalyDetection_train[0].configure(command=lambda: self.OnClickTrainStateButton(0))
		self.button_AnomalyDetection_train[1].configure(command=lambda: self.OnClickTrainStateButton(1))
		self.button_AnomalyDetection_train[2].configure(command=lambda: self.OnClickTrainStateButton(2))
		self.button_AnomalyDetection_train[3].configure(command=lambda: self.OnClickTrainStateButton(3))

		self.button_ML1_Initialize = tk.Button(self.Frame_ML1, command=self.OnClickResetButton, text="Clear Presets", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
		self.button_ML1_Initialize.grid(row=6, column=0, columnspan=4, sticky="we")

	def AnomalyDetection_NullifyProgressbars(self):
		for i in range(0, len(self.button_AnomalyDetection_train)):
			self.update_progressbar_AnomalyDetection_pdf(index=i, value=0)
		self.update_progressbar_AnomalyDetection_Anomaly(value=0)

	def OnClickResetButton(self):
		self.AnomalyDetection_which_model_is_currently_trained = None
		self.AnomalyDetection_train_count = 0
		self.AnomalyDetection_NullifyProgressbars()
		self.update_progressbar_AnomalyDetection_Anomaly(value=0)
		common.app.iCOMOX_Data.send_msg(msg=iCOMOX_messages.OUT_MSG_SetConfiguration(
			ConfigBitmask=0,
			ConfigModulesBitmask=iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection,
			ActiveModules=iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection,
			LocalTimestamp=0,
			Common=0,
			RawData_Sensors=0, Repetition=0, IntervalInMinutes=0,
			AnomalyDetection_Command=iCOMOX_messages.cANOMALY_DETECTION_COMMAND_Reset, AnomalyDetection_Sensors=iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL356, AnomalyDetection_StateToTrain=0
		))

	def OnClickTrainStateButton(self, stateToTrain):
		self.AnomalyDetection_which_model_is_currently_trained = stateToTrain
		self.AnomalyDetection_train_count = 0
		self.AnomalyDetection_NullifyProgressbars()
		self.update_progressbar_AnomalyDetection_Anomaly(value=0)
		common.app.iCOMOX_Data.send_msg(msg=iCOMOX_messages.OUT_MSG_SetConfiguration(
			ConfigBitmask=0, ConfigModulesBitmask=iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection, ActiveModules=iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection,
			LocalTimestamp=0,
			Common=0,
			RawData_Sensors=0, Repetition=0xFF, IntervalInMinutes=0xFFFF,
			AnomalyDetection_Command=iCOMOX_messages.cANOMALY_DETECTION_COMMAND_Train, AnomalyDetection_Sensors=0, AnomalyDetection_StateToTrain=stateToTrain
		))
		common.app.StatusBar.set("Sent train request to anomaly detection module")

	def update_progressbar_AnomalyDetection_pdf(self, index, value):
		if (self.AnomalyDetection_which_model_is_currently_trained is not None) and (index == self.AnomalyDetection_which_model_is_currently_trained):
			style = "Train.Horizontal.TProgressbar"
			value = self.AnomalyDetection_train_count * 20
		else:
			style = "Infer.Horizontal.TProgressbar"
		self.progressbar_AnomalyDetection_states[index].configure(value=value, style=style)

	def update_progressbar_AnomalyDetection_Anomaly(self, value):
		if value > 100:
			value = 100
		elif value < 0:
			value = 0
		if value <= 50:
			style = "Anomaly.Green.Horizontal.TProgressbar"
		elif value <= 75:
			style = "Anomaly.Yellow.Horizontal.TProgressbar"
		else:
			style = "Anomaly.Red.Horizontal.TProgressbar"
		self.progressbar_AnomalyDetection_Anomaly.configure(value=value, style=style)

	def AnomalyDetection_Report_Handler(self, ValAnomaly, probState, ReportStatus, Sensors, Result):
		# model_learned = (ReportStatus & 0x01) != 0
		if Result != 0:
			# self.AnomalyDetection_train_count = 0
			# self.AnomalyDetection_which_model_is_currently_trained = None
			# for i in range(0, len(probState)):
			#     self.update_progressbar_AnomalyDetection_pdf(index=i, value=0)
			common.app.StatusBar.set(text="Get anomaly detection module report error {}".format(Result))
			return
		IsTrainReport = (ReportStatus & 0x80) != 0   # Train ("1") or Inference ("0")  result
		if IsTrainReport:
			if self.AnomalyDetection_train_count < 5:
				self.AnomalyDetection_train_count += 1
			self.AnomalyDetection_NullifyProgressbars()
			self.update_progressbar_AnomalyDetection_Anomaly(value=0)
			common.app.StatusBar.set("Get anomaly detection module train report")
		else:
			if self.AnomalyDetection_train_count < 5:
				return
			# self.AnomalyDetection_train_count = 0
			self.AnomalyDetection_which_model_is_currently_trained = None
			probStateSum = sum(probState)
			if (probStateSum == 0) or (len([*filter(lambda state : state < 0, probState)]) > 0):
				# prevent division by zero exception in case normalization is impossible and/or negative probState[] values
				for i in range(0, len(probState)):
					self.update_progressbar_AnomalyDetection_pdf(index=i, value=0)
				helpers.OUT("LiveData.AnomalyDetection_Report_Handler(): Illegal states distribution received")
			else:
				for i in range(0, len(probState)):
					self.update_progressbar_AnomalyDetection_pdf(index=i, value=int(100 * probState[i] / probStateSum))
			self.update_progressbar_AnomalyDetection_Anomaly(value=ValAnomaly)
			common.app.StatusBar.set(text="Get anomaly detection module inference report")

	def update_wideband_acc_plot_data(self, freq_Acc, Pxxf_Acc_dBg):
		self.freq_Wideband_Acc = freq_Acc
		self.Pxxf_Wideband_Acc_X_dBg = Pxxf_Acc_dBg

	def update_wideband_acc_plot(self):
		if len(self.Pxxf_Wideband_Acc_X_dBg) == 1:
			self.PlotWidebandAccFFT.updatePlot([1E-1], self.Pxxf_Wideband_Acc_X_dBg)
		else:
			self.PlotWidebandAccFFT.updatePlot(self.freq_Wideband_Acc, self.Pxxf_Wideband_Acc_X_dBg)

	def update_lowpower_acc_plot_data(self, freq_Acc, Pxxf_Acc_X_dBg, Pxxf_Acc_Y_dBg, Pxxf_Acc_Z_dBg):
		self.freq_LowPower_Acc = freq_Acc
		self.Pxxf_LowPower_Acc_X_dBg = Pxxf_Acc_X_dBg
		self.Pxxf_LowPower_Acc_Y_dBg = Pxxf_Acc_Y_dBg
		self.Pxxf_LowPower_Acc_Z_dBg = Pxxf_Acc_Z_dBg

	def update_lowpower_acc_plot(self):
		if self.axis_LowPowerAccPlot.get() == iCOMOX_messages.cAXIS_X:
			values = self.Pxxf_LowPower_Acc_X_dBg
		elif self.axis_LowPowerAccPlot.get() == iCOMOX_messages.cAXIS_Y:
			values = self.Pxxf_LowPower_Acc_Y_dBg
		else:
			values = self.Pxxf_LowPower_Acc_Z_dBg

		if len(values) == 1:
			self.PlotLowPowerAccFFT.updatePlot([1E-1], values)
		else:
			self.PlotLowPowerAccFFT.updatePlot(self.freq_LowPower_Acc, values)

	def update_acc_plot_data(self, axis, freq_Acc, Pxxf_Acc_dBg):
		self.freq_Acc = freq_Acc
		if axis == iCOMOX_messages.cAXIS_X:
			self.Pxxf_Acc_X_dBg = Pxxf_Acc_dBg
		elif axis == iCOMOX_messages.cAXIS_Y:
			self.Pxxf_Acc_Y_dBg = Pxxf_Acc_dBg
		else:
			self.Pxxf_Acc_Z_dBg = Pxxf_Acc_dBg

	def update_acc_plot(self):
		if self.axis_accPlot.get() == iCOMOX_messages.cAXIS_X:
			values = self.Pxxf_Acc_X_dBg
		elif self.axis_accPlot.get() == iCOMOX_messages.cAXIS_Y:
			values = self.Pxxf_Acc_Y_dBg
		else:
			values = self.Pxxf_Acc_Z_dBg

		if len(values) == 1:
			self.PlotAccFFT.updatePlot([1E-1], values)
		else:
			self.PlotAccFFT.updatePlot(self.freq_Acc, values)

	def update_mag_plot_data(self, freq_Mag, Pxxf_Mag_X_dB, Pxxf_Mag_Y_dB, Pxxf_Mag_Z_dB, Pxxf_X_noise_floor=None, Pxxf_Y_noise_floor=None, Pxxf_Z_noise_floor=None):
		self.freq_Mag = freq_Mag
		self.Pxxf_Mag_X_dB = Pxxf_Mag_X_dB
		self.Pxxf_Mag_Y_dB = Pxxf_Mag_Y_dB
		self.Pxxf_Mag_Z_dB = Pxxf_Mag_Z_dB
		self.Pxxf_X_noise_floor = Pxxf_X_noise_floor
		self.Pxxf_Y_noise_floor = Pxxf_Y_noise_floor
		self.Pxxf_Z_noise_floor = Pxxf_Z_noise_floor

	def update_mag_plot(self):
		if self.axis_magPlot.get() == iCOMOX_messages.cAXIS_X:
			self.PlotMagFFT.updatePlot(self.freq_Mag, self.Pxxf_Mag_X_dB, self.Pxxf_X_noise_floor)
		elif self.axis_magPlot.get() == iCOMOX_messages.cAXIS_Y:
			self.PlotMagFFT.updatePlot(self.freq_Mag, self.Pxxf_Mag_Y_dB, self.Pxxf_Y_noise_floor)
		else:
			self.PlotMagFFT.updatePlot(self.freq_Mag, self.Pxxf_Mag_Z_dB, self.Pxxf_Z_noise_floor)

	def clear_all(self):
		self.update_lowpower_acc_plot_data(freq_Acc=[1E-1], Pxxf_Acc_X_dBg=[-100], Pxxf_Acc_Y_dBg=[-100], Pxxf_Acc_Z_dBg=[-100])
		self.update_acc_plot_data(axis=iCOMOX_messages.cAXIS_X, freq_Acc=[1E-1], Pxxf_Acc_dBg=[-100])
		self.update_acc_plot_data(axis=iCOMOX_messages.cAXIS_Y, freq_Acc=[1E-1], Pxxf_Acc_dBg=[-100])
		self.update_acc_plot_data(axis=iCOMOX_messages.cAXIS_Z, freq_Acc=[1E-1], Pxxf_Acc_dBg=[-100])
		self.update_mag_plot_data([1E-1], [-60], [-60], [-60])
		self.update_wideband_acc_plot_data(freq_Acc=[1E-1], Pxxf_Acc_dBg=[-100])
		self.PlotLowPowerAccFFT.updatePlot(x=[], y=[])
		self.PlotAccFFT.updatePlot(x=[], y=[])
		self.PlotWidebandAccFFT.updatePlot(x=[], y=[])
		self.PlotMagFFT.updatePlot(x=[], y=[])
		self.PlotMicFFT.updatePlot(x=[], y=[])
		self.speed_widget.set_value(number=0)
		self.temp_widget.display(temp=None)
		self.AnomalyDetection_NullifyProgressbars()

	def EnableButtons(self, enable):
		state = "normal" if enable else "disabled"
		self.button_ML1_Initialize.configure(state=state)
		for button in self.button_AnomalyDetection_train:
			button.configure(state=state)

class Statistics(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent
		label = tk.Label(self.parent, text="Mean", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=0, column=1, ipadx=50, ipady=10)
		label = tk.Label(self.parent, text="SD", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=0, column=2, ipadx=50, ipady=10)
		label = tk.Label(self.parent, text="Min", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=0, column=3, ipadx=50, ipady=10)
		label = tk.Label(self.parent, text="Max", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=0, column=4, ipadx=50, ipady=10)
		ttk.Separator(self.parent, orient="horizontal").grid(row=1, column=1, columnspan=4, sticky="we")

		label = tk.Label(self.parent, text="Low power accelerometer X [g]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=1, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Low power accelerometer Y [g]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=2, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Low power accelerometer Z [g]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=3, column=0, ipadx=50, ipady=10, sticky="we")

		label = tk.Label(self.parent, text="Accelerometer X [g]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=4, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Accelerometer Y [g]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=5, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Accelerometer Z [g]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=6, column=0, ipadx=50, ipady=10, sticky="we")

		label = tk.Label(self.parent, text="Magnetometer X [μT]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=7, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Magnetometer Y [μT]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=8, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Magnetometer Z [μT]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=9, column=0, ipadx=50, ipady=10, sticky="we")

		label = tk.Label(self.parent, text="Microphone [dB SPL]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE, width=15)
		label.grid(row=10, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Temperature [°C]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE, width=15)
		label.grid(row=11, column=0, ipadx=50, ipady=10, sticky="we")
		label = tk.Label(self.parent, text="Wideband accelerometer X [g]: ", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		label.grid(row=12, column=0, ipadx=50, ipady=10, sticky="we")

		self.lowPowerAccMeanX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMeanX.grid(row=1, column=1, sticky="nwse")
		self.lowPowerAccMeanY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMeanY.grid(row=2, column=1, sticky="nwse")
		self.lowPowerAccMeanZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMeanZ.grid(row=3, column=1, sticky="nwse")

		self.lowPowerAccMeanX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMeanX.grid(row=1, column=1, sticky="nwse")
		self.lowPowerAccMeanY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMeanY.grid(row=2, column=1, sticky="nwse")
		self.lowPowerAccMeanZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMeanZ.grid(row=3, column=1, sticky="nwse")
		self.lowPowerAccSDX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccSDX.grid(row=1, column=2, sticky="nwse")
		self.lowPowerAccSDY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccSDY.grid(row=2, column=2, sticky="nwse")
		self.lowPowerAccSDZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccSDZ.grid(row=3, column=2, sticky="nwse")
		self.lowPowerAccMinX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMinX.grid(row=1, column=3, sticky="nwse")
		self.lowPowerAccMinY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMinY.grid(row=2, column=3, sticky="nwse")
		self.lowPowerAccMinZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMinZ.grid(row=3, column=3, sticky="nwse")
		self.lowPowerAccMaxX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMaxX.grid(row=1, column=4, sticky="nwse")
		self.lowPowerAccMaxY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMaxY.grid(row=2, column=4, sticky="nwse")
		self.lowPowerAccMaxZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.lowPowerAccMaxZ.grid(row=3, column=4, sticky="nwse")

		self.accMeanX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMeanX.grid(row=4, column=1, sticky="nwse")
		self.accMeanY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMeanY.grid(row=5, column=1, sticky="nwse")
		self.accMeanZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMeanZ.grid(row=6, column=1, sticky="nwse")
		self.accSDX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accSDX.grid(row=4, column=2, sticky="nwse")
		self.accSDY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accSDY.grid(row=5, column=2, sticky="nwse")
		self.accSDZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accSDZ.grid(row=6, column=2, sticky="nwse")
		self.accMinX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMinX.grid(row=4, column=3, sticky="nwse")
		self.accMinY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMinY.grid(row=5, column=3, sticky="nwse")
		self.accMinZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMinZ.grid(row=6, column=3, sticky="nwse")
		self.accMaxX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMaxX.grid(row=4, column=4, sticky="nwse")
		self.accMaxY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMaxY.grid(row=5, column=4, sticky="nwse")
		self.accMaxZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.accMaxZ.grid(row=6, column=4, sticky="nwse")

		self.magMeanX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMeanX.grid(row=7, column=1, sticky="nwse")
		self.magMeanY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMeanY.grid(row=8, column=1, sticky="nwse")
		self.magMeanZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMeanZ.grid(row=9, column=1, sticky="nwse")
		self.magSDX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magSDX.grid(row=7, column=2, sticky="nwse")
		self.magSDY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magSDY.grid(row=8, column=2, sticky="nwse")
		self.magSDZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magSDZ.grid(row=9, column=2, sticky="nwse")
		self.magMinX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMinX.grid(row=7, column=3, sticky="nwse")
		self.magMinY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMinY.grid(row=8, column=3, sticky="nwse")
		self.magMinZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMinZ.grid(row=9, column=3, sticky="nwse")
		self.magMaxX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMaxX.grid(row=7, column=4, sticky="nwse")
		self.magMaxY = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMaxY.grid(row=8, column=4, sticky="nwse")
		self.magMaxZ = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.magMaxZ.grid(row=9, column=4, sticky="nwse")

		self.micMean = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.micMean.grid(row=10, column=1, sticky="nwse")
		self.micSD = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.micSD.grid(row=10, column=2, sticky="nwse")
		self.micMin = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.micMin.grid(row=10, column=3, sticky="nwse")
		self.micMax = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.micMax.grid(row=10, column=4, sticky="nwse")

		self.tempMean = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.tempMean.grid(row=11, column=1, sticky="nwse")
		self.tempSD = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.tempSD.grid(row=11, column=2, sticky="nwse")
		self.tempMin = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.tempMin.grid(row=11, column=3, sticky="nwse")
		self.tempMax = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.tempMax.grid(row=11, column=4, sticky="nwse")

		self.widebandAccMeanX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.widebandAccMeanX.grid(row=12, column=1, sticky="nwse")
		self.widebandAccSDX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.widebandAccSDX.grid(row=12, column=2, sticky="nwse")
		self.widebandAccMinX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.widebandAccMinX.grid(row=12, column=3, sticky="nwse")
		self.widebandAccMaxX = tk.Label(self.parent, text="", font=("Verdana", 10), bd=1, relief=tk.RIDGE)
		self.widebandAccMaxX.grid(row=12, column=4, sticky="nwse")

	def update_lowpower_acc(self):
		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_lowpower_acc_x)
		if count > 0:
			self.lowPowerAccMeanX.config(text=str(round(mean, 2)))
			self.lowPowerAccSDX.config(text=str(round(std, 2)))
			self.lowPowerAccMinX.config(text=str(round(minimum, 2)))
			self.lowPowerAccMaxX.config(text=str(round(maximum, 2)))
		else:
			self.lowPowerAccMeanX.config(text="")
			self.lowPowerAccSDX.config(text="")
			self.lowPowerAccMinX.config(text="")
			self.lowPowerAccMaxX.config(text="")

		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_lowpower_acc_y)
		if count > 0:
			self.lowPowerAccMeanY.config(text=str(round(mean, 2)))
			self.lowPowerAccSDY.config(text=str(round(std, 2)))
			self.lowPowerAccMinY.config(text=str(round(minimum, 2)))
			self.lowPowerAccMaxY.config(text=str(round(maximum, 2)))
		else:
			self.lowPowerAccMeanY.config(text="")
			self.lowPowerAccSDY.config(text="")
			self.lowPowerAccMinY.config(text="")
			self.lowPowerAccMaxY.config(text="")

		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_lowpower_acc_z)
		if count > 0:
			self.lowPowerAccMeanZ.config(text=str(round(mean, 2)))
			self.lowPowerAccSDZ.config(text=str(round(std, 2)))
			self.lowPowerAccMinZ.config(text=str(round(minimum, 2)))
			self.lowPowerAccMaxZ.config(text=str(round(maximum, 2)))
		else:
			self.lowPowerAccMeanZ.config(text="")
			self.lowPowerAccSDZ.config(text="")
			self.lowPowerAccMinZ.config(text="")
			self.lowPowerAccMaxZ.config(text="")

	def update_acc(self):
		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_acc_x)
		if count > 0:
			self.accMeanX.config(text=str(round(mean, 2)))
			self.accSDX.config(text=str(round(std, 2)))
			self.accMinX.config(text=str(round(minimum, 2)))
			self.accMaxX.config(text=str(round(maximum, 2)))
		else:
			self.accMeanX.config(text="")
			self.accSDX.config(text="")
			self.accMinX.config(text="")
			self.accMaxX.config(text="")

		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_acc_y)
		if count > 0:
			self.accMeanY.config(text=str(round(mean, 2)))
			self.accSDY.config(text=str(round(std, 2)))
			self.accMinY.config(text=str(round(minimum, 2)))
			self.accMaxY.config(text=str(round(maximum, 2)))
		else:
			self.accMeanY.config(text="")
			self.accSDY.config(text="")
			self.accMinY.config(text="")
			self.accMaxY.config(text="")

		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_acc_z)
		if count > 0:
			self.accMeanZ.config(text=str(round(mean, 2)))
			self.accSDZ.config(text=str(round(std, 2)))
			self.accMinZ.config(text=str(round(minimum, 2)))
			self.accMaxZ.config(text=str(round(maximum, 2)))
		else:
			self.accMeanZ.config(text="")
			self.accSDZ.config(text="")
			self.accMinZ.config(text="")
			self.accMaxZ.config(text="")

	def update_mag(self):
		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_mag_x)
		if count > 0:
			self.magMeanX.config(text=str(round(mean, 2)))
			self.magSDX.config(text=str(round(std, 2)))
			self.magMinX.config(text=str(round(minimum, 2)))
			self.magMaxX.config(text=str(round(maximum, 2)))
		else:
			self.magMeanX.config(text="")
			self.magSDX.config(text="")
			self.magMinX.config(text="")
			self.magMaxX.config(text="")

		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_mag_y)
		if count > 0:
			self.magMeanY.config(text=str(round(mean, 2)))
			self.magSDY.config(text=str(round(std, 2)))
			self.magMinY.config(text=str(round(minimum, 2)))
			self.magMaxY.config(text=str(round(maximum, 2)))
		else:
			self.magMeanY.config(text="")
			self.magSDY.config(text="")
			self.magMinY.config(text="")
			self.magMaxY.config(text="")

		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_mag_z)
		if count > 0:
			self.magMeanZ.config(text=str(round(mean, 2)))
			self.magSDZ.config(text=str(round(std, 2)))
			self.magMinZ.config(text=str(round(minimum, 2)))
			self.magMaxZ.config(text=str(round(maximum, 2)))
		else:
			self.magMeanZ.config(text="")
			self.magSDZ.config(text="")
			self.magMinZ.config(text="")
			self.magMaxZ.config(text="")

	def update_mic(self):
		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_mic)
		if count > 0:
			# self.micMean.config(text=str(round(mean, 2)))
			# self.micSD.config(text=str(round(std, 2)))
			# self.micMin.config(text=str(round(minimum, 2)))
			# self.micMax.config(text=str(round(maximum, 2)))
			epsilon = 0.1
			self.micMean.config(text=str(round(20*np.log10(abs(mean)+epsilon), 2)))
			self.micSD.config(text=str(round(20*np.log10(abs(std)+epsilon), 2)))
			self.micMin.config(text=str(round(20*np.log10(abs(minimum)+epsilon), 2)))
			self.micMax.config(text=str(round(20*np.log10(abs(maximum)+epsilon), 2)))
		else:
			self.micMean.config(text="")
			self.micSD.config(text="")
			self.micMin.config(text="")
			self.micMax.config(text="")

	def update_temp(self):
		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_temp)
		if count > 0:
			self.tempMean.config(text=str(round(mean, 2)))
			self.tempSD.config(text=str(round(std, 2)))
			self.tempMin.config(text=str(round(minimum, 2)))
			self.tempMax.config(text=str(round(maximum, 2)))
		else:
			self.tempMean.config(text="")
			self.tempSD.config(text="")
			self.tempMin.config(text="")
			self.tempMax.config(text="")

	def update_wideband_acc(self):
		[mean, std, minimum, maximum, count] = common.app.statisticsLogger.retrieve_statistics(data_arr_index=statistics_data_logger.stat_data_logger_index_wideband_acc_x)
		if count > 0:
			self.widebandAccMeanX.config(text=str(round(mean, 2)))
			self.widebandAccSDX.config(text=str(round(std, 2)))
			self.widebandAccMinX.config(text=str(round(minimum, 2)))
			self.widebandAccMaxX.config(text=str(round(maximum, 2)))
		else:
			self.widebandAccMeanX.config(text="")
			self.widebandAccSDX.config(text="")
			self.widebandAccMinX.config(text="")
			self.widebandAccMaxX.config(text="")

	def update_all(self):
		self.update_lowpower_acc()
		self.update_acc()
		self.update_mag()
		self.update_mic()
		self.update_temp()
		self.update_wideband_acc()


def timestamp_ns_to_str(timestamp_ns):
	sec_since_epoch, ns_since_last_sec = divmod(timestamp_ns, 1000000000)
	time_struct = time.gmtime(sec_since_epoch)
	timestamp = time.mktime(time_struct)
	return " " + datetime.datetime.fromtimestamp(timestamp=timestamp).strftime("%y-%m-%d %H:%M:%S") + ".{0:06d}".format(int(ns_since_last_sec))


class Configuration(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent

		self.lowPowerAccFile = tk.IntVar()
		self.accFile = tk.IntVar()
		self.magFile = tk.IntVar()
		self.micFile = tk.IntVar()
		self.tempFile = tk.IntVar()
		self.widebandAccFile = tk.IntVar()

		self.lowPowerAccAcq = tk.BooleanVar(False)
		self.accAcq = tk.BooleanVar(False)
		self.magAcq = tk.BooleanVar(False)
		self.micAcq = tk.BooleanVar(False)
		self.tempAcq = tk.BooleanVar(False)
		self.widebandAccAcq = tk.BooleanVar(False)

		self.config_AbsoluteTime = tk.BooleanVar(False)
		self.config_Common = tk.BooleanVar(False)
		self.config_PeriodicActivation = tk.BooleanVar(False)
		self.ConfigModule_RawData = tk.BooleanVar(False)
		self.ConfigModule_AnomalyDetection = tk.BooleanVar(False)
		self.Module_AnomalyDetection_sensors = tk.IntVar(0)

		self.radioCommChannel = tk.IntVar(0)
		self.enableVibrator = tk.BooleanVar(False)
		self.saveToSD = tk.BooleanVar(False)
		self.transmit = tk.BooleanVar(False)
		# self.ML1_stateToTrain = tk.IntVar(0)
		# self.ML1_Command = tk.IntVar(iCOMOX_messages.cANOMALY_DETECTION_COMMAND_Train)

		self.ActivateModule_RawData = tk.BooleanVar(False)
		self.ActivateModule_AnomalyDetection = tk.BooleanVar(False)

		self.PeriodicActivation_IntervalInMinutes = tk.IntVar()
		self.PeriodicActivation_Repetitions = tk.IntVar()

		# self.stream_schedule = tk.IntVar()
		# self.schedule = tk.StringVar()
		# self.time = tk.StringVar()
		self.reset_type = tk.IntVar(0)
		self.save_to_file = tk.IntVar()
		self.dirname = None
		self.start_pressed = False
		self.stop_pressed = False
		self.workbook = None
		self.WS_ADXL362 = None
		self.WS_ADXL356 = None
		self.WS_BMM150 = None
		self.WS_ADT7410 = None
		self.WS_IM69D130 = None
		self.WS_ADXL1002 = None
		self.logfilename = None
		self.time_to_save = None
		self.WS_ADXL356_Batches = [0, 0, 0]

		# # Label frames definition
		acqFrame = tk.LabelFrame(self.parent, text="Data acquisition", width=330, height=500)
		acqFrame.grid(row=0, column=0, rowspan=4, padx=10, pady=(20, 10), sticky="wn")
		if True:
			self.cb_ConfigAbsoluteTime = tk.Checkbutton(acqFrame, text="Configure absolute time", variable=self.config_AbsoluteTime, font=("Verdana", 8), state="normal")
			self.cb_ConfigAbsoluteTime.grid(row=0, column=0, padx=0, sticky=tk.W)
			self.config_AbsoluteTime.set(True)
			self.cb_ConfigAbsoluteTime.grid_forget()

			# Common options
			self.cb_ConfigCommon = tk.Checkbutton(acqFrame, text="Configure common options:", variable=self.config_Common, font=("Verdana", 8), state="normal")
			self.cb_ConfigCommon.grid(row=1, column=0, padx=0, sticky=tk.W)
			self.config_Common.set(True)
			self.cb_ConfigCommon.grid_forget()

			# tk.Label(acqFrame, text="Output channel:", font=("Verdana", 8)).place(x=35, y=5)
			frameOutputChannel = tk.Frame(acqFrame)
			frameOutputChannel.grid(row=2, column=0, padx=0, sticky=tk.W)
			if True:
				self.cb_transmit = tk.Checkbutton(frameOutputChannel, text="Output channel:", variable=self.transmit, font=("Verdana", 8), state="normal")
				self.cb_transmit.grid(row=0, column=0, padx=0, sticky=tk.W)
				self.transmit.set(True)
				frameOutputChannelRadioGroup = tk.Frame(frameOutputChannel)
				frameOutputChannelRadioGroup.grid(row=0, column=1, padx=0, sticky=tk.E)
				if True:
					self.radioCommChannel_USB = tk.Radiobutton(frameOutputChannelRadioGroup, text="USB", command=None, variable=self.radioCommChannel, value=0, font=("Verdana", 8))
					self.radioCommChannel_USB.grid(row=0, column=0, padx=(20,0), sticky=tk.E)
					self.radioCommChannel_AUX = tk.Radiobutton(frameOutputChannelRadioGroup, text="TCP/IP / SMIP", command=None, variable=self.radioCommChannel, value=1, font=("Verdana", 8))
					self.radioCommChannel_AUX.grid(row=0, column=1, padx=(0,20), sticky=tk.W)

			self.cb_enableVibrator = tk.Checkbutton(acqFrame, text="Enable vibration motor", variable=self.enableVibrator, font=("Verdana", 8), state="normal")
			self.cb_enableVibrator.grid(row=3, column=0, padx=0, sticky=tk.W)
			self.cb_enableVibrator.grid_forget()

			self.cb_saveToSD = tk.Checkbutton(acqFrame, text="Save to file in SD card", variable=self.saveToSD, font=("Verdana", 8), state="normal")
			self.cb_saveToSD.grid(row=4, column=0, padx=0, sticky=tk.W)

			self.cb_config_Module_RawData = tk.Checkbutton(acqFrame, text="Configure raw data module", variable=self.ConfigModule_RawData, font=("Verdana", 8), state="normal")
			self.cb_config_Module_RawData.grid(row=5, column=0, padx=0, sticky=tk.W)
			self.ConfigModule_RawData.set(True)
			self.cb_config_Module_RawData.grid_forget()

			# self.cb_config_Module_AnomalyDetection = tk.Checkbutton(acqFrame, text="Configure anomaly detection module", variable=self.ConfigModule_AnomalyDetection, font=("Verdana", 8), state="normal")
			# self.cb_config_Module_AnomalyDetection.grid(row=6, column=0, padx=0, sticky=tk.W)
			# self.ConfigModule_AnomalyDetection.set(False)
			# self.cb_config_Module_AnomalyDetection.grid_forget()

			# tk.Label(acqFrame, text="Modules to activate:", font=("Verdana", 8)).place(x=5, y=110)
			self.cb_enable_Module_RawData = tk.Checkbutton(acqFrame, text="Raw sensors data to acquire:", variable=self.ActivateModule_RawData, font=("Verdana", 8), state="normal")
			self.cb_enable_Module_RawData.grid(row=7, column=0, padx=0, sticky=tk.W)
			self.ActivateModule_RawData.set(True)
			# Raw data module
			# self.cb_ConfigAbsoluteTime = tk.Checkbutton(acqFrame, text="Raw sensors data to acquire:", variable=self.enable_RawData, font=("Verdana", 8), state="normal")
			# self.cb_ConfigAbsoluteTime.place(x=5, y=210, anchor="w")
			frameSensors = tk.Frame(acqFrame)
			frameSensors.grid(row=8, column=0, padx=10, sticky=tk.W)
			if True:
				self.lowPowerAccCheckAcq = tk.Checkbutton(frameSensors, text="Low power accelerometer", variable=self.lowPowerAccAcq, font=("Verdana", 8), state="normal")
				self.lowPowerAccCheckAcq.grid(row=0, column=0, padx=0, sticky=tk.W)
				self.lowPowerAccAcq.set(1)

				self.accCheckAcq = tk.Checkbutton(frameSensors, text="Accelerometer", variable=self.accAcq, font=("Verdana", 8), state="normal")
				self.accCheckAcq.grid(row=1, column=0, padx=0, sticky=tk.W)
				self.accAcq.set(1)

				self.tempCheckAcq = tk.Checkbutton(frameSensors, text="Thermometer", variable=self.tempAcq, font=("Verdana", 8), state="normal")
				self.tempCheckAcq.grid(row=2, column=0, padx=0, sticky=tk.W)
				self.tempAcq.set(1)

				self.magCheckAcq = tk.Checkbutton(frameSensors, text="Magnetometer", variable=self.magAcq, font=("Verdana", 8), state="normal")
				self.magCheckAcq.grid(row=3, column=0, padx=0, sticky=tk.W)
				self.magAcq.set(1)

				self.micCheckAcq = tk.Checkbutton(frameSensors, text="Microphone", variable=self.micAcq, font=("Verdana", 8), state="normal")
				self.micCheckAcq.grid(row=4, column=0, padx=0, sticky=tk.W)
				self.micAcq.set(1)

				self.widebandAccCheckAcq = tk.Checkbutton(frameSensors, text="Wideband accelerometer (optional)", variable=self.widebandAccAcq, font=("Verdana", 8), state="normal")
				self.widebandAccCheckAcq.grid(row=5, column=0, padx=0, sticky=tk.W)
				# self.widebandAccAcq.set(1)
				self.widebandAccCheckAcq.grid_forget()

			# Anomaly detection module
			self.cb_configure_Module_AnomalyDetection = tk.Checkbutton(acqFrame, text="Configure anomaly detection algorithm", variable=self.ConfigModule_AnomalyDetection, font=("Verdana", 8), state="normal")
			self.cb_configure_Module_AnomalyDetection.grid(row=9, column=0, padx=0, sticky=tk.W)
			self.cb_enable_Module_AnomalyDetection = tk.Checkbutton(acqFrame, text="Enable anomaly detection algorithm", variable=self.ActivateModule_AnomalyDetection, font=("Verdana", 8), state="normal")
			self.cb_enable_Module_AnomalyDetection.grid(row=10, column=0, padx=10, sticky=tk.W)

			# Periodic activation
			self.cb_configure_PeriodicActivation = tk.Checkbutton(acqFrame, text="Configure periodic activation", variable=self.config_PeriodicActivation, font=("Verdana", 8), state="normal")
			self.cb_configure_PeriodicActivation.grid(row=11, column=0, padx=0, sticky=tk.W)
			frameActivationPeriod = tk.Frame(acqFrame)
			frameActivationPeriod.grid(row=12, column=0, padx=10, sticky=tk.W)
			if True:
				label = tk.Label(frameActivationPeriod, text="Activation period [minutes]:", font=("Verdana", 8))
				label.grid(row=0, column=0, padx=(0,5), sticky=tk.W)
				self.entry_PeriodicActivation_IntervalInMinutes = tk.Entry(frameActivationPeriod, textvariable=self.PeriodicActivation_IntervalInMinutes, width=5)
				self.entry_PeriodicActivation_IntervalInMinutes.grid(row=0, column=1, padx=(5,0), sticky=tk.W)
			frameRepetitions = tk.Frame(acqFrame)
			frameRepetitions.grid(row=13, column=0, padx=10, sticky=tk.W)
			if True:
				label = tk.Label(frameRepetitions, text="Repetitions in each activation period:", font=("Verdana", 8))
				label.grid(row=0, column=0, padx=(0,5), sticky=tk.W)
				self.entry_PeriodicActivation_Repetitions = tk.Entry(frameRepetitions, textvariable=self.PeriodicActivation_Repetitions, width=3)
				self.entry_PeriodicActivation_Repetitions.grid(row=0, column=1, padx=(5,0), sticky=tk.W)

			self.checkbox_Module_AnomalyDetection_Command = ttk.Combobox(acqFrame, values=["Train", "Reset"], width=10, font=("Verdana", 8), state="readonly")  # variable=self.ML1_Command,
			self.checkbox_Module_AnomalyDetection_Command.current(0)
			self.checkbox_Module_AnomalyDetection_Command.grid(row=14, column=0, padx=0, sticky=tk.W)
			self.checkbox_Module_AnomalyDetection_Command.grid_forget()

			#tk.Checkbutton(acqFrame, text="Reset models", variable=self.ML1_Command, font=("Verdana", 8), command=self.StateBrowse, state="normal")
			frameAnomaly = tk.Frame(acqFrame)
			frameAnomaly.grid(row=15, column=0, padx=10, sticky=tk.W)
			if True:
				label = tk.Label(frameAnomaly, text="Command:", font=("Verdana", 8))
				label.grid(row=0, column=0, padx=(0,5), sticky=tk.W)
				label.grid_forget()
				self.checkbox_Module_AnomalyDetection_Command = ttk.Combobox(frameAnomaly, values=["Train", "Reset"], width=10, font=("Verdana", 8), state="readonly")  # variable=self.ML1_Command,
				self.checkbox_Module_AnomalyDetection_Command.current(0)
				self.checkbox_Module_AnomalyDetection_Command.grid(row=0, column=1, padx=(5,0), sticky=tk.W)
				self.checkbox_Module_AnomalyDetection_Command.grid_forget()

				# tk.Label(acqFrame, text="Machine Learning algorithm 1:", font=("Verdana", 8)).place(x=15, y=260)
				label = tk.Label(frameAnomaly, text="Preset to train:", font=("Verdana", 8))
				label.grid(row=1, column=0, padx=(0,5), sticky=tk.W)
				label.grid_forget()
				self.combobox_Module_AnomalyDetection_stateToTrain = ttk.Combobox(frameAnomaly, values=["Preset 1", "Preset 2", "Preset 3", "Preset 4"], width=10, font=("Verdana", 8), state="readonly")   # variable=self.ML1_stateToTrain,
				self.combobox_Module_AnomalyDetection_stateToTrain.current(0)
				self.combobox_Module_AnomalyDetection_stateToTrain.grid(row=1, column=1, padx=(5,0), sticky=tk.W)
				self.combobox_Module_AnomalyDetection_stateToTrain.grid_forget()

				label = tk.Label(frameAnomaly, text="Sensors:", font=("Verdana", 8))
				label.grid(row=2, column=0, padx=(0,5), sticky=tk.W)
				label.grid_forget()
				self.entry_Module_AnomalyDetection_sensors = tk.Entry(frameAnomaly, textvariable=self.Module_AnomalyDetection_sensors, width=5, state="normal")
				self.entry_Module_AnomalyDetection_sensors.grid(row=3, column=0, padx=(5,0), sticky=tk.W)
				self.entry_Module_AnomalyDetection_sensors.grid_forget()

			frameSetGet = tk.Frame(acqFrame)
			frameSetGet.grid(row=16, column=0, padx=10, pady=(70,10), sticky=tk.S)
			self.buttonSetConfiguration = tk.Button(frameSetGet, command=self.SetConfiguration, text="Set", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			self.buttonSetConfiguration.config(height=1, width=3)
			self.buttonSetConfiguration.grid(row=0, column=0, padx=(0,30), sticky=tk.W)

			self.buttonGetConfiguration = tk.Button(frameSetGet, command=self.GetConfiguration, text="Get", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			self.buttonGetConfiguration.config(height=1, width=3)
			self.buttonGetConfiguration.grid(row=0, column=1, padx=(30,0), sticky=tk.W)

			# self.cb_ConfigAbsoluteTime = tk.Checkbutton(acqFrame, text="Configure absolute time", variable=self.config_AbsoluteTime, font=("Verdana", 8), state="normal")
			# self.cb_ConfigAbsoluteTime.place(x=5, y=10, anchor="w")
			# self.config_AbsoluteTime.set(True)
			# self.cb_ConfigAbsoluteTime.place_forget()
			#
			# # Common options
			# self.cb_ConfigCommon = tk.Checkbutton(acqFrame, text="Configure common options:", variable=self.config_Common, font=("Verdana", 8), state="normal")
			# self.cb_ConfigCommon.place(x=5, y=30, anchor="w")
			# self.config_Common.set(True)
			# self.cb_ConfigCommon.place_forget()
			#
			# # tk.Label(acqFrame, text="Output channel:", font=("Verdana", 8)).place(x=35, y=5)
			# self.cb_transmit = tk.Checkbutton(acqFrame, text="Output channel:", variable=self.transmit, font=("Verdana", 8), state="normal")
			# self.cb_transmit.place(x=10, y=15, anchor="w")
			# self.transmit.set(True)
			# self.radioCommChannel_USB = tk.Radiobutton(acqFrame, text="USB", command=None, variable=self.radioCommChannel, value=0, font=("Verdana", 8))
			# self.radioCommChannel_USB.place(x=140, y=16, anchor="w")
			# self.radioCommChannel_AUX = tk.Radiobutton(acqFrame, text="TCP/IP / SMIP", command=None, variable=self.radioCommChannel, value=1, font=("Verdana", 8))
			# self.radioCommChannel_AUX.place(x=190, y=16, anchor="w")
			#
			# self.cb_enableVibrator = tk.Checkbutton(acqFrame, text="Enable vibration motor", variable=self.enableVibrator, font=("Verdana", 8), state="normal")
			# self.cb_enableVibrator.place(x=30, y=70, anchor="w")
			# self.cb_enableVibrator.place_forget()
			#
			# self.cb_saveToSD = tk.Checkbutton(acqFrame, text="Save to file in SD card", variable=self.saveToSD, font=("Verdana", 8), state="normal")
			# self.cb_saveToSD.place(x=10, y=35, anchor="w")
			#
			# self.cb_config_Module_RawData = tk.Checkbutton(acqFrame, text="Configure raw data module", variable=self.ConfigModule_RawData, font=("Verdana", 8), state="normal")
			# self.cb_config_Module_RawData.place(x=30, y=150, anchor="w")
			# self.ConfigModule_RawData.set(True)
			# self.cb_config_Module_RawData.place_forget()
			#
			# # self.cb_config_Module_AnomalyDetection = tk.Checkbutton(acqFrame, text="Configure anomaly detection module", variable=self.ConfigModule_AnomalyDetection, font=("Verdana", 8), state="normal")
			# # self.cb_config_Module_AnomalyDetection.place(x=30, y=330, anchor="w")
			# # self.ConfigModule_AnomalyDetection.set(False)
			# # self.cb_config_Module_AnomalyDetection.place_forget()
			#
			# # tk.Label(acqFrame, text="Modules to activate:", font=("Verdana", 8)).place(x=5, y=110)
			# self.cb_enable_Module_RawData = tk.Checkbutton(acqFrame, text="Raw sensors data to acquire:", variable=self.ActivateModule_RawData, font=("Verdana", 8), state="normal")
			# self.cb_enable_Module_RawData.place(x=10, y=55, anchor="w")
			# self.ActivateModule_RawData.set(True)
			# # Raw data module
			# # self.cb_ConfigAbsoluteTime = tk.Checkbutton(acqFrame, text="Raw sensors data to acquire:", variable=self.enable_RawData, font=("Verdana", 8), state="normal")
			# # self.cb_ConfigAbsoluteTime.place(x=5, y=210, anchor="w")
			#
			# self.lowPowerAccCheckAcq = tk.Checkbutton(acqFrame, text="Low power accelerometer", variable=self.lowPowerAccAcq, font=("Verdana", 8), state="normal")
			# self.lowPowerAccCheckAcq.place(x=30, y=75, anchor="w")
			# self.lowPowerAccAcq.set(1)
			#
			# self.accCheckAcq = tk.Checkbutton(acqFrame, text="Accelerometer", variable=self.accAcq, font=("Verdana", 8), state="normal")
			# self.accCheckAcq.place(x=30, y=95, anchor="w")
			# self.accAcq.set(1)
			#
			# self.tempCheckAcq = tk.Checkbutton(acqFrame, text="Thermometer", variable=self.tempAcq, font=("Verdana", 8), state="normal")
			# self.tempCheckAcq.place(x=30, y=115, anchor="w")
			# self.tempAcq.set(1)
			#
			# self.magCheckAcq = tk.Checkbutton(acqFrame, text="Magnetometer", variable=self.magAcq, font=("Verdana", 8), state="normal")
			# self.magCheckAcq.place(x=30, y=135, anchor="w")
			# self.magAcq.set(1)
			#
			# self.micCheckAcq = tk.Checkbutton(acqFrame, text="Microphone", variable=self.micAcq, font=("Verdana", 8), state="normal")
			# self.micCheckAcq.place(x=30, y=155, anchor="w")
			# self.micAcq.set(1)
			#
			# self.widebandAccCheckAcq = tk.Checkbutton(acqFrame, text="Wideband accelerometer (optional)", variable=self.widebandAccAcq, font=("Verdana", 8), state="normal")
			# self.widebandAccCheckAcq.place(x=30, y=175, anchor="w")
			# # self.widebandAccAcq.set(1)
			#
			# # Anomaly detection module
			# self.cb_configure_Module_AnomalyDetection = tk.Checkbutton(acqFrame, text="Configure anomaly detection algorithm", variable=self.ConfigModule_AnomalyDetection, font=("Verdana", 8), state="normal")
			# self.cb_configure_Module_AnomalyDetection.place(x=10, y=195, anchor="w")
			# self.cb_enable_Module_AnomalyDetection = tk.Checkbutton(acqFrame, text="Enable anomaly detection algorithm", variable=self.ActivateModule_AnomalyDetection, font=("Verdana", 8), state="normal")
			# self.cb_enable_Module_AnomalyDetection.place(x=30, y=215, anchor="w")
			# # self.cb_enable_Module_AnomalyDetection.place_forget()
			#
			# # Periodic activation
			# self.cb_configure_PeriodicActivation = tk.Checkbutton(acqFrame, text="Configure periodic activation", variable=self.config_PeriodicActivation, font=("Verdana", 8), state="normal")
			# self.cb_configure_PeriodicActivation.place(x=10, y=235, anchor="w")
			# label = tk.Label(acqFrame, text="Activation period [minutes]:", font=("Verdana", 8))
			# label.place(x=30, y=245)
			# self.entry_PeriodicActivation_IntervalInMinutes = tk.Entry(acqFrame, textvariable=self.PeriodicActivation_IntervalInMinutes, width=5)
			# self.entry_PeriodicActivation_IntervalInMinutes.place(x=200, y=245)
			# label = tk.Label(acqFrame, text="Repetitions in each activation period:", font=("Verdana", 8))
			# label.place(x=30, y=265)
			# self.entry_PeriodicActivation_Repetitions = tk.Entry(acqFrame, textvariable=self.PeriodicActivation_Repetitions, width=3)
			# self.entry_PeriodicActivation_Repetitions.place(x=250, y=265)
			#
			# self.checkbox_Module_AnomalyDetection_Command = ttk.Combobox(acqFrame, values=["Train", "Reset"], width=10, font=("Verdana", 8), state="readonly")  # variable=self.ML1_Command,
			# self.checkbox_Module_AnomalyDetection_Command.current(0)
			# self.checkbox_Module_AnomalyDetection_Command.place(x=150, y=390, anchor="w")
			# self.checkbox_Module_AnomalyDetection_Command.place_forget()
			#
			# #tk.Checkbutton(acqFrame, text="Reset models", variable=self.ML1_Command, font=("Verdana", 8), command=self.StateBrowse, state="normal")
			# label = tk.Label(acqFrame, text="Command:", font=("Verdana", 8))
			# label.place(x=50, y=380)
			# label.place_forget()
			# self.checkbox_Module_AnomalyDetection_Command = ttk.Combobox(acqFrame, values=["Train", "Reset"], width=10, font=("Verdana", 8), state="readonly")  # variable=self.ML1_Command,
			# self.checkbox_Module_AnomalyDetection_Command.current(0)
			# self.checkbox_Module_AnomalyDetection_Command.place(x=150, y=390, anchor="w")
			# self.checkbox_Module_AnomalyDetection_Command.place_forget()
			#
			# # tk.Label(acqFrame, text="Machine Learning algorithm 1:", font=("Verdana", 8)).place(x=15, y=260)
			# label = tk.Label(acqFrame, text="Preset to train:", font=("Verdana", 8))
			# label.place(x=50, y=400)
			# label.place_forget()
			# self.combobox_Module_AnomalyDetection_stateToTrain = ttk.Combobox(acqFrame, values=["Preset 1", "Preset 2", "Preset 3", "Preset 4"], width=10, font=("Verdana", 8), state="readonly")   # variable=self.ML1_stateToTrain,
			# self.combobox_Module_AnomalyDetection_stateToTrain.current(0)
			# self.combobox_Module_AnomalyDetection_stateToTrain.place(x=150, y=410, anchor="w")
			# self.combobox_Module_AnomalyDetection_stateToTrain.place_forget()
			#
			# label = tk.Label(acqFrame, text="Sensors:", font=("Verdana", 8))
			# label.place(x=50, y=420)
			# label.place_forget()
			#
			# self.entry_Module_AnomalyDetection_sensors = tk.Entry(acqFrame, textvariable=self.Module_AnomalyDetection_sensors, width=5, state="normal")
			# self.entry_Module_AnomalyDetection_sensors.place(x=150, y=430, anchor="w")
			# self.entry_Module_AnomalyDetection_sensors.place_forget()
			#
			# self.buttonSetConfiguration = tk.Button(acqFrame, command=self.SetConfiguration, text="Set", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			# self.buttonSetConfiguration.config(height=1, width=3)
			# self.buttonSetConfiguration.place(x=60, y=450)
			#
			# self.buttonGetConfiguration = tk.Button(acqFrame, command=self.GetConfiguration, text="Get", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			# self.buttonGetConfiguration.config(height=1, width=3)
			# self.buttonGetConfiguration.place(x=180, y=450)

		miscFrame = tk.LabelFrame(self.parent, text="Miscellaneous", width=250, height=150)
		miscFrame.grid(row=0, column=1, padx=(20, 10), pady=(20, 10), sticky="wn")
		if True:
			self.button_reset = tk.Button(miscFrame, command=self.Reset, text="Reset", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			self.button_reset.grid(row=0, column=0, padx=(10, 50), ipadx=40, sticky=tk.W)
			self.radio_reset_type_software = tk.Radiobutton(miscFrame, text="Software", variable=self.reset_type, value=iCOMOX_messages.cCOMOX_RESET_TYPE_Software, font=("Verdana", 8))
			self.radio_reset_type_software.grid(row=1, column=0, padx=10, sticky=tk.W)
			self.radio_reset_type_hardware = tk.Radiobutton(miscFrame, text="Hardware", variable=self.reset_type, value=iCOMOX_messages.cCOMOX_RESET_TYPE_Hardware, font=("Verdana", 8))
			self.radio_reset_type_hardware.grid(row=2, column=0, padx=10, sticky=tk.W)
			self.radio_reset_type_bootloader = tk.Radiobutton(miscFrame, text="Firmware update", variable=self.reset_type, value=iCOMOX_messages.cCOMOX_RESET_TYPE_FirmwareUpdate, font=("Verdana", 8))
			self.radio_reset_type_bootloader.grid(row=3, column=0, padx=10, sticky=tk.W)
			self.radio_reset_type_bootloader.grid_forget()

			self.button_bin_to_text_converter = tk.Button(miscFrame, command=self.BinToTextConverter, text="Binary file to text file converter...", font=("Verdana", 8), bg="gray80", fg="black", state="normal")
			self.button_bin_to_text_converter.grid(row=4, column=0, padx=10, pady=(20,5), sticky=tk.W)

		# Save to File
		fileFrame = tk.LabelFrame(self.parent, text="Save to File", width=380, height=200)
		fileFrame.grid(row=0, column=2, padx=(20, 10), pady=(20, 10), sticky="wn")
		if True:
			self.SaveToFile = tk.Checkbutton(fileFrame, text="Save to File", variable=self.save_to_file, font=("Verdana", 8), command=self.StateBrowse, state="normal")
			self.SaveToFile.grid(row=0, column=0, padx=10, sticky=tk.W,)
			self.save_to_file.set(True)      # Enable Save_to_file button

			frameSelectDir = tk.Frame(fileFrame)
			frameSelectDir.grid(row=1, column=0, padx=10, sticky=tk.W)
			tk.Label(frameSelectDir, text="Select the directory:", font=("Verdana", 8)).grid(row=0, column=0, padx=(0, 5), sticky=tk.W)
			self.SaveToDirButton = tk.Button(frameSelectDir, command=self.FileDir, text="Browse", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			self.SaveToDirButton.config(height=1, width=6)
			self.SaveToDirButton.grid(row=0, column=1, padx=0, ipadx=5, sticky=tk.W)

			tk.Label(fileFrame, text="Select sensor data to save:", font=("Verdana", 8)).grid(row=2, column=0, padx=10, sticky=tk.W)
			frameSensors = tk.Frame(fileFrame)
			frameSensors.grid(row=3, column=0, padx=10, sticky=tk.W)
			if True:
				self.lowPowerAccCheckFile = tk.Checkbutton(frameSensors, text="Low power accelerometer", variable=self.lowPowerAccFile, font=("Verdana", 8), command=self.StateStartStop, state="disabled")
				self.lowPowerAccCheckFile.grid(row=0, column=0, padx=0, sticky=tk.W)
				self.accCheckFile = tk.Checkbutton(frameSensors, text="Accelerometer", variable=self.accFile, font=("Verdana", 8), command=self.StateStartStop, state="disabled")
				self.accCheckFile.grid(row=1, column=0, padx=0, sticky=tk.W)
				self.accFile.set(True)      # Enable Accelerometer button

				self.magCheckFile = tk.Checkbutton(frameSensors, text="Magnetometer", variable=self.magFile, font=("Verdana", 8), command=self.StateStartStop, state="disabled")
				self.magCheckFile.grid(row=2, column=0, padx=0, sticky=tk.W)
				self.micCheckFile = tk.Checkbutton(frameSensors, text="Microphone", variable=self.micFile, font=("Verdana", 8), command=self.StateStartStop, state="disabled")
				self.micCheckFile.grid(row=0, column=1, padx=0, sticky=tk.W)
				self.tempCheckFile = tk.Checkbutton(frameSensors, text="Thermometer", variable=self.tempFile, font=("Verdana", 8), command=self.StateStartStop, state="disabled")
				self.tempCheckFile.grid(row=1, column=1, padx=0, sticky=tk.W)
				self.widebandAccCheckFile = tk.Checkbutton(frameSensors, text="Wideband accelerometer", variable=self.widebandAccFile, font=("Verdana", 8), command=self.StateStartStop, state="disabled")
				self.widebandAccCheckFile.grid(row=2, column=1, padx=0, sticky=tk.W)

			frameStartStopLog = tk.Frame(fileFrame)
			frameStartStopLog.grid(row=4, column=0, padx=10, pady=10, sticky=tk.W)
			self.StartLogButton = tk.Button(frameStartStopLog, command=self.StartPressed, text="Start", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			self.StartLogButton.grid(row=0, column=0, ipadx=10, padx=(0,10), sticky=tk.W)
			self.StopLogButton = tk.Button(frameStartStopLog, command=self.StopPressed, text="Stop", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
			self.StopLogButton.grid(row=0, column=1, ipadx=10, padx=(10,0), sticky=tk.W)

		# scheduleFrame = tk.LabelFrame(self.parent, text="Transmission mode", width=250, height=150)
		# scheduleFrame.grid(row=0, column=1, padx=10, pady=(20, 10), sticky="wn")
		# transmission mode options (Schedule)
		# self.stream_schedule.set(1)
		# self.radio_stream = tk.Radiobutton(scheduleFrame, text="Stream", command=self.StateStreamSchedule, variable=self.stream_schedule, value=1, font=("Verdana", 8))
		# self.radio_stream.place(x=10, y=15, anchor="w")
		# self.radio_schedule = tk.Radiobutton(scheduleFrame, text="On schedule", command=self.StateStreamSchedule, variable=self.stream_schedule, value=2, font=("Verdana", 8))
		# self.radio_schedule.place(x=10, y=35, anchor="w")
		# tk.Label(scheduleFrame, text="Every ", font=("Verdana", 8)).place(x=30, y=45)
		# self.scheduleEntry = tk.Entry(scheduleFrame, textvariable=self.schedule, width=5, state="disabled")
		# self.scheduleEntry.place(x=75, y=55, anchor="w")
		# tk.Label(scheduleFrame, text=" hours", font=("Verdana", 8)).place(x=110, y=45)
		# tk.Label(scheduleFrame, text="for ", font=("Verdana", 8)).place(x=40, y=65)
		# self.timeEntry = tk.Entry(scheduleFrame, textvariable=self.time, width=5, state="disabled")
		# self.timeEntry.place(x=75, y=75, anchor="w")
		# tk.Label(scheduleFrame, text=" minutes", font=("Verdana", 8)).place(x=110, y=65)

	def EnableButtons(self, enable):
		state = "normal" if enable else "disabled"
		self.button_reset.config(state=state)
		self.buttonGetConfiguration.config(state=state)
		self.buttonSetConfiguration.config(state=state)

	def Reset(self):
		common.app.iCOMOX_Data.send_msg(msg=iCOMOX_messages.OUT_MSG_Reset(ResetType=self.reset_type.get()))

	def MsgToWindow(self, getConfigurationInMsg):
		Common, \
		Repetition, IntervalInMinutes, \
		ActiveModules, \
		RawDataSensors, \
		ML1_target, ML1_selSensor, ML1_countLearned, ML1_modelLearned, \
		MaintenanceModule_sensors, MaintenanceModule_flags, MaintenanceModule_alpha = iCOMOX_messages.IN_MSG_GetConfiguration(msg=getConfigurationInMsg)
		self.CommonToWindow(Common=Common)
		self.PeriodicActivationToWindow(IntervalInMinutes=IntervalInMinutes, Repetition=Repetition)
		self.ActiveModulesToWindow(ActiveModules=ActiveModules)
		self.ModuleRawDataToWindow(RawDataSensors=RawDataSensors)
		self.ModuleAnomalyDetectionToWindow(stateToTrain=ML1_target, sensors=ML1_selSensor)
		# helpers.OUT(
		#     "IN_MSG_GetConfiguration:\nCommon = 0x{:02X}\nActive modules = 0x{:02X}\nRawDataSensors = 0x{:02X}\nTransmit repetition = {}\nTransmit interval in minutes = {}\nML1: target={}, selSensor=0x{:02X}, count learned={}, model learned={}".format(
		#         Common,
		#         ConfigBitmask,
		#         RawDataSensors,
		#         TransmitRepetition,
		#         IntervalInMinutes,
		#         ML1_target, ML1_selSensor, ML1_countLearned, ML1_modelLearned
		#     ))

	def WindowToMsg(self):
		ConfigBitmask = self.WindowToConfigBitmask()
		ConfigModulesBitmask = self.WindowToConfigModulesBitmask()
		LocalTimestamp = int(time.time() + (datetime.datetime.now() - datetime.datetime.utcnow()).total_seconds())   # mktime(time.localtime()))
		IntervalInMinutes, Repetition = self.WindowToPeriodicActivation()
		Common = self.WindowToCommon()
		ActiveModules = self.WindowToActiveModules()
		RawData_Sensors = self.WindowToModuleRawData()
		# Module_AnomalyDetection_Command, Module_AnomalyDetection_stateToTrain, Module_AnomalyDetection_Sensors = self.WindowToModuleAnomalyDetection()
		Module_AnomalyDetection_Sensors = iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL356
		Module_AnomalyDetection_stateToTrain = 0
		Module_AnomalyDetection_Command = iCOMOX_messages.cANOMALY_DETECTION_COMMAND_Reset

		return iCOMOX_messages.OUT_MSG_SetConfiguration(ConfigBitmask=ConfigBitmask, ConfigModulesBitmask=ConfigModulesBitmask, \
														LocalTimestamp=LocalTimestamp, \
														Common=Common, \
														Repetition=Repetition, IntervalInMinutes=IntervalInMinutes, \
														ActiveModules=ActiveModules, \
														RawData_Sensors=RawData_Sensors, \
														AnomalyDetection_Command=Module_AnomalyDetection_Command, AnomalyDetection_Sensors=Module_AnomalyDetection_Sensors, AnomalyDetection_StateToTrain=Module_AnomalyDetection_stateToTrain)

	def SetConfiguration(self):
		common.app.StatusBar.set("Sent SetConfiguration request")
		common.app.iCOMOX_Data.send_msg(self.WindowToMsg())

	def GetConfiguration(self):
		common.app.StatusBar.set("Sent GetConfiguration request")
		common.app.iCOMOX_Data.send_msg(iCOMOX_messages.OUT_MSG_GetConfiguration())

	def BinToTextConverter(self):
		bin_file_name = filedialog.askopenfilename(initialdir="", title="select binary file to convert", filetypes=(("binary files", "iCOMOX_*.bin"), ("all files", "*.*")))
		if bin_file_name == "":
			return
		text_file_name = os.path.splitext(bin_file_name)[0] + ".txt"
		conversionError, reportsCount = BinFileConversion.BinFileConversion(bin_file_name=bin_file_name, text_file_name=text_file_name)
		if BinFileConversion.cBIN_FILE_CONVERSION_OK == conversionError:
			messagebox.showinfo(title="Conversion information", message="Conversion has been succeeded. {} report messages has been found".format(reportsCount))
		else:
			if BinFileConversion.cBIN_FILE_CONVERSION_UnrecognizedMsg == conversionError:
				error_msg = "Unrecognized message has been found in file {}".format(bin_file_name)
			elif BinFileConversion.cBIN_FILE_CONVERSION_NonReportMsgFound == conversionError:
				error_msg = "Non report messages has been found in file {}".format(bin_file_name)
			elif BinFileConversion.cBIN_FILE_CONVERSION_NotEnoughBytesInTheLastMsg == conversionError:
				error_msg = "No enough bytes in the last message stored in file {}".format(bin_file_name)
			elif BinFileConversion.cBIN_FILE_CONVERSION_NotEnoughBytes == conversionError:
				error_msg = "No enough bytes in the file {}".format(bin_file_name)
			elif BinFileConversion.cBIN_FILE_CONVERSION_UnrecognizedHeader == conversionError:
				error_msg = "File {} has an unrecognized header".format(bin_file_name)
			else:
				Exception("Unrecognized return code from BinToTextConverter()")
			messagebox.showerror(title="Conversion error", message=error_msg)

	def WindowToActiveModules(self):
		ActiveModules = 0
		if self.ActivateModule_RawData.get():
			ActiveModules |= iCOMOX_messages.cMODULE_BITMASK_RawData
		if self.ActivateModule_AnomalyDetection.get():
			ActiveModules |= iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection
		return ActiveModules

	def ActiveModulesToWindow(self, ActiveModules):
		self.ActivateModule_RawData.set((ActiveModules & iCOMOX_messages.cMODULE_BITMASK_RawData) != 0)
		self.ActivateModule_AnomalyDetection.set((ActiveModules & iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection) != 0)

	def WindowToConfigBitmask(self):
		ConfigBitmask = 0
		if self.config_AbsoluteTime.get():
			ConfigBitmask |= iCOMOX_messages.cCOMOX_CONFIG_BITMASK_AbsoluteTime
		if self.config_Common.get():
			ConfigBitmask |= iCOMOX_messages.cCOMOX_CONFIG_BITMASK_Common
		if self.config_PeriodicActivation.get():
			ConfigBitmask |= iCOMOX_messages.cCOMOX_CONFIG_BITMASK_PeriodicActivation
		return ConfigBitmask

	def ConfigBitmaskToWindow(self, ConfigBitmask):
		self.config_AbsoluteTime.set((ConfigBitmask & iCOMOX_messages.cCOMOX_CONFIG_BITMASK_AbsoluteTime) != 0)
		self.config_Common.set((ConfigBitmask & iCOMOX_messages.cCOMOX_CONFIG_BITMASK_Common) != 0)
		self.config_PeriodicActivation((ConfigBitmask & iCOMOX_messages.cCOMOX_CONFIG_BITMASK_PeriodicActivation) != 0)

	def WindowToConfigModulesBitmask(self):
		ConfigModulesBitmask = 0
		if self.ConfigModule_RawData.get():
			ConfigModulesBitmask |= iCOMOX_messages.cMODULE_BITMASK_RawData
		if self.ConfigModule_AnomalyDetection.get():
			ConfigModulesBitmask |= iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection
		return ConfigModulesBitmask

	def ConfigModulesBitmaskToWindow(self, ConfigModulesBitmask):
		self.cb_config_Module_RawData.set((ConfigModulesBitmask & iCOMOX_messages.cMODULE_BITMASK_RawData) != 0)
		self.cb_config_Module_AnomalyDetection.set((ConfigModulesBitmask & iCOMOX_messages.cMODULE_BITMASK_AnomalyDetection) != 0)

	def WindowToPeriodicActivation(self):
		return int(self.PeriodicActivation_IntervalInMinutes.get()), int(self.PeriodicActivation_Repetitions.get())

	def PeriodicActivationToWindow(self, IntervalInMinutes, Repetition):
		self.PeriodicActivation_IntervalInMinutes.set(IntervalInMinutes)
		self.PeriodicActivation_Repetitions.set(Repetition)

	def WindowToModuleRawData(self):
		RawDataSensors = 0
		if self.lowPowerAccAcq.get():
			RawDataSensors |= iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL362
		if self.accAcq.get():
			RawDataSensors |= iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL356
		if self.magAcq.get():
			RawDataSensors |= iCOMOX_messages.cCOMOX_SENSOR_BITMASK_BMM150
		if self.micAcq.get():
			RawDataSensors |= iCOMOX_messages.cCOMOX_SENSOR_BITMASK_IM69D130
		if self.tempAcq.get():
			RawDataSensors |= iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADT7410
		if self.widebandAccAcq.get():
			RawDataSensors |= iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL1002
		return RawDataSensors

	def ModuleRawDataToWindow(self, RawDataSensors):
		self.lowPowerAccAcq.set((RawDataSensors & iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL362) != 0)
		self.accAcq.set((RawDataSensors & iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL356) != 0)
		self.magAcq.set((RawDataSensors & iCOMOX_messages.cCOMOX_SENSOR_BITMASK_BMM150) != 0)
		self.micAcq.set((RawDataSensors & iCOMOX_messages.cCOMOX_SENSOR_BITMASK_IM69D130) != 0)
		self.tempAcq.set((RawDataSensors & iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADT7410) != 0)
		self.widebandAccAcq.set((RawDataSensors & iCOMOX_messages.cCOMOX_SENSOR_BITMASK_ADXL1002) != 0)

	def WindowToCommon(self):
		Common = 0
		if self.enableVibrator.get():
			Common |= iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_Vibrator
		if self.saveToSD.get():
			Common |= iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_SaveToFile
		if self.transmit.get():
			Common |= iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_Transmit
		if self.radioCommChannel.get():
			Common |= iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_CommChannelIsAux
		return Common

	def CommonToWindow(self, Common):
		self.enableVibrator.set((Common & iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_Vibrator) != 0)
		self.saveToSD.set((Common & iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_SaveToFile) != 0)
		self.transmit.set((Common & iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_Transmit) != 0)
		self.radioCommChannel.set((Common & iCOMOX_messages.cCOMOX_CONFIGURATION_Activate_CommChannelIsAux) != 0)

	def ModuleAnomalyDetectionToWindow(self, sensors, stateToTrain):
		self.Module_AnomalyDetection_sensors.set(sensors)
		self.combobox_Module_AnomalyDetection_stateToTrain.current(stateToTrain)

	def WindowToModuleAnomalyDetection(self):
		Command = self.checkbox_Module_AnomalyDetection_Command.current()
		stateToTrain = self.combobox_Module_AnomalyDetection_stateToTrain.current()
		Sensors = int(self.Module_AnomalyDetection_sensors.get())
		return Command, stateToTrain, Sensors

	def EnableSensorsToSaveToExcel(self, enabled):
		if enabled:
			state = "normal"
		else:
			state = "disabled"
		self.lowPowerAccCheckFile.config(state=state)
		self.accCheckFile.config(state=state)
		self.magCheckFile.config(state=state)
		self.micCheckFile.config(state=state)
		self.tempCheckFile.config(state=state)
		self.widebandAccCheckFile.config(state=state)

	def FileDir(self):
		# self.dirname = filedialog.askdirectory(initialdir=os.getcwd(), title='Please select the directory')
		self.dirname = "C:/Users/Public/Documents/"
		self.EnableSensorsToSaveToExcel(enabled=self.dirname is not None)

	def StateStartStop(self):
		if (self.lowPowerAccFile.get() + self.accFile.get() + self.magFile.get() + self.micFile.get() + self.tempFile.get() + self.widebandAccFile.get()) > 0:
			self.StartLogButton.configure(state="normal")
			self.StopLogButton.configure(state="disabled")
		else:
			self.StartLogButton.configure(state="disabled")
			self.StopLogButton.configure(state="disabled")

	def StateBrowse(self):
		if self.save_to_file.get() == 1:
			self.SaveToDirButton.configure(state="normal")
		if self.save_to_file.get() == 0:
			self.SaveToDirButton.configure(state="disabled")
			self.StartLogButton.configure(state="disabled")
			self.StopLogButton.configure(state="disabled")
			self.EnableSensorsToSaveToExcel(enabled=False)
			self.start_pressed = False
			self.stop_pressed = False

	def StartPressed(self):
		if self.start_pressed == False:
			self.start_pressed = True
			self.stop_pressed = False
			self.StartLogButton.configure(state="disabled")
			self.StopLogButton.configure(state="normal")
			self.EnableSensorsToSaveToExcel(enabled=False)
			self.SaveToFile.config(state="disabled")
			self.SaveToDirButton.config(state="disabled")
			right_now = datetime.datetime.now()
			self.logfilename = self.dirname + "/iCOMOX_{:04}_{:02}_{:02}_{:02}_{:02}_{:02}.xlsx".format(right_now.year, right_now.month, right_now.day, right_now.hour, right_now.minute, right_now.second)
			self.time_to_save = None
			self.workbook = openpyxl.Workbook()
			self.workbook.active.title = 'ID'
			if common.app.iCOMOX_Data.logfile_info is not None:
				self.Workbook_UpdateID(logfile_info=common.app.iCOMOX_Data.logfile_info)

			# Create headers in the tabsheets of the workbook
			if self.lowPowerAccFile.get() == 1:
				self.WS_ADXL362 = self.workbook.create_sheet("Low power accelerometer")
				if self.WS_ADXL362 is not None:
					self.WS_ADXL362.append(["Timestamp", "Time [sec]", "Ax [g]", "Ay [g]", "Az [g]"])
			else:
				self.WS_ADXL362 = None

			if self.accFile.get() == 1:
				self.WS_ADXL356 = self.workbook.create_sheet("Accelerometer")
				if self.WS_ADXL356 is not None:
					self.WS_ADXL356.append(["Timestamp", "Time [sec]", "Ax [g]", "Timestamp", "Time [sec]", "Ay [g]", "Timestamp", "Time [sec]", "Az [g]"])
			else:
				self.WS_ADXL356 = None

			if self.magFile.get() == 1:
				self.WS_BMM150 = self.workbook.create_sheet("Magnetometer")
				if self.WS_BMM150 is not None:
					self.WS_BMM150.append(["Timestamp", "Time [sec]", "Bx [μTesla]", "By [μTesla]", "Bz [μTesla]"])
			else:
				self.WS_BMM150 = None

			if self.micFile.get() == 1:
				self.WS_IM69D130 = self.workbook.create_sheet("Microphone")
				if self.WS_IM69D130 is not None:
					self.WS_IM69D130.append(["Timestamp", "Time [sec]", "Sound [SPL]"])
			else:
				self.WS_IM69D130 = None

			if self.tempFile.get() == 1:
				self.WS_ADT7410 = self.workbook.create_sheet("Temperature")
				if self.WS_ADT7410 is not None:
					self.WS_ADT7410.append(["Timestamp", "Temperature [°C]"])
			else:
				self.WS_ADT7410 = None

			if self.widebandAccFile.get() == 1:
				self.WS_ADXL1002 = self.workbook.create_sheet("Wideband accelerometer")
				if self.WS_ADXL1002 is not None:
					self.WS_ADXL1002.append(["Timestamp", "Time [sec]", "Ax [g]"])
			else:
				self.WS_ADXL1002 = None

	def StopPressed(self):
		if self.stop_pressed == False:
			self.stop_pressed = True
			self.start_pressed = False
			self.StopLogButton.configure(state="disabled")
			self.StartLogButton.configure(state="normal")
			self.lowPowerAccCheckFile.config(state="normal")
			self.accCheckFile.config(state="normal")
			self.magCheckFile.config(state="normal")
			self.micCheckFile.config(state="normal")
			self.tempCheckFile.config(state="normal")
			self.widebandAccCheckFile.config(state="normal")
			self.SaveToFile.config(state="normal")
			self.SaveToDirButton.config(state="normal")
			#self.time_to_save = time.sleep(10)

			# self.workbook.remove(self.workbook["Sheet"])
			self.workbook.save(self.logfilename)

			import xlrd         # convert xlsx to txt
			import csv
			with xlrd.open_workbook(self.logfilename) as wb:
				sh = wb.sheet_by_name('Accelerometer')
				with open(self.dirname+'/iCOMOX-data.txt', 'w') as f:
					c = csv.writer(f)
					for r in range(sh.nrows):
						c.writerow(sh.row_values(r))

			# self.xlsx2csv()
			self.logfilename = None
			self.time_to_save = None
			self.workbook = None
			self.WS_ADXL362 = None
			self.WS_ADXL356 = None
			self.WS_BMM150 = None
			self.WS_ADT7410 = None
			self.WS_IM69D130 = None
			self.WS_ADXL1002 = None

			self.WS_ADXL356_Batches = [0, 0, 0]

	# def StateStreamSchedule(self):
	#     if self.stream_schedule.get() == 1:
	#         self.scheduleEntry.configure(state="disabled")
	#         self.timeEntry.configure(state="disabled")
	#     if self.stream_schedule.get() == 2:
	#         self.scheduleEntry.configure(state="normal")
	#         self.timeEntry.configure(state="normal")

	def Workbook_UpdateID(self, logfile_info):
		self.logfile_info = logfile_info
		if self.workbook is not None:
			if self.time_to_save is None:
				ws = self.workbook["ID"]
				for i in range(0, len(self.logfile_info)):
					ws["A{}".format(i+1)] = self.logfile_info[i]
				self.time_to_save = time.time()

	def Workbook_Update_ADXL362(self, time_ns_acquired, acc_x_units_g, acc_y_units_g, acc_z_units_g):
		if self.workbook is not None and self.WS_ADXL362 is not None:
			self.WS_ADXL362.append([timestamp_ns_to_str(timestamp_ns=time_ns_acquired), 0.0, acc_x_units_g[0], acc_y_units_g[0], acc_z_units_g[0]])
			t = 0.0
			delta_t = 1/common.ADXL362.Fs_Acc
			for i in range(1, iCOMOX_messages.ADXL362_SAMPLES_NUM):
				t += delta_t
				self.WS_ADXL362.append(["", t, acc_z_units_g[i], acc_y_units_g[i], acc_z_units_g[i]])

	def Workbook_Update_ADXL356(self, time_ns_acquired, axis, acc_units_g, adxl356_smip):
		if self.workbook is not None and self.WS_ADXL356 is not None:
			if adxl356_smip:
				samples_number_per_axis = iCOMOX_messages.ADXL356_SMIP_SAMPLES_NUM_PER_PIN
				delta_t = 1/common.ADXL356_SMIP.Fs_Acc
			else:
				samples_number_per_axis = iCOMOX_messages.ADXL356_SAMPLES_NUM_PER_PIN
				delta_t = 1/common.ADXL356.Fs_Acc
			base_row = 2 + samples_number_per_axis * self.WS_ADXL356_Batches[axis]
			timestamp_col = chr(ord("A")+3*axis)
			time_col = chr(ord(timestamp_col)+1)
			data_col = chr(ord(time_col)+1)

			self.WS_ADXL356["{}{}".format(timestamp_col, base_row)] = timestamp_ns_to_str(timestamp_ns=time_ns_acquired)
			t = 0.0
			for row in range(base_row, base_row + samples_number_per_axis):
				self.WS_ADXL356["{}{}".format(time_col, row)] = t
				self.WS_ADXL356["{}{}".format(data_col, row)] = acc_units_g[row - base_row]
				t += delta_t

			# Mark for the next update
			self.WS_ADXL356_Batches[axis] += 1

	def Workbook_Update_BMM150(self, time_ns_acquired, mag_x_units_tesla, mag_y_units_tesla, mag_z_units_tesla):
		if self.workbook is not None and self.WS_BMM150 is not None:
			self.WS_BMM150.append([timestamp_ns_to_str(timestamp_ns=time_ns_acquired), 0.0, mag_x_units_tesla[0], mag_y_units_tesla[0], mag_z_units_tesla[0]])
			t = 0.0
			delta_t = 1/common.BMM150.Fs_Mag
			for i in range(1, iCOMOX_messages.BMM150_SAMPLES_NUM):
				t += delta_t
				self.WS_BMM150.append(["", t, mag_x_units_tesla[i], mag_y_units_tesla[i], mag_z_units_tesla[i]])

	def Workbook_Update_ADT7410(self, time_ns_acquired, temp_unit_celsius):
		if self.workbook is not None and self.WS_ADT7410 is not None:
			self.WS_ADT7410.append([timestamp_ns_to_str(timestamp_ns=time_ns_acquired), temp_unit_celsius])

	def Workbook_Update_IM69D130(self, time_ns_acquired, sound_units_SPL):
		if self.workbook is not None and self.WS_IM69D130 is not None:
			self.WS_IM69D130.append([timestamp_ns_to_str(timestamp_ns=time_ns_acquired), 0.0, sound_units_SPL[0]])
			t = 0
			delta_t = 1/common.IM69D130.Fs_Mic
			for i in range(1, iCOMOX_messages.IM69D130_SAMPLES_NUM):
				t += delta_t
				self.WS_IM69D130.append(["", t, sound_units_SPL[i]])

	def Workbook_Update_ADXL1002(self, time_ns_acquired, acc_units_g):
		if self.workbook is not None and self.WS_ADXL1002 is not None:
			self.WS_ADXL1002.append([timestamp_ns_to_str(timestamp_ns=time_ns_acquired), 0.0, acc_units_g[0]])
			t = 0.0
			delta_t = 1/common.ADXL1002.Fs_Acc
			for i in range(1, iCOMOX_messages.ADXL1002_SAMPLES_NUM):
				t += delta_t
				self.WS_ADXL1002.append(["", t, acc_units_g[i]])


class Information(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent
		self.label_iCOMOX_version_title = tk.Label(self.parent, text="", justify="center", anchor="center", font=("Verdana", 10), bd=0)
		self.label_iCOMOX_version_title.grid(row=0, column=0, ipadx=20, ipady=10, sticky="we")
		self.label_iCOMOX_version = tk.Label(self.parent, justify="left", anchor="w", text="", font=("Verdana", 10), bd=0)
		self.label_iCOMOX_version.grid(row=1, column=0, ipadx=20, ipady=10, sticky="we")

		self.label_dongle_version_title = tk.Label(self.parent, text="", justify="center", anchor="center", font=("Verdana", 10), bd=0)
		self.label_dongle_version_title.grid(row=2, column=0, ipadx=20, ipady=10, sticky="we")
		self.label_dongle_version = tk.Label(self.parent, justify="left", anchor="w", text="", font=("Verdana", 10), bd=0)
		self.label_dongle_version.grid(row=3, column=0, ipadx=20, ipady=10, sticky="we")

		self.dongle_version_available = False
		self.dongle_hwModel = 0
		self.dongle_hwRev = 0
		self.dongle_swMajor = 0
		self.dongle_swMinor = 0
		self.dongle_swPatch = 0
		self.dongle_swBuild = 0

		self.icomox_version_available = False
		self.icomox_board_type = 0
		self.icomox_board_version_major = 0
		self.icomox_board_version_minor = 0
		self.icomox_mcu_serial_number = 0
		self.icomox_firmware_release_version_major = 0
		self.icomox_firmware_release_version_minor = 0
		self.icomox_firmware_release_version_patch = 0
		self.icomox_firmware_release_version_branch = 0
		self.icomox_firmware_build_version_year = 0
		self.icomox_firmware_build_version_month = 0
		self.icomox_firmware_build_version_day = 0
		self.icomox_firmware_build_version_hour = 0
		self.icomox_firmware_build_version_min = 0
		self.icomox_firmware_build_version_sec = 0

		self.icomox_production_serial_number = bytearray()
		self.icomox_product_part_number = bytearray()
		self.icomox_name = bytearray()

		self.icomox_bit_status = 0

		self.icomox_smip_swMajor = 0
		self.icomox_smip_swMinor = 0
		self.icomox_smip_swPatch = 0
		self.icomox_smip_swBuild = 0

	def clear(self, clear_iCOMOX_only=False):
		if not clear_iCOMOX_only:
			self.update_dongle_version(dongle_version_available=False)
		self.update_iCOMOX_version(icomox_version_available=False)

	def get_iCOMOX_version_string(self):
		BoardVersionStr = messages_utils.iCOMOX_BoardVersion_to_Str(board_version_major=self.icomox_board_version_major, board_version_minor=self.icomox_board_version_minor)
		if BoardVersionStr != "":
			BoardVersionStr = "iCOMOX board version: {}\n".format(BoardVersionStr)
		if self.icomox_board_type == iCOMOX_messages.cCOMOX_BOARD_TYPE_SMIP:
			SmipVersionStr = "\nSMIP software version: {}.{}.{}.{}".format(
				self.icomox_smip_swMajor,
				self.icomox_smip_swMinor,
				self.icomox_smip_swPatch,
				self.icomox_smip_swBuild
			)
		elif self.icomox_board_type == iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT:
			SmipVersionStr = "\nNB-IOT test: 0x{:02X}".format(self.icomox_smip_swMajor)
		else:
			SmipVersionStr = ""
		return "iCOMOX board type: {}\n{}MCU serial number: {:^32}\nFirmware release version: {}.{}.{} {}\nFirmware build version: {:02}.{:02}.{:04} {:02}:{:02}:{:02}\nProduction serial number: {}\nPart number: {}\nName: {}\nBIT state code: 0x{:02X}{}".format(
			messages_utils.iCOMOX_BoardType_to_Str(self.icomox_board_type),
			BoardVersionStr,
			helpers.u8s_to_str(arr=self.icomox_mcu_serial_number, separator=""),
			self.icomox_firmware_release_version_major, self.icomox_firmware_release_version_minor, self.icomox_firmware_release_version_patch,
			messages_utils.iCOMOX_firmware_release_version_branch_to_Str(firmware_release_version_branch=self.icomox_firmware_release_version_branch),
			self.icomox_firmware_build_version_day, self.icomox_firmware_build_version_month,
			self.icomox_firmware_build_version_year, self.icomox_firmware_build_version_hour,
			self.icomox_firmware_build_version_min, self.icomox_firmware_build_version_sec,
			helpers.bytearrayToString(self.icomox_production_serial_number),
			helpers.bytearrayToString(self.icomox_product_part_number),
			helpers.bytearrayToString(self.icomox_name),
			self.icomox_bit_status,
			SmipVersionStr
		)

	def update_iCOMOX_version(self, icomox_version_available=False, board_type=0, board_version_major=0, board_version_minor=0, mcu_serial_number=0, firmware_release_version_major=0, firmware_release_version_minor=0, firmware_release_version_patch=0, firmware_release_version_branch=0, firmware_build_version_year=0, firmware_build_version_month=0, firmware_build_version_day=0, firmware_build_version_hour=0, firmware_build_version_min=0, firmware_build_version_sec=0, icomox_name="", bit_status=0, product_part_number="", production_serial_number="", name="", smip_swMajor=0, smip_swMinor=0, smip_swPatch=0, smip_swBuild=0):
		self.icomox_version_available = icomox_version_available
		if icomox_version_available:
			self.icomox_board_type = board_type
			self.icomox_board_version_major = board_version_major
			self.icomox_board_version_minor = board_version_minor
			self.icomox_mcu_serial_number = mcu_serial_number
			self.icomox_firmware_release_version_major = firmware_release_version_major
			self.icomox_firmware_release_version_minor = firmware_release_version_minor
			self.icomox_firmware_release_version_patch = firmware_release_version_patch
			self.icomox_firmware_release_version_branch = firmware_release_version_branch
			self.icomox_firmware_build_version_year = firmware_build_version_year
			self.icomox_firmware_build_version_month = firmware_build_version_month
			self.icomox_firmware_build_version_day = firmware_build_version_day
			self.icomox_firmware_build_version_hour = firmware_build_version_hour
			self.icomox_firmware_build_version_min = firmware_build_version_min
			self.icomox_firmware_build_version_sec = firmware_build_version_sec
			self.icomox_bit_status = bit_status
			self.icomox_product_part_number=product_part_number
			self.icomox_production_serial_number = production_serial_number
			self.icomox_name = name
			if self.icomox_board_type == iCOMOX_messages.cCOMOX_BOARD_TYPE_SMIP:
				self.icomox_smip_swMajor = smip_swMajor
				self.icomox_smip_swMinor = smip_swMinor
				self.icomox_smip_swPatch = smip_swPatch
				self.icomox_smip_swBuild = smip_swBuild
			elif self.icomox_board_type == iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT:
				self.icomox_smip_swMajor = smip_swMajor

			self.label_iCOMOX_version_title.configure(text="iCOMOX:")
			self.label_iCOMOX_version.configure(text=self.get_iCOMOX_version_string())
		else:
			self.label_iCOMOX_version_title.configure(text="")
			self.label_iCOMOX_version.configure(text="")

	def update_dongle_version(self, dongle_version_available=False, hwModel=0, hwRev=0, swMajor=0, swMinor=0, swPatch=0, swBuild=0):
		self.dongle_version_available = dongle_version_available
		if dongle_version_available:
			self.dongle_hwModel = hwModel
			self.dongle_hwRev = hwRev
			self.dongle_swMajor = swMajor
			self.dongle_swMinor = swMinor
			self.dongle_swPatch = swPatch
			self.dongle_swBuild = swBuild
			self.label_dongle_version_title.configure(text="SmartMesh IP dongle:")
			self.label_dongle_version.configure(text="Dongle hardware model: {}\nDongle hardware revision: {}\nDongle software version: {}.{}.{}.{}".format(self.dongle_hwModel, self.dongle_hwRev, self.dongle_swMajor, self.dongle_swMinor, self.dongle_swPatch, self.dongle_swBuild))
		else:
			self.label_dongle_version_title.configure(text="")
			self.label_dongle_version.configure(text="")


class EEPROM(tk.Frame):
	cSTATE_INIT         = 0
	cSTATE_OK           = 1
	cSTATE_ERROR        = 2
	cSTATE_FINISH_OK    = 3
	cSTATE_FINISH_FAIL  = 4

	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent
		self.PartNumber = tk.StringVar(value=MemMap.list_iCOMOX_PartNumber[1])
		self.SerialNumber = tk.StringVar(value="")
		self.BoardVersionMajor = tk.IntVar(value=1)
		self.BoardVersionMinor = tk.IntVar(value=3)

		self.Name = tk.StringVar(value="iCOMOX NB-IoT")

		# NBIOT specific variables
		self.NBIOT_DefaultSIMConfig = tk.BooleanVar(value=True)
		self.NBIOT_BitmaskAcT = [tk.BooleanVar(value=False) for _ in range(3)]   # EGPRS & LTE cat. M1
		self.NBIOT_ScanBands = [tk.StringVar(value="0000") for _ in range(3)]
		self.NBIOT_EnableRoaming = tk.BooleanVar(value=True)

		self.NBIOT_ManualOperatorSelection = tk.BooleanVar(value=False)
		self.NBIOT_PLMN = tk.StringVar(value="42503")

		self.NBIOT_ApnAccessName = tk.StringVar(value="sphone.pelephone.net.il")
		self.NBIOT_ApnUser = tk.StringVar(value="pcl@3g")
		self.NBIOT_ApnPassword = tk.StringVar(value="rl")

		self.NBIOT_SSL_Enable = tk.BooleanVar(value=False)
		self.NBIOT_SSL_IgnoreLocalTime = tk.BooleanVar(value=True)
		self.NBIOT_SSL_NegotiateTimeoutSec = tk.IntVar(value=300)  # 10 to 300 seconds

		# POE variables
		self.POE_StaticIpAddr = tk.StringVar(value="10.0.0.110")
		self.POE_MaskAddress = tk.StringVar(value="255.255.255.0")
		self.POE_DnsServerAddr = tk.StringVar(value="10.0.0.138")
		self.POE_Gateway = tk.StringVar(value="10.0.0.138")
		self.POE_DHCP = tk.BooleanVar(value=False)
		self.POE_IPv6 = tk.BooleanVar(value=False)
		self.POE_IPv6LocalLinkOnly = tk.BooleanVar(value=False)

		self.MemMap = MemMap.cMemMap()
		self.address = 0    # next address to read/program/verify
		self.state = self.cSTATE_INIT
		self.after_id = None

		frame_Production = tk.Frame(self.parent, bd=0)
		frame_Production.grid(row=0, column=0, sticky="w")
		if True:
			self.label_PartNumber = tk.Label(frame_Production, justify="left", text="Part number:", font=("Verdana", 10), bd=0)
			self.label_PartNumber.grid(row=1, column=0, ipadx=10, ipady=2, pady=(5,0), sticky="w")
			self.combobox_PartNumber = ttk.Combobox(frame_Production, textvariable=self.PartNumber, values=MemMap.list_iCOMOX_PartNumber, width=75, font=("Verdana", 8), state="readonly")
			self.combobox_PartNumber.current(iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT)
			self.combobox_PartNumber.grid(row=1, column=1, pady=(5,0), sticky="w")
			# self.entry_PartNumber = tk.Entry(self.parent, textvariable=self.PartNumber, width=75)
			# self.entry_PartNumber.grid(row=1, column=1, sticky="w")

			self.label_SerialNumber = tk.Label(frame_Production, justify="left", text="Serial number:", font=("Verdana", 10), bd=0)
			self.label_SerialNumber.grid(row=2, column=0, ipadx=10, ipady=2, sticky="w")
			self.entry_SerialNumber = tk.Entry(frame_Production, textvariable=self.SerialNumber, width=75)
			self.entry_SerialNumber.grid(row=2, column=1, sticky="w")

			self.label_BoardType = tk.Label(frame_Production, justify="left", text="Board type:", font=("Verdana", 10), bd=0)
			self.label_BoardType.grid(row=3, column=0, ipadx=10, ipady=2, sticky="w")
			self.combobox_BoardType = ttk.Combobox(frame_Production, values=["SMIP", "NBIOT", "POE"], width=10, font=("Verdana", 8), state="readonly")
			self.combobox_BoardType.current(iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT)
			self.combobox_BoardType.grid(row=3, column=1, sticky="w")

			self.label_BoardVersion = tk.Label(frame_Production, justify="left", text="Board version:", font=("Verdana", 10), bd=0)
			self.label_BoardVersion.grid(row=4, column=0, ipadx=10, ipady=2, sticky="w")
			self.entry_BoardVersionMajor = tk.Entry(frame_Production, textvariable=self.BoardVersionMajor, width=5)
			self.entry_BoardVersionMajor.grid(row=4, column=1, sticky="w")
			self.label_BoardVersionPeriod = tk.Label(frame_Production, justify="left", text=".", font=("Verdana", 10), bd=0)
			self.label_BoardVersionPeriod.grid(row=4, column=1, padx=40, sticky="w")
			self.entry_BoardVersionMinor = tk.Entry(frame_Production, textvariable=self.BoardVersionMinor, width=5)
			self.entry_BoardVersionMinor.grid(row=4, column=1, padx=50, sticky="w")

			self.label_Name = tk.Label(frame_Production, justify="left", text="Name:", font=("Verdana", 10), bd=0)
			self.label_Name.grid(row=5, column=0, pady=(10,0), ipadx=10, ipady=2, sticky="w")
			self.entry_Name = tk.Entry(frame_Production, textvariable=self.Name, width=75)
			self.entry_Name.grid(row=5, column=1, pady=(10,0), sticky="w")

		self.notebook = ttk.Notebook(self.parent)
		self.notebook.grid(row=6, column=0, columnspan=2, padx=0, sticky="we")
		self.tab_smip = tk.Frame(self.notebook)
		self.tab_nbiot = tk.Frame(self.notebook)
		self.tab_poe = tk.Frame(self.notebook)
		self.notebook.add(self.tab_smip, text="SMIP")
		self.notebook.add(self.tab_nbiot, text="NB-IOT")
		self.notebook.add(self.tab_poe, text="POE")
		self.notebook.select(tab_id=self.tab_nbiot)

		self.frame_NBIOT_DefSimConfig = tk.Frame(self.tab_nbiot, bd=1, relief=tk.GROOVE)
		self.frame_NBIOT_DefSimConfig.grid(row=0, column=0, sticky="we", padx=0)
		if True:
			self.checkbox_NBIOT_DefSimConfig_Enable = tk.Checkbutton(self.frame_NBIOT_DefSimConfig, text="Use default SIM configuration", variable=self.NBIOT_DefaultSIMConfig, command=self.on_NBIOT_UseDefSimConfig, font=("Verdana", 10), state="normal")
			self.checkbox_NBIOT_DefSimConfig_Enable.grid(row=0, column=0, padx=2, ipadx=0, ipady=0, sticky="w")

			self.frame_NBIOT_DefSimConfig_AccessTechnology = tk.Frame(self.frame_NBIOT_DefSimConfig)
			self.frame_NBIOT_DefSimConfig_AccessTechnology.grid(row=1, column=0, columnspan=2, padx=10, ipady=0, sticky="we")
			if True:
				self.label_NBIOT_DefSimConfig_AccessTechnology = tk.Label(self.frame_NBIOT_DefSimConfig_AccessTechnology, justify="left", text="Access technologies:", font=("Verdana", 10), bd=0, state="disabled")
				self.label_NBIOT_DefSimConfig_AccessTechnology.grid(row=0, column=0, ipadx=0, ipady=0, sticky="w")

				self.checkbox_NBIOT_DefSimConfig_GSM_Enable = tk.Checkbutton(self.frame_NBIOT_DefSimConfig_AccessTechnology, text="GSM", variable=self.NBIOT_BitmaskAcT[0], font=("Verdana", 10), state="disabled")
				self.checkbox_NBIOT_DefSimConfig_GSM_Enable.grid(row=0, column=1, ipadx=0, ipady=0, sticky="w")

				self.checkbox_NBIOT_DefSimConfig_CatM1_Enable = tk.Checkbutton(self.frame_NBIOT_DefSimConfig_AccessTechnology, text="LTE cat. M1", variable=self.NBIOT_BitmaskAcT[1], font=("Verdana", 10), state="disabled")
				self.checkbox_NBIOT_DefSimConfig_CatM1_Enable.grid(row=0, column=2, ipadx=0, ipady=0, sticky="w")

				self.checkbox_NBIOT_DefSimConfig_CatNB1_Enable = tk.Checkbutton(self.frame_NBIOT_DefSimConfig_AccessTechnology, text="LTE cat. NB1", variable=self.NBIOT_BitmaskAcT[2], font=("Verdana", 10), state="disabled")
				self.checkbox_NBIOT_DefSimConfig_CatNB1_Enable.grid(row=0, column=3, ipadx=0, ipady=0, sticky="w")

			frame_NBIOT_DefSimConfig_Params = tk.Frame(self.frame_NBIOT_DefSimConfig)
			frame_NBIOT_DefSimConfig_Params.grid(row=2, column=0, padx=10, ipady=0, sticky="we")
			self.entry_NBIOT_ScanBands = [None]*3
			self.label_NBIOT_DefSimConfig_ScanBands = [None]*3
			self.button_NBIOT_ScanBands = [None]*3
			AcT_names = ["GSM", "LTE cat. M1", "LTE cat. NB1"]
			button_commands = [(lambda : self.on_button_NBIOT_SetBand(AcT=0)), (lambda : self.on_button_NBIOT_SetBand(AcT=1)), (lambda : self.on_button_NBIOT_SetBand(AcT=2))]
			for AcT in range(0, 3):
				if True:
					self.label_NBIOT_DefSimConfig_ScanBands[AcT] = tk.Label(frame_NBIOT_DefSimConfig_Params, justify="left", text="{} bands:".format(AcT_names[AcT]), font=("Verdana", 10), bd=0, state="disabled")
					self.label_NBIOT_DefSimConfig_ScanBands[AcT].grid(row=AcT, column=0, ipadx=0, ipady=0, sticky="w")
					self.entry_NBIOT_ScanBands[AcT] = tk.Entry(frame_NBIOT_DefSimConfig_Params, textvariable=self.NBIOT_ScanBands[AcT], width=6, state="disabled")
					self.entry_NBIOT_ScanBands[AcT].grid(row=AcT, column=1, sticky="w")
					self.button_NBIOT_ScanBands[AcT] = tk.Button(frame_NBIOT_DefSimConfig_Params, text="...", command=button_commands[AcT], state="disabled")
					self.button_NBIOT_ScanBands[AcT].grid(row=AcT, column=2, padx=(1,0), sticky="w")

		frame_NBIOT_DefSimConfig_ScanOrder = tk.Frame(self.frame_NBIOT_DefSimConfig)
		frame_NBIOT_DefSimConfig_ScanOrder.grid(row=3, column=0, padx=10, ipady=2, sticky="we")
		if True:
			self.label_NBIOT_DefSimConfig_ScanOrder = tk.Label(frame_NBIOT_DefSimConfig_ScanOrder, justify="left", text="Scan order:", font=("Verdana", 10), bd=0, state="disabled")
			self.label_NBIOT_DefSimConfig_ScanOrder.grid(row=5, column=0, ipadx=0, ipady=0, sticky="w")
			self.combobox_NBIOT_DefSimConfig_ScanOrder = ttk.Combobox(frame_NBIOT_DefSimConfig_ScanOrder, width=35, values=["LTE cat. M1 -> LTE cat. NB1 -> GSM", "LTE cat. M1 -> GSM -> LTE cat. NB1", "GSM -> LTE cat. M1 -> LTE cat. NB1", "LTE cat. NB1 -> LTE cat. M1 -> GSM", "LTE cat. NB1 -> GSM -> LTE cat. M1", "GSM -> LTE cat. NB1 -> LTE cat. M1"], font=("Verdana", 8), state="disabled")
			self.combobox_NBIOT_DefSimConfig_ScanOrder.grid(row=5, column=1, sticky="w")
			self.combobox_NBIOT_DefSimConfig_ScanOrder.current(0)

		frame_NBIOT_DefSimConfig_ServiceDomain = tk.Frame(self.frame_NBIOT_DefSimConfig)
		frame_NBIOT_DefSimConfig_ServiceDomain.grid(row=4, column=0, padx=10, ipady=2, sticky="we")
		if True:
			self.label_NBIOT_DefSimConfig_ServiceDomain = tk.Label(frame_NBIOT_DefSimConfig_ServiceDomain, justify="left", text="Service domain:", font=("Verdana", 10), bd=0, state="disabled")
			self.label_NBIOT_DefSimConfig_ServiceDomain.grid(row=0, column=0, ipadx=0, ipady=0, sticky="w")
			self.combobox_NBIOT_DefSimConfig_ServiceDomain = ttk.Combobox(frame_NBIOT_DefSimConfig_ServiceDomain, width=11, values=["PS only", "CS & PS"], font=("Verdana", 8), state="disabled")
			self.combobox_NBIOT_DefSimConfig_ServiceDomain.grid(row=0, column=1, sticky="w")
			self.combobox_NBIOT_DefSimConfig_ServiceDomain.current(1)

		self.frame_NBIOT_ManualOperatorSelection = tk.Frame(self.tab_nbiot, bd=1, relief=tk.GROOVE)
		self.frame_NBIOT_ManualOperatorSelection.grid(row=2, column=0, sticky="we", ipadx=0, ipady=2)
		# self.frame_NBIOT_ManualOperatorSelection.grid_forget()	# Remove in 2.8.1
		if True:
			self.checkbox_NBIOT_ManualOperatorSelection_Enable = tk.Checkbutton(self.frame_NBIOT_ManualOperatorSelection, text="Manual operator selection", variable=self.NBIOT_ManualOperatorSelection, command=self.on_NBIOT_ManualOperatorSelection, font=("Verdana", 10), state="normal")
			self.checkbox_NBIOT_ManualOperatorSelection_Enable.grid(row=0, column=0, ipadx=2, ipady=0, sticky="w")
			# self.checkbox_NBIOT_ManualOperatorSelection_Enable.configure(state="disabled")	# Remove in 2.8.1

			self.frame_NBIOT_ManualOperatorSelection_PLMN = tk.Frame(self.frame_NBIOT_ManualOperatorSelection)
			# self.frame_NBIOT_ManualOperatorSelection_PLMN.grid(row=1, column=0, sticky="we", ipadx=0, ipady=0)
			self.frame_NBIOT_ManualOperatorSelection_PLMN.grid(row=0, column=1, sticky="we", ipadx=0, ipady=0)
			if True:
				self.label_NBIOT_ManualOperatorSelection_PLMN = tk.Label(self.frame_NBIOT_ManualOperatorSelection_PLMN, justify="left", text="PLMN:", font=("Verdana", 10), bd=0, state="disabled")
				self.label_NBIOT_ManualOperatorSelection_PLMN.grid(row=1, column=0, ipadx=10, ipady=0, sticky="w")
				self.entry_NBIOT_ManualOperatorSelection_PLMN = tk.Entry(self.frame_NBIOT_ManualOperatorSelection_PLMN, textvariable=self.NBIOT_PLMN, width=10, state="disabled")
				self.entry_NBIOT_ManualOperatorSelection_PLMN.grid(row=1, column=1, sticky="w")

		self.frame_NBIOT_Context = tk.Frame(self.tab_nbiot, bd=1, relief=tk.GROOVE)
		self.frame_NBIOT_Context.grid(row=3, column=0, sticky="we", padx=0, ipadx=0, ipady=2)
		if True:
			self.checkbox_NBIOT_EnableRoaming = tk.Checkbutton(self.frame_NBIOT_Context, width=0, bd=0, text="Enable roaming", variable=self.NBIOT_EnableRoaming, font=("Verdana", 10), state="normal")
			self.checkbox_NBIOT_EnableRoaming.grid(row=0, column=0, padx=2, ipadx=0, ipady=0, sticky="w")

			self.label_NBIOT_ContextType = tk.Label(self.frame_NBIOT_Context, justify="left", text="Context type:", font=("Verdana", 10), bd=0)
			self.label_NBIOT_ContextType.grid(row=2, column=0, padx=2, ipadx=0, ipady=2, sticky="w")
			self.combobox_NBIOT_ContextType = ttk.Combobox(self.frame_NBIOT_Context, width=11, values=["IPv4", "IPv6", "IPv4 & IPv6"], font=("Verdana", 8), state="readonly")
			self.combobox_NBIOT_ContextType.grid(row=2, column=1, sticky="w")
			self.combobox_NBIOT_ContextType.current(0)

			label_NBIOT_Authentication = tk.Label(self.frame_NBIOT_Context, justify="left", text="Authentication:", font=("Verdana", 10), bd=0)
			label_NBIOT_Authentication.grid(row=3, column=0, padx=2, ipadx=0, ipady=2, sticky="w")
			self.combobox_NBIOT_Authentication = ttk.Combobox(self.frame_NBIOT_Context, width=11, values=["None", "PAP", "CHAP", "PAP & CHAP"], font=("Verdana", 8), state="readonly")
			self.combobox_NBIOT_Authentication.grid(row=3, column=1, sticky="w")
			self.combobox_NBIOT_Authentication.current(0)

			label_NBIOT_ApnAccessName = tk.Label(self.frame_NBIOT_Context, justify="left", text="APN Access Name:", font=("Verdana", 10), bd=0)
			label_NBIOT_ApnAccessName.grid(row=4, column=0, padx=2, ipadx=0, ipady=2, sticky="w")
			self.entry_NBIOT_ApnAccessName = tk.Entry(self.frame_NBIOT_Context, textvariable=self.NBIOT_ApnAccessName, width=45)
			self.entry_NBIOT_ApnAccessName.grid(row=4, column=1, sticky="w")

			label_NBIOT_ApnUser = tk.Label(self.frame_NBIOT_Context, justify="left", text="APN User:", font=("Verdana", 10), bd=0)
			label_NBIOT_ApnUser.grid(row=5, column=0, padx=2, ipadx=0, ipady=2, sticky="w")
			self.entry_NBIOT_ApnUser = tk.Entry(self.frame_NBIOT_Context, textvariable=self.NBIOT_ApnUser, width=45)
			self.entry_NBIOT_ApnUser.grid(row=5, column=1, sticky="w")

			label_NBIOT_ApnPassword = tk.Label(self.frame_NBIOT_Context, justify="left", text="APN Password:", font=("Verdana", 10), bd=0)
			label_NBIOT_ApnPassword.grid(row=6, column=0, padx=2, ipadx=0, ipady=2, sticky="w")
			self.entry_NBIOT_ApnPassword = tk.Entry(self.frame_NBIOT_Context, textvariable=self.NBIOT_ApnPassword, width=45)
			self.entry_NBIOT_ApnPassword.grid(row=6, column=1, sticky="w")

		self.frame_NBIOT_SSL = tk.Frame(self.tab_nbiot, bd=1, relief=tk.GROOVE)
		self.frame_NBIOT_SSL.grid(row=0, column=1, rowspan=4, sticky="wesn", ipady=2)
		# self.frame_NBIOT_SSL.grid_forget()	# Remove in 2.8.1
		if True:
			self.checkbox_NBIOT_SSL_Enable = tk.Checkbutton(self.frame_NBIOT_SSL, text="Use SSL", variable=self.NBIOT_SSL_Enable, command=self.on_NBIOT_SSL_Enable, font=("Verdana", 10), state="normal")
			self.checkbox_NBIOT_SSL_Enable.grid(row=0, column=0, padx=0, ipadx=2, ipady=2, sticky="w")
			self.checkbox_NBIOT_SSL_Enable.configure(state="disabled")	# Remove in 2.8.1

			self.frame_NBIOT_SSL_Params = tk.Frame(self.frame_NBIOT_SSL)
			self.frame_NBIOT_SSL_Params.grid(row=1, column=0, columnspan=2, padx=10, ipady=0, sticky="we")
			if True:
				self.label_NBIOT_SSL_Version = tk.Label(self.frame_NBIOT_SSL_Params, justify="left", text="Version:", font=("Verdana", 10), bd=0, state="disabled")
				self.label_NBIOT_SSL_Version.grid(row=0, column=0, ipadx=0, ipady=0, sticky="w")
				self.combobox_NBIOT_SSL_Version = ttk.Combobox(self.frame_NBIOT_SSL_Params, values=["SSL 3.0", "TLS 1.0", "TLS 1.1", "TLS 1.2", "All"], width=7, font=("Verdana", 8), state="disabled")
				self.combobox_NBIOT_SSL_Version.grid(row=0, column=1, pady=2, sticky="w")
				self.combobox_NBIOT_SSL_Version.current(3) # TLS 1.2

				self.label_NBIOT_SSL_SecurityLevel = tk.Label(self.frame_NBIOT_SSL_Params, justify="left", text="Security level:", font=("Verdana", 10), bd=0, state="disabled")
				self.label_NBIOT_SSL_SecurityLevel.grid(row=1, column=0, ipadx=0, ipady=0, sticky="w")
				self.combobox_NBIOT_SSL_SecurityLevel = ttk.Combobox(self.frame_NBIOT_SSL_Params, values=["None", "Server authentication", "Server and client authentication if remotely requested"], width=40, font=("Verdana", 8), state="disabled")
				self.combobox_NBIOT_SSL_SecurityLevel.grid(row=1, column=1, pady=2, sticky="w")
				self.combobox_NBIOT_SSL_SecurityLevel.current(0) # None

				self.label_NBIOT_SSL_CipherSuite = tk.Label(self.frame_NBIOT_SSL_Params, justify="left", text="Cipher suite:", font=("Verdana", 10), bd=0, state="disabled")
				self.label_NBIOT_SSL_CipherSuite.grid(row=2, column=0, ipadx=0, ipady=0, sticky="w")
				self.combobox_NBIOT_SSL_CipherSuite = ttk.Combobox(self.frame_NBIOT_SSL_Params, values=[
					"TLS_RSA_WITH_AES_256_CBC_SHA",
					"TLS_RSA_WITH_AES_128_CBC_SHA",
					"TLS_RSA_WITH_RC4_128_SHA",
					"TLS_RSA_WITH_RC4_128_MD5",
					"TLS_RSA_WITH_3DES_EDE_CBC_SHA",
					"TLS_RSA_WITH_AES_256_CBC_SHA256",
					"TLS_ECDHE_RSA_WITH_RC4_128_SHA",
					"TLS_ECDHE_RSA_WITH_3DES_EDE_CBC_SHA",
					"TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA",
					"TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA",
					"TLS_ECDHE_RSA_WITH_AES_128_CBC_SHA256",
					"TLS_ECDHE_RSA_WITH_AES_256_CBC_SHA384",
					"TLS_ECDHE_RSA_WITH_AES_128_GCM_SHA256",
					"Support all"], width=40, font=("Verdana", 8), state="disabled")
				self.combobox_NBIOT_SSL_CipherSuite.grid(row=2, column=1, pady=2, sticky="w")
				self.combobox_NBIOT_SSL_CipherSuite.current(0) # RSA + AES256 + CBC + SHA

				self.label_NBIOT_SSL_NegotiateTimeout = tk.Label(self.frame_NBIOT_SSL_Params, justify="left", text="Negotiate timeout [sec]:", font=("Verdana", 10), bd=0, state="disabled")
				self.label_NBIOT_SSL_NegotiateTimeout.grid(row=3, column=0, ipadx=0, ipady=0, sticky="w")
				self.spinbox_NBIOT_SSL_NegotiateTimeout = tk.Spinbox(self.frame_NBIOT_SSL_Params, from_=10, increment=1, to=300, wrap=True, width=3, state="disabled", textvariable=self.NBIOT_SSL_NegotiateTimeoutSec)
				self.spinbox_NBIOT_SSL_NegotiateTimeout.grid(row=3, column=1, ipadx=0, ipady=0, sticky="w")

				# self.label_NBIOT_SSL_CACertFile = tk.Label(self.frame_NBIOT_SSL_Params, justify="left", text="CA certificate file:", font=("Verdana", 10), bd=0, state="disabled")
				# self.label_NBIOT_SSL_CACertFile.grid(row=4, column=0, ipadx=0, ipady=2, sticky="w")
				# self.entry_NBIOT_SSL_CACertFile = tk.Entry(self.frame_NBIOT_SSL_Params, textvariable=self.NBIOT_SSL_CACertFile, width=45, state="disabled")
				# self.entry_NBIOT_SSL_CACertFile.grid(row=4, column=1, sticky="w")
				#
				# self.label_NBIOT_SSL_ClientCertFile = tk.Label(self.frame_NBIOT_SSL_Params, justify="left", text="Client certificate file:", font=("Verdana", 10), bd=0, state="disabled")
				# self.label_NBIOT_SSL_ClientCertFile.grid(row=5, column=0, ipadx=0, ipady=2, sticky="w")
				# self.entry_NBIOT_SSL_ClientCertFile = tk.Entry(self.frame_NBIOT_SSL_Params, textvariable=self.NBIOT_SSL_ClientCertFile, width=45, state="disabled")
				# self.entry_NBIOT_SSL_ClientCertFile.grid(row=5, column=1, sticky="w")
				#
				# self.label_NBIOT_SSL_ClientKeyFile = tk.Label(self.frame_NBIOT_SSL_Params, justify="left", text="Client key file:", font=("Verdana", 10), bd=0, state="disabled")
				# self.label_NBIOT_SSL_ClientKeyFile.grid(row=6, column=0, ipadx=0, ipady=2, sticky="w")
				# self.entry_NBIOT_SSL_ClientKeyFile = tk.Entry(self.frame_NBIOT_SSL_Params, textvariable=self.NBIOT_SSL_ClientKeyFile, width=45, state="disabled")
				# self.entry_NBIOT_SSL_ClientKeyFile.grid(row=6, column=1, sticky="w")

				self.checkbox_NBIOT_SSL_IgnoreLocalTime = tk.Checkbutton(self.frame_NBIOT_SSL_Params, text="Ignore local time", variable=self.NBIOT_SSL_IgnoreLocalTime, font=("Verdana", 10), state="disabled")
				self.checkbox_NBIOT_SSL_IgnoreLocalTime.grid(row=7, column=0, ipadx=0, ipady=2, sticky="w")

		self.frame_NBIOT_backbone = self.create_backbone(parent_tab=self.tab_nbiot)

		# POE tab
		self.frame_POE_interface = tk.Frame(self.tab_poe, bd=1, relief=tk.GROOVE)
		self.frame_POE_interface.grid(row=0, column=0, rowspan=4, sticky="nswe", ipady=2, ipadx=2, padx=5)
		if True:
			self.label_POE_StaticIPAddr = tk.Label(self.frame_POE_interface, justify="left", text="Client static IP address:", font=("Verdana", 10), bd=0)
			self.label_POE_StaticIPAddr.grid(row=1, column=0, ipadx=10, ipady=2, sticky="w")
			self.entry_POE_StaticIpAddr = tk.Entry(self.frame_POE_interface, textvariable=self.POE_StaticIpAddr, width=45)
			self.entry_POE_StaticIpAddr.grid(row=1, column=1, sticky="w")

			self.label_POE_mask = tk.Label(self.frame_POE_interface, justify="left", text="Mask address:", font=("Verdana", 10), bd=0)
			self.label_POE_mask.grid(row=2, column=0, ipadx=10, ipady=2, sticky="w")
			self.entry_POE_mask = tk.Entry(self.frame_POE_interface, textvariable=self.POE_MaskAddress, width=45)
			self.entry_POE_mask.grid(row=2, column=1, sticky="w")

			self.label_POE_DnsServerAddr = tk.Label(self.frame_POE_interface, justify="left", text="DNS server address:", font=("Verdana", 10), bd=0)
			self.label_POE_DnsServerAddr.grid(row=3, column=0, ipadx=10, ipady=2, sticky="w")
			self.entry_POE_DnsServerAddr = tk.Entry(self.frame_POE_interface, textvariable=self.POE_DnsServerAddr, width=45)
			self.entry_POE_DnsServerAddr.grid(row=3, column=1, sticky="w")

			self.label_POE_gateway = tk.Label(self.frame_POE_interface, justify="left", text="Default gateway:", font=("Verdana", 10), bd=0)
			self.label_POE_gateway.grid(row=4, column=0, ipadx=10, ipady=2, sticky="w")
			self.entry_POE_gateway = tk.Entry(self.frame_POE_interface, textvariable=self.POE_Gateway, width=45)
			self.entry_POE_gateway.grid(row=4, column=1, sticky="w")

			self.frame_POE_flags = tk.Frame(self.frame_POE_interface, bd=0, relief=tk.GROOVE)
			self.frame_POE_flags.grid(row=5, column=0, rowspan=1, columnspan=2, sticky="nswe", ipady=2, ipadx=2)
			if True:
				self.checkbox_POE_DHCP = tk.Checkbutton(self.frame_POE_flags, text="DHCP", variable=self.POE_DHCP, command=self.on_POE_DHCP, font=("Verdana", 8), state="normal")
				self.checkbox_POE_DHCP.grid(row=0, column=0, padx=5, sticky="w")
				self.checkbox_POE_IPv6 = tk.Checkbutton(self.frame_POE_flags, text="IPv6", variable=self.POE_IPv6, font=("Verdana", 8), state="normal")
				self.checkbox_POE_IPv6.grid(row=0, column=1, padx=5, sticky="w")
				self.checkbox_POE_IPv6.grid_forget()	# Remove in 2.8.1
				self.checkbox_POE_IPv6LocalLinkOnly = tk.Checkbutton(self.frame_POE_flags, text="IPv6 local link only", variable=self.POE_IPv6LocalLinkOnly, font=("Verdana", 8), state="normal")
				self.checkbox_POE_IPv6LocalLinkOnly.grid(row=0, column=2, padx=5, sticky="w")
				self.checkbox_POE_IPv6LocalLinkOnly.grid_forget()	# Remove in 2.8.1
				# self.checkbox_POE_TLS = tk.Checkbutton(self.frame_POE_interface, text="TLS", variable=self.POE_TLS, font=("Verdana", 8), state="normal")
				# self.checkbox_POE_TLS.grid(row=7, column=1, padx=325, sticky="w")

		self.frame_POE_backbone = self.create_backbone(parent_tab=self.tab_poe)

		# Read/Program/Verify buttons
		self.frame = tk.Frame(self.parent)
		self.frame.grid(row=20, column=0, rowspan=1, columnspan=4)

		self.button_ReadEEPROM = tk.Button(self.frame, command=self.ReadEEPROM, width=20, text="Read from EEPROM", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
		self.button_ReadEEPROM.grid(row = 0, column=0, padx=10)

		self.button_ProgramEEPROM = tk.Button(self.frame, command=self.ProgramEEPROM, width=20, text="Program EEPROM", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
		self.button_ProgramEEPROM.grid(row = 0, column=1, padx=10)

		self.button_VerifyEEPROM = tk.Button(self.frame, command=self.VerifyEEPROM, width=20, text="Verify EEPROM", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
		self.button_VerifyEEPROM.grid(row = 0, column=2, padx=10)

		if common_symbols.__PRODUCTION_SUPPORT__:
			self.button_Default = tk.Button(self.frame, command=self.Default, width=25, text="Default", font=("Verdana", 8), bg="gray80", fg="black", state="normal")
			self.button_Default.grid(row = 0, column=4, padx=10)

		self.progressbar = ttk.Progressbar(self.parent, style="green.Horizontal.TProgressbar", orient="horizontal", mode="determinate", value=0, maximum=MemMap.cMEMMAP_SIZE)
		self.progressbar.grid(row=21, column=0, rowspan=1, columnspan=4, padx=5, pady=10, sticky="we")

		if not common_symbols.__PRODUCTION_SUPPORT__:
			# self.label_PartNumber.grid_remove()
			self.combobox_PartNumber.configure(state="disabled")
			# self.label_SerialNumber.grid_remove()
			self.entry_SerialNumber.configure(state="disabled")
			# self.label_BoardType.grid_remove()
			self.combobox_BoardType.configure(state="disabled")
			# self.label_BoardVersion.grid_remove()
			# self.label_BoardVersionPeriod.grid_remove()
			self.entry_BoardVersionMajor.configure(state="disabled")
			self.entry_BoardVersionMinor.configure(state="disabled")
		else:
			self.entry_SerialNumber.focus_set()

	def create_backbone(self, parent_tab):
		backbone = \
		{ \
			"TCP_ServerURL": tk.StringVar(value="31.154.35.78"), \
			"TCP_ServerPort": tk.IntVar(value=1526), \
			"TCP_delayBeforeReconnectingSec": tk.IntVar(value=5), \

			"MQTT_ServerURL": tk.StringVar(value=""), \
			"MQTT_UserName": tk.StringVar(value=""), \
			"MQTT_Password": tk.StringVar(value=""), \
			"MQTT_ClientID": tk.StringVar(value=""), \
			"MQTT_ServerPort": tk.IntVar(value=8000), \
			"MQTT_Flags_ProvideUserName": tk.BooleanVar(value=True), \
			"MQTT_Flags_ProvidePassword": tk.BooleanVar(value=True), \
			"MQTT_Keepalive": tk.IntVar(value=400), \

			"frame_backbone": tk.Frame(parent_tab, bd=1, relief=tk.GROOVE) \
		}

		backbone["frame_backbone"].grid(row=0, column=2, rowspan=4, sticky="wesn", ipady=2)
		if True:
			backbone["frame_backbone_protocol"] = tk.Frame(backbone["frame_backbone"])
			backbone["frame_backbone_protocol"].grid(row=0, column=0, sticky="w", ipady=0)
			if True:
				backbone["label_backbone"] = tk.Label(backbone["frame_backbone_protocol"], justify="left", text="Backbone protocol: ", font=("Verdana", 10), bd=0)
				backbone["label_backbone"].grid(row=0, column=0, padx=(2,0), ipady=2, sticky="w")
				backbone["combobox_backbone"] = ttk.Combobox(backbone["frame_backbone_protocol"], values=["TCP"], width=10, font=("Verdana", 8), state="readonly")
				backbone["combobox_backbone"].grid(row=0, column=1, pady=2, sticky="w")
				backbone["combobox_backbone"].current(0) # TCP
				backbone["label_backbone"].configure(state="disabled")	# Remove in 2.8.1
				# self.combobox_NBIOT_backbone.configure(state="disabled")	# Remove in 2.8.1

			backbone["notebook_backbone"] = ttk.Notebook(backbone["frame_backbone"])
			backbone["notebook_backbone"].grid(row=1, column=0, columnspan=2, sticky="ns")
			backbone["tab_backbone_tcp"] = tk.Frame(backbone["notebook_backbone"])
			backbone["tab_backbone_mqtt"] = tk.Frame(backbone["notebook_backbone"])
			backbone["notebook_backbone"].add(backbone["tab_backbone_tcp"], text="TCP")
			backbone["notebook_backbone"].add(backbone["tab_backbone_mqtt"], text="MQTT")
			backbone["tab_backbone_arr"] = [backbone["tab_backbone_tcp"], backbone["tab_backbone_mqtt"]]
			backbone["notebook_backbone"].tab(backbone["notebook_backbone"].tabs()[1], state="disabled")	# Remove in 2.8.1

			if True:
				backbone["label_TCP_ServerURL"] = tk.Label(backbone["tab_backbone_tcp"], justify="left", text="Server URL:", font=("Verdana", 10), bd=0)
				backbone["label_TCP_ServerURL"].grid(row=0, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_TCP_ServerURL"] = tk.Entry(backbone["tab_backbone_tcp"], textvariable=backbone["TCP_ServerURL"], width=45)
				backbone["entry_TCP_ServerURL"].grid(row=0, column=1, sticky="w")
				# backbone["checkbox_TCP_IPv6"] = tk.Checkbutton(backbone["tab_backbone_tcp"], text="IPv6", variable=backbone["TCP_IPv6"], font=("Verdana", 10), state="normal")
				# backbone["checkbox_TCP_IPv6"].grid(row=0, column=2, ipadx=0, ipady=2, sticky="w")

				backbone["label_TCP_ServerPort"] = tk.Label(backbone["tab_backbone_tcp"], justify="left", text="Server port:", font=("Verdana", 10), bd=0)
				backbone["label_TCP_ServerPort"].grid(row=1, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_TCP_ServerPort"] = tk.Entry(backbone["tab_backbone_tcp"], textvariable=backbone["TCP_ServerPort"], width=5)
				backbone["entry_TCP_ServerPort"].grid(row=1, column=1, sticky="w")

				backbone["label_TCP_DelayBeforeReconnecting"] = tk.Label(backbone["tab_backbone_tcp"], justify="left", text="Delay before reconnecting [sec]:", font=("Verdana", 10), bd=0)
				backbone["label_TCP_DelayBeforeReconnecting"].grid(row=9, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_TCP_DelayBeforeReconnecting"] = tk.Entry(backbone["tab_backbone_tcp"], textvariable=backbone["TCP_delayBeforeReconnectingSec"], width=5)
				backbone["entry_TCP_DelayBeforeReconnecting"].grid(row=9, column=1, sticky="w")

			if True:
				backbone["label_MQTT_ServerURL"] = tk.Label(backbone["tab_backbone_mqtt"], justify="left", text="Server URL:", font=("Verdana", 10), bd=0)
				backbone["label_MQTT_ServerURL"].grid(row=0, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_MQTT_ServerURL"] = tk.Entry(backbone["tab_backbone_mqtt"], textvariable=backbone["MQTT_ServerURL"], width=67)
				backbone["entry_MQTT_ServerURL"].grid(row=0, column=1, columnspan=2, sticky="w")

				backbone["label_MQTT_UserName"] = tk.Label(backbone["tab_backbone_mqtt"], justify="left", text="User name:", font=("Verdana", 10), bd=0)
				backbone["label_MQTT_UserName"].grid(row=1, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_MQTT_UserName"] = tk.Entry(backbone["tab_backbone_mqtt"], textvariable=backbone["MQTT_UserName"], width=32)
				backbone["entry_MQTT_UserName"].grid(row=1, column=1, sticky="w")
				backbone["checkbox_MQTT_ProvideUserName"] = tk.Checkbutton(backbone["tab_backbone_mqtt"], text="Provide user name on CONNECT", variable=backbone["MQTT_Flags_ProvideUserName"], font=("Verdana", 8), state="normal")
				backbone["checkbox_MQTT_ProvideUserName"].grid(row=1, column=2, ipadx=2, ipady=0, padx=0, pady=0, sticky="w")

				backbone["label_MQTT_Password"] = tk.Label(backbone["tab_backbone_mqtt"], justify="left", text="Password:", font=("Verdana", 10), bd=0)
				backbone["label_MQTT_Password"].grid(row=2, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_MQTT_Password"] = tk.Entry(backbone["tab_backbone_mqtt"], textvariable=backbone["MQTT_Password"], width=32)
				backbone["entry_MQTT_Password"].grid(row=2, column=1, sticky="w")
				backbone["checkbox_MQTT_ProvidePassword"] = tk.Checkbutton(backbone["tab_backbone_mqtt"], text="Provide password on CONNECT", variable=backbone["MQTT_Flags_ProvidePassword"], font=("Verdana", 8), state="normal")
				backbone["checkbox_MQTT_ProvidePassword"].grid(row=2, column=2, ipadx=2, ipady=0, padx=0, pady=0, sticky="w")

				backbone["label_MQTT_ClientID"] = tk.Label(backbone["tab_backbone_mqtt"], justify="left", text="Client ID:", font=("Verdana", 10), bd=0)
				backbone["label_MQTT_ClientID"].grid(row=3, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_MQTT_ClientID"] = tk.Entry(backbone["tab_backbone_mqtt"], textvariable=backbone["MQTT_ClientID"], width=32)
				backbone["entry_MQTT_ClientID"].grid(row=3, column=1, sticky="w")

				backbone["label_MQTT_ServerPort"] = tk.Label(backbone["tab_backbone_mqtt"], justify="left", text="Server port:", font=("Verdana", 10), bd=0)
				backbone["label_MQTT_ServerPort"].grid(row=4, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_MQTT_ServerPort"] = tk.Entry(backbone["tab_backbone_mqtt"], textvariable=backbone["MQTT_ServerPort"], width=5)
				backbone["entry_MQTT_ServerPort"].grid(row=4, column=1, sticky="w")

				backbone["label_MQTT_Keepalive"] = tk.Label(backbone["tab_backbone_mqtt"], justify="left", text="MNC:", font=("Verdana", 10), bd=0, state="disabled")
				backbone["label_MQTT_Keepalive"].grid(row=5, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["entry_MQTT_Keepalive"] = tk.Entry(backbone["tab_backbone_mqtt"], textvariable=backbone["MQTT_Keepalive"], width=10, state="disabled")
				backbone["entry_MQTT_Keepalive"].grid(row=5, column=1, sticky="w")

				backbone["label_MQTT_Cloud"] = tk.Label(backbone["tab_backbone_mqtt"], justify="left", text="Cloud:", font=("Verdana", 10), bd=0)
				backbone["label_MQTT_Cloud"].grid(row=6, column=0, ipadx=10, ipady=2, sticky="w")
				backbone["combobox_MQTT_Cloud"] = ttk.Combobox(backbone["tab_backbone_mqtt"], width=10, values=["Reserved", "T-Systems"], font=("Verdana", 8), state="readonly")
				backbone["combobox_MQTT_Cloud"].grid(row=6, column=1, sticky="w")
				backbone["combobox_MQTT_Cloud"].current(0)

				backbone["notebook_MQTT_clouds"] = ttk.Notebook(backbone["tab_backbone_mqtt"])
				backbone["notebook_MQTT_clouds"].grid(row=7, column=0, columnspan=2, sticky="wns")
				backbone["tab_mqtt_cloud_reserved"] = tk.Frame(backbone["notebook_MQTT_clouds"])
				backbone["tab_mqtt_cloud_T_Systems"] = tk.Frame(backbone["notebook_MQTT_clouds"])
				backbone["notebook_MQTT_clouds"].add(backbone["tab_mqtt_cloud_reserved"], text="Reserved")
				backbone["notebook_MQTT_clouds"].add(backbone["tab_mqtt_cloud_T_Systems"], text="T-Systems")
				backbone["tab_mqtt_cloud_arr"] = [backbone["tab_mqtt_cloud_reserved"], backbone["tab_mqtt_cloud_T_Systems"]]

		return backbone

	def on_NBIOT_UseDefSimConfig(self):
		state_DN = "disabled" if self.NBIOT_DefaultSIMConfig.get() else "normal"
		state_DR = "disabled" if self.NBIOT_DefaultSIMConfig.get() else "readonly"
		self.label_NBIOT_DefSimConfig_AccessTechnology.configure(state=state_DN)
		self.checkbox_NBIOT_DefSimConfig_GSM_Enable.configure(state=state_DN)
		self.checkbox_NBIOT_DefSimConfig_CatM1_Enable.configure(state=state_DN)
		self.checkbox_NBIOT_DefSimConfig_CatNB1_Enable.configure(state=state_DN)
		self.label_NBIOT_DefSimConfig_ServiceDomain.configure(state=state_DN)
		self.combobox_NBIOT_DefSimConfig_ServiceDomain.configure(state=state_DR)
		state_DN = "disabled"	# Remove in 2.8.1
		state_DR = "disabled"	# Remove in 2.8.1
		for AcT in range(0, 3):
			self.label_NBIOT_DefSimConfig_ScanBands[AcT].configure(state=state_DN)
			self.entry_NBIOT_ScanBands[AcT].configure(state=state_DR)
			self.button_NBIOT_ScanBands[AcT].configure(state=state_DN)
		self.label_NBIOT_DefSimConfig_ScanOrder.configure(state=state_DN)
		self.combobox_NBIOT_DefSimConfig_ScanOrder.configure(state=state_DR)

	def on_NBIOT_ManualOperatorSelection(self):
		# self.NBIOT_ManualOperatorSelection.set(False)	# Remove in 2.8.1
		state = "normal" if self.NBIOT_ManualOperatorSelection.get() else "disabled"
		self.label_NBIOT_ManualOperatorSelection_PLMN.configure(state=state)
		self.entry_NBIOT_ManualOperatorSelection_PLMN.configure(state=state)

	def on_NBIOT_SSL_Enable(self):
		self.NBIOT_SSL_Enable.set(False)	# REMOVE in 2.8.1
		state = "normal" if self.NBIOT_SSL_Enable.get() else "disabled"
		self.label_NBIOT_SSL_Version.configure(state=state)
		self.label_NBIOT_SSL_SecurityLevel.configure(state=state)
		self.label_NBIOT_SSL_CipherSuite.configure(state=state)
		self.label_NBIOT_SSL_NegotiateTimeout.configure(state=state)
		self.spinbox_NBIOT_SSL_NegotiateTimeout.configure(state=state)
		self.checkbox_NBIOT_SSL_IgnoreLocalTime.configure(state=state)

		state = "readonly" if self.NBIOT_SSL_Enable.get() else "disabled"
		self.combobox_NBIOT_SSL_Version.configure(state=state)
		self.combobox_NBIOT_SSL_Version.configure(state=state)
		self.combobox_NBIOT_SSL_SecurityLevel.configure(state=state)
		self.combobox_NBIOT_SSL_CipherSuite.configure(state=state)

	def on_button_NBIOT_SetBand(self, AcT):
		initial_value = int(self.NBIOT_ScanBands[AcT].get(), 16)
		if AcT == 0:
			Dialog = DialogBands.cDialog_GSM_bands(self, initial_value=initial_value)
		elif AcT == 1:
			Dialog = DialogBands.cDialog_LTE_cat_M1_bands(self, initial_value=initial_value)
		elif AcT == 2:
			Dialog = DialogBands.cDialog_LTE_cat_NB1_bands(self, initial_value=initial_value)
		else:
			raise Exception("on_button_NBIOT_SetBand(): Illegal AcT argument")
		if Dialog.result is None:
			return
		self.NBIOT_ScanBands[AcT].set("{:04X}".format(Dialog.result))

	def on_POE_DHCP(self):
		if self.POE_DHCP.get():
			state = "disabled"
		else:
			state = "normal"
		self.label_POE_StaticIPAddr.configure(state=state)
		self.entry_POE_StaticIpAddr.configure(state=state)
		self.label_POE_mask.configure(state=state)
		self.entry_POE_mask.configure(state=state)
		self.label_POE_DnsServerAddr.configure(state=state)
		self.entry_POE_DnsServerAddr.configure(state=state)
		self.label_POE_gateway.configure(state=state)
		self.entry_POE_gateway.configure(state=state)

	def SetDefaultFocus(self):
		if common_symbols.__PRODUCTION_SUPPORT__:
			self.entry_Name.focus_set()
		else:
			self.entry_SerialNumber.focus_set()

	def Default(self):
		# Common configuration
		board_type = self.notebook.index(tab_id="current")
		self.combobox_PartNumber.current(board_type)
		self.combobox_BoardType.current(board_type)
		self.SerialNumber.set("")
		if board_type == iCOMOX_messages.cCOMOX_BOARD_TYPE_SMIP:
			board_name = "iCOMOX SMIP"
			board_version_major = 1
			board_version_minor = 7
		elif board_type == iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT:
			board_name = "iCOMOX NB-IoT"
			board_version_major = 1
			board_version_minor = 5
		else:
			board_name = "iCOMOX PoE"
			board_version_major = 2
			board_version_minor = 0
		self.Name.set(board_name)
		self.BoardVersionMajor.set(board_version_major)
		self.BoardVersionMinor.set(board_version_minor)

		# NB-IoT default configuration
		self.NBIOT_DefaultSIMConfig.set(True)
		for AcT in range(0,3):
			self.NBIOT_BitmaskAcT[AcT].set(False)
			self.NBIOT_ScanBands[AcT].set("0000")
		self.combobox_NBIOT_DefSimConfig_ScanOrder.current(0)
		self.combobox_NBIOT_DefSimConfig_ServiceDomain.current(1)
		self.on_NBIOT_UseDefSimConfig()

		self.NBIOT_ManualOperatorSelection.set(False)
		self.NBIOT_PLMN.set("42503")
		self.on_NBIOT_ManualOperatorSelection()

		self.NBIOT_EnableRoaming.set(True)
		self.combobox_NBIOT_ContextType.current(0)
		self.NBIOT_ApnAccessName.set("sphone.pelephone.net.il")
		self.NBIOT_ApnUser.set("pcl@3g")
		self.NBIOT_ApnPassword.set("rl")

		self.NBIOT_SSL_Enable.set(False)
		self.NBIOT_SSL_NegotiateTimeoutSec.set(300)
		self.NBIOT_SSL_IgnoreLocalTime.set(True)
		self.combobox_NBIOT_SSL_Version.current(3)
		self.combobox_NBIOT_SSL_CipherSuite.current(13)
		self.combobox_NBIOT_SSL_SecurityLevel.current(0)
		self.on_NBIOT_SSL_Enable()

		self.frame_NBIOT_backbone["combobox_backbone"].current(0)
		self.frame_NBIOT_backbone["TCP_ServerURL"].set("31.154.35.78")
		self.frame_NBIOT_backbone["TCP_ServerPort"].set(1201)
		self.frame_NBIOT_backbone["TCP_delayBeforeReconnectingSec"].set(5)
		self.frame_NBIOT_backbone["MQTT_ServerURL"].set("31.154.35.78")
		self.frame_NBIOT_backbone["MQTT_ServerPort"].set(1883)
		self.frame_NBIOT_backbone["MQTT_ClientID"].set("")
		self.frame_NBIOT_backbone["MQTT_UserName"].set("")
		self.frame_NBIOT_backbone["MQTT_Password"].set("")
		self.frame_NBIOT_backbone["MQTT_Flags_ProvideUserName"].set(True)
		self.frame_NBIOT_backbone["MQTT_Flags_ProvidePassword"].set(True)
		self.frame_NBIOT_backbone["MQTT_Keepalive"].set(400)
		self.frame_NBIOT_backbone["combobox_MQTT_Cloud"].current(0)

		# PoE default configuration
		self.POE_StaticIpAddr.set("192.168.15.150")
		self.POE_MaskAddress.set("255.255.255.0")
		self.POE_DnsServerAddr.set("192.168.15.254")
		self.POE_Gateway.set("192.168.15.254")

		self.POE_DHCP.set(False)
		self.POE_IPv6.set(False)
		self.POE_IPv6LocalLinkOnly.set(False)
		self.POE_DHCP.set(False)
		self.on_POE_DHCP()

		self.frame_POE_backbone["combobox_backbone"].current(0)
		self.frame_POE_backbone["TCP_ServerURL"].set("192.168.15.110")
		self.frame_POE_backbone["TCP_ServerPort"].set(1201)
		self.frame_POE_backbone["TCP_delayBeforeReconnectingSec"].set(5)
		self.frame_POE_backbone["MQTT_ServerURL"].set("192.168.15.110")
		self.frame_POE_backbone["MQTT_ServerPort"].set(1883)
		self.frame_POE_backbone["MQTT_ClientID"].set("")
		self.frame_POE_backbone["MQTT_UserName"].set("")
		self.frame_POE_backbone["MQTT_Password"].set("")
		self.frame_POE_backbone["MQTT_Flags_ProvideUserName"].set(True)
		self.frame_POE_backbone["MQTT_Flags_ProvidePassword"].set(True)
		self.frame_POE_backbone["MQTT_Keepalive"].set(400)
		self.frame_POE_backbone["combobox_MQTT_Cloud"].current(0)

		self.SetDefaultFocus()

	def progress_change_color(self, color):
		common.app.style.configure("EEPROM.Horizontal.TProgressbar", background=color)
		self.progressbar.configure(style="EEPROM.Horizontal.TProgressbar")
		self.progressbar.update()

	def __progressbar_update__(self):
		self.progressbar.configure(value=self.address)
		self.progressbar.update()

	def EnableButtons(self, enable):
		state = "normal" if enable else "disabled"
		self.button_ReadEEPROM.configure(state=state)
		self.button_ProgramEEPROM.configure(state=state)
		self.button_VerifyEEPROM.configure(state=state)

	def StartCycle(self):
		self.EnableButtons(enable=False)
		self.progress_change_color(color="green")
		self.progressbar.update()

	def StopCycle(self):
		if self.after_id is not None:
			self.after_cancel(id=self.after_id)
			self.after_id = None
		if self.state == self.cSTATE_ERROR:
			if self.address < 0:
				self.address = 0
			self.progressbar.configure(value=self.address)
			self.progress_change_color(color="red")
		else:
			self.__progressbar_update__()
		self.EnableButtons(enable=True)

	def ValidateWin_Common(self):
		if (self.combobox_BoardType.current() >= iCOMOX_messages.cCOMOX_BOARD_TYPE_COUNT):
			return self.combobox_BoardType, "board type is incorrect"
		if (not isinstance(self.BoardVersionMajor.get(), (int)) or (self.BoardVersionMajor.get()) < 0) or (self.BoardVersionMajor.get() > 255):
			return self.BoardVersionMajor, "board major version must be an integer in the range of 0 to 255"
		if (not isinstance(self.BoardVersionMinor.get(), (int)) or (self.BoardVersionMinor.get()) < 0) or (self.BoardVersionMinor.get() > 255):
			return self.BoardVersionMinor, "board minor version must be an integer in the range of 0 to 255"
		if not any(s in self.PartNumber.get() for s in MemMap.list_iCOMOX_PartNumber):
			return self.PartNumber, "part number is not supported"
		if len(self.SerialNumber.get()) < 11:
			return self.SerialNumber, "serial number must have at least 11 characters"
		elif len(self.SerialNumber.get()) > 32:
			return self.SerialNumber, "serial number must have at most 32 characters"
		return None, None

	def ValidateWin_NBIOT(self):
		if self.NBIOT_ManualOperatorSelection.get():
			if not MemMap.ValidPLMNString(self.NBIOT_PLMN.get()):
				return self.entry_NBIOT_ManualOperatorSelection_PLMN, "PLMN must have 5 or 6 digits (3 digits for MCC + 2 or 3 digits for MNC)"
		if len(self.NBIOT_ApnAccessName.get()) > 32:
			return self.entry_NBIOT_ApnAccessName, "APN Access Name must have at most 32 characters"
		# elif not urllib.parse.urlparse(self.NBIOT_ApnAccessName.get()):
		# 	return self.entry_NBIOT_ApnAccessName, "APN Access Name must be a valid URL"
		if len(self.NBIOT_ApnUser.get()) > 32:
			return self.entry_NBIOT_ApnUser, "APN User Name must have at most 32 characters"
		if len(self.NBIOT_ApnPassword.get()) > 32:
			return self.entry_NBIOT_ApnPassword, "APN password must have at most 32 characters"
		return None, None

	def ValidateWin_NBIOT_SSL(self):
		if self.NBIOT_SSL_Enable.get():
			if (self.NBIOT_SSL_NegotiateTimeoutSec.get() < 10) or (self.NBIOT_SSL_NegotiateTimeoutSec.get() > 300):
				return self.spinbox_NBIOT_SSL_NegotiateTimeout, "SSL negotiate time must be an integer in the range of 10 to 300"
		return None, None

	def ValidateWin_POE(self):
		if not self.POE_DHCP.get():
			client = helpers.stringToIpAddress(self.POE_StaticIpAddr.get())
			if client is None:
				return self.entry_POE_StaticIpAddr, "invalid static IP address of iCOMOX"
			mask = helpers.stringToIpAddress(self.POE_MaskAddress.get())
			if mask is None:
				return self.entry_POE_mask, "invalid static mask address"
			dns = helpers.stringToIpAddress(self.POE_DnsServerAddr.get())
			if dns is None:
				return self.entry_POE_DnsServerAddr, "invalid DNS server address"
			gateway = helpers.stringToIpAddress(self.POE_Gateway.get())
			if self.POE_IPv6.get():
				if client.version != 6:
					return self.entry_POE_StaticIpAddr, "static IP address must be of version 6"
				if mask.version != 6:
					return self.entry_POE_mask, "mask address must be of version 6"
				if dns.version != 6:
					return self.entry_POE_DnsServerAddr, "DNS server must be of version 6"
				if (gateway is not None) and (gateway.version != 6):
					return self.entry_POE_gateway, "gateway address must be of version 6 (or empty)"
			else:
				if client.version != 4:
					return self.entry_POE_StaticIpAddr, "static IP address must be of version 4"
				if mask.version != 4:
					return self.entry_POE_mask, "mask address must be of version 4"
				if dns.version != 4:
					return self.entry_POE_DnsServerAddr, "DNS server must be of version 4"
				if gateway is None:
					return self.entry_POE_gateway, "Invalid gateway address"
				elif gateway.version != 4:
					return self.entry_POE_gateway, "gateway address must be of version 4"

		return None, None

	def ValidateWin_POE_SSL(self):
		return None, None

	def ValidateWin_Backbone(self, frame_backbone):
		if frame_backbone["combobox_backbone"].current() == 0:	# TCP
			if (len(frame_backbone["TCP_ServerURL"].get()) == 0) or (len(frame_backbone["TCP_ServerURL"].get()) > 128):
				return frame_backbone["entry_TCP_ServerURL"], "server URL must be non empty and limited to 128 characters"
			if (frame_backbone["TCP_ServerPort"].get() < 1025) or (frame_backbone["TCP_ServerPort"].get() > 65535):
				return frame_backbone["entry_TCP_ServerPort"], "server port must be in the range of 1025 to 65535"
			if (frame_backbone["TCP_delayBeforeReconnectingSec"].get() < 0) or (frame_backbone["TCP_delayBeforeReconnectingSec"].get() > 255):
				return frame_backbone["entry_TCP_DelayBeforeReconnecting"], "delay before reconnecting must be in the range of 0 to 255 seconds"
		elif frame_backbone["combobox_backbone"].current() == 1: # MQTT
			if (len(frame_backbone["MQTT_ServerURL"].get()) == 0) or (len(frame_backbone["MQTT_ServerURL"].get()) > 128):
				return frame_backbone["entry_MQTT_ServerURL"], "MQTT server URL must be non empty and limited to 128 characters"
			if (frame_backbone["MQTT_ServerPort"].get() < 1024) or (frame_backbone["MQTT_ServerPort"].get() > 65535):
				return frame_backbone["entry_MQTT_ServerPort"], "server port must be in the range of 1025 to 65535"
			if len(frame_backbone["MQTT_ClientID"].get()) > 32:
				return frame_backbone["entry_MQTT_clientID"], "client ID is limited to 32 characters"
			if len(frame_backbone["MQTT_UserName"].get()) > 32:
				return frame_backbone["entry_MQTT_UserName"], "user name is limited to 32 characters"
			if len(frame_backbone["MQTT_Password"].get()) > 32:
				return frame_backbone["entry_MQTT_Password"], "password is limited to 32 characters"
			if (len(frame_backbone["MQTT_Keepalive"].get()) < 0) or (len(frame_backbone["MQTT_Keepalive"].get()) > 65535):
				return frame_backbone["entry_MQTT_Keepalive"], "Keepalive must be in the range of 0 to 65535"

		else:
			raise Exception("ValidateWin_Backbone(): Unknown combobox_backbone value")

		return None, None

	def ValidateWin(self, ProductionSupport=False):
		widget = None
		error_description = None
		if ProductionSupport:
			widget, error_description = self.ValidateWin_Common()
		while widget is None:
			if (not ProductionSupport) and (common.iCOMOXs.current is not None) and (common.iCOMOXs.current.board_type() != self.combobox_BoardType.current()):
				widget = self.notebook
				error_description = "board type configuration does not match the currently connected iCOMOX"
				break

			if self.combobox_BoardType.current() == iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT:
				widget, error_description = self.ValidateWin_NBIOT()
				if widget is not None:
					break
				widget, error_description = self.ValidateWin_NBIOT_SSL()
				if widget is not None:
					break
				widget, error_description = self.ValidateWin_Backbone(self.frame_NBIOT_backbone)

			elif self.combobox_BoardType.current() == iCOMOX_messages.cCOMOX_BOARD_TYPE_POE:
				widget, error_description = self.ValidateWin_POE()
				if widget is not None:
					break
				widget, error_description = self.ValidateWin_POE_SSL()
				if widget is not None:
					break
				widget, error_description = self.ValidateWin_Backbone(self.frame_POE_backbone)

			break	# one must exit the artifical while ()

		return widget, error_description

	def WinToMemMap(self, ProductionSupport=False):
		result = True
		try:
			self.MemMap.Clear()
			self.MemMap.WinToMemMap_Common(
				MemMapVersion=MemMap.cMEMMAP_VERSION,
				BoardType=self.combobox_BoardType.current(),
				BoardVersionMajor=self.BoardVersionMajor.get(),
				BoardVersionMinor=self.BoardVersionMinor.get(),
				PartNumber=self.PartNumber.get(),
				SerialNumber=self.SerialNumber.get(),
				Name=self.Name.get(),
				ProductionSupport=ProductionSupport
			)
			if self.combobox_BoardType.current() == iCOMOX_messages.cCOMOX_BOARD_TYPE_SMIP:
				pass
			elif self.combobox_BoardType.current() == iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT:
				frame_backbone = self.frame_NBIOT_backbone
				self.MemMap.WinToMemMap_NBIOT(
					UseDefSimConfiguration=self.NBIOT_DefaultSIMConfig.get(),
					AccessTechnologiesBitmask=self.NBIOT_BitmaskAcT[0].get() + (self.NBIOT_BitmaskAcT[1].get() << 1) + (self.NBIOT_BitmaskAcT[2].get() << 2),
					GsmBandsBitmask=int(self.NBIOT_ScanBands[0].get(), 16),
					LteCatM1BandsBitmask=int(self.NBIOT_ScanBands[1].get(), 16),
					LteCatNB1BandsBitmask=int(self.NBIOT_ScanBands[2].get(), 16),
					ScanOrder=self.combobox_NBIOT_DefSimConfig_ScanOrder.current(),
					ServiceDomain=self.combobox_NBIOT_DefSimConfig_ServiceDomain.current(),
					ManualOperatorSelection=self.NBIOT_ManualOperatorSelection.get(),
					PLMN=self.NBIOT_PLMN.get(),
					EnableRoaming=self.NBIOT_EnableRoaming.get(),
					ContextType=self.combobox_NBIOT_ContextType.current(),
					Authentication=self.combobox_NBIOT_Authentication.current(),
					ApnAccessName=self.NBIOT_ApnAccessName.get(),
					ApnUser=self.NBIOT_ApnUser.get(),
					ApnPassword=self.NBIOT_ApnPassword.get(),
					SSL_Enable=self.NBIOT_SSL_Enable.get(),
					SSL_version=self.combobox_NBIOT_SSL_Version.current(),
					SSL_SecLevel=self.combobox_NBIOT_SSL_SecurityLevel.current(),
					SSL_CipherSuite=self.combobox_NBIOT_SSL_CipherSuite.current(),
					SSL_NegotiateTimeout=self.NBIOT_SSL_NegotiateTimeoutSec.get(),
					SSL_IgnoreLocalTime=self.NBIOT_SSL_IgnoreLocalTime.get(),
					BackboneProtocol=frame_backbone["combobox_backbone"].current(),
					TCP_ServerURL=frame_backbone["TCP_ServerURL"].get(), TCP_ServerPort=frame_backbone["TCP_ServerPort"].get(), TCP_delayBeforeReconnectingSec=frame_backbone["TCP_delayBeforeReconnectingSec"].get(),
					MQTT_ServerURL=frame_backbone["MQTT_ServerURL"].get(), MQTT_UserName=frame_backbone["MQTT_UserName"].get(), MQTT_Password=frame_backbone["MQTT_Password"].get(), MQTT_ClientID=frame_backbone["MQTT_ClientID"].get(), MQTT_ServerPort=frame_backbone["MQTT_ServerPort"].get(), MQTT_ProvideUserName=frame_backbone["MQTT_Flags_ProvideUserName"].get(), MQTT_ProvidePassword=frame_backbone["MQTT_Flags_ProvidePassword"].get(), MQTT_Keepalive=frame_backbone["MQTT_Keepalive"].get(), MQTT_Cloud=frame_backbone["combobox_MQTT_Cloud"].current()
				)
			elif self.combobox_BoardType.current() == iCOMOX_messages.cCOMOX_BOARD_TYPE_POE:
				frame_backbone = self.frame_POE_backbone
				self.MemMap.WinToMemMap_POE(
					StaticIpAddr=self.POE_StaticIpAddr.get(),
					StaticMaskAddr=self.POE_MaskAddress.get(),
					StaticDnsServerAddr=self.POE_DnsServerAddr.get(),
					StaticGatewayAddr=self.POE_Gateway.get(),
					IPv6=False,	# self.POE_IPv6.get(),								# Remove in 2.8.1
					IPv6LocalLinkOnly=False, # self.POE_IPv6LocalLinkOnly.get(),	# Remove in 2.8.1
					# TLS=False,  # self.POE_TLS.get()
					DHCP=self.POE_DHCP.get(),
					BackboneProtocol = frame_backbone["combobox_backbone"].current(),
					TCP_ServerURL=frame_backbone["TCP_ServerURL"].get(), TCP_ServerPort=frame_backbone["TCP_ServerPort"].get(), TCP_delayBeforeReconnectingSec=frame_backbone["TCP_delayBeforeReconnectingSec"].get(),
					MQTT_ServerURL=frame_backbone["MQTT_ServerURL"].get(), MQTT_UserName=frame_backbone["MQTT_UserName"].get(), MQTT_Password=frame_backbone["MQTT_Password"].get(), MQTT_ClientID=frame_backbone["MQTT_ClientID"].get(), MQTT_ServerPort=frame_backbone["MQTT_ServerPort"].get(), MQTT_ProvideUserName=frame_backbone["MQTT_Flags_ProvideUserName"].get(), MQTT_ProvidePassword=frame_backbone["MQTT_Flags_ProvidePassword"].get(), MQTT_Keepalive=frame_backbone["MQTT_Keepalive"].get(), MQTT_Cloud=frame_backbone["combobox_MQTT_Cloud"].current()
				)
		except Exception as ex:
			if hasattr(ex, 'message'):
				helpers.OUT(ex.message)
			result = False
		finally:
			pass
		return result

	def ValidateMemMap_Common(self, BoardType, BoardVersionMajor, BoardVersionMinor, PartNumber, SerialNumber):
		if (BoardType >= iCOMOX_messages.cCOMOX_BOARD_TYPE_COUNT):
			return self.combobox_BoardType, "board type can not recognized"
		if (BoardVersionMajor == 255) and (BoardVersionMinor == 255):
			return self.BoardVersionMajor, "board version is illegal"
		if not any(s in PartNumber for s in MemMap.list_iCOMOX_PartNumber):
			return self.PartNumber, "part number can not recognized"
		if len(SerialNumber) < 11:
			return self.SerialNumber, "serial number must contain at least 11 characters"
		return None, None

	def MemMapToWin(self):
		MemMapVersion, BoardType, BoardVersionMajor, BoardVersionMinor, PartNumber, SerialNumber, Name = self.MemMap.MemMapToWin_Common()
		self.BoardVersionMajor.set(BoardVersionMajor)
		self.BoardVersionMinor.set(BoardVersionMinor)
		self.PartNumber.set(PartNumber)
		self.SerialNumber.set(SerialNumber)
		self.Name.set(Name)
		frame_backbone = None

		widget, error_description = self.ValidateMemMap_Common(BoardType=BoardType, BoardVersionMajor=BoardVersionMajor, BoardVersionMinor=BoardVersionMinor, PartNumber=PartNumber, SerialNumber=SerialNumber)
		if widget is not None:
			return widget, error_description

		if BoardType <= iCOMOX_messages.cCOMOX_BOARD_TYPE_POE:
			self.combobox_BoardType.current(BoardType)
			if BoardType == iCOMOX_messages.cCOMOX_BOARD_TYPE_SMIP:
				self.notebook.select(self.tab_smip)
			elif BoardType == iCOMOX_messages.cCOMOX_BOARD_TYPE_NB_IOT:
				frame_backbone = self.frame_NBIOT_backbone

				UseDefSimConfiguration, AccessTechnologiesBitmask, GsmBandsBitmask, LteCatM1BandsBitmask, LteCatNB1BandsBitmask, ScanOrder, ServiceDomain, \
				ManualOperatorSelection, PLMN, \
				EnableRoaming, ContextType, Authentication, ApnAccessName, ApnUser, ApnPassword, \
				SSL_Enable, SSL_version, SSL_SecLevel, SSL_CipherSuite, SSL_NegotiateTimeout, SSL_IgnoreLocalTime, \
				BackboneProtocol, \
				TCP_ServerURL, TCP_ServerPort, TCP_delayBeforeReconnectingSec, \
				MQTT_ServerURL, MQTT_UserName, MQTT_Password, MQTT_ClientID, MQTT_ServerPort, MQTT_ProvideUserName, MQTT_ProvidePassword, MQTT_Keepalive, MQTT_Cloud = self.MemMap.MemMapToWin_NBIOT()

				self.NBIOT_DefaultSIMConfig.set(UseDefSimConfiguration)
				self.NBIOT_BitmaskAcT[0].set(0 != (AccessTechnologiesBitmask & 0x01))
				self.NBIOT_BitmaskAcT[1].set(0 != (AccessTechnologiesBitmask & 0x02))
				self.NBIOT_BitmaskAcT[2].set(0 != (AccessTechnologiesBitmask & 0x04))
				self.NBIOT_ScanBands[0].set("{:04X}".format(GsmBandsBitmask))
				self.NBIOT_ScanBands[1].set("{:04X}".format(LteCatM1BandsBitmask))
				self.NBIOT_ScanBands[2].set("{:04X}".format(LteCatNB1BandsBitmask))
				self.combobox_NBIOT_DefSimConfig_ScanOrder.current(ScanOrder)
				self.combobox_NBIOT_DefSimConfig_ServiceDomain.current(ServiceDomain)
				self.on_NBIOT_UseDefSimConfig()

				self.NBIOT_ManualOperatorSelection.set(ManualOperatorSelection)
				self.on_NBIOT_ManualOperatorSelection()
				self.NBIOT_PLMN.set(PLMN)

				self.NBIOT_EnableRoaming.set(EnableRoaming)
				self.combobox_NBIOT_ContextType.current(ContextType)
				self.combobox_NBIOT_Authentication.current(Authentication)
				self.NBIOT_ApnAccessName.set(ApnAccessName)
				self.NBIOT_ApnUser.set(ApnUser)
				self.NBIOT_ApnPassword.set(ApnPassword)

				self.NBIOT_SSL_Enable.set(SSL_Enable)
				self.combobox_NBIOT_SSL_Version.current(SSL_version)
				self.combobox_NBIOT_SSL_SecurityLevel.current(SSL_SecLevel)
				self.combobox_NBIOT_SSL_CipherSuite.current(SSL_CipherSuite)
				self.NBIOT_SSL_NegotiateTimeoutSec.set(SSL_NegotiateTimeout)
				self.NBIOT_SSL_IgnoreLocalTime.set(SSL_IgnoreLocalTime)
				self.on_NBIOT_SSL_Enable()

				self.notebook.select(self.tab_nbiot)
			else:
				frame_backbone = self.frame_POE_backbone

				StaticIpAddr, StaticMaskAddr, StaticDnsServerAddr, StaticGatewayAddr, IPv6, IPv6LocalLinkOnly, DHCP, \
				BackboneProtocol, \
				TCP_ServerURL, TCP_ServerPort, TCP_delayBeforeReconnectingSec, \
				MQTT_ServerURL, MQTT_UserName, MQTT_Password, MQTT_ClientID, MQTT_ServerPort, MQTT_ProvideUserName, MQTT_ProvidePassword, MQTT_Keepalive, MQTT_Cloud = self.MemMap.MemMapToWin_POE()

				self.POE_StaticIpAddr.set(StaticIpAddr)
				self.POE_MaskAddress.set(StaticMaskAddr)
				self.POE_DnsServerAddr.set(StaticDnsServerAddr)
				self.POE_Gateway.set(StaticGatewayAddr)
				self.POE_DHCP.set(DHCP)
				self.on_POE_DHCP()	# Disable/enable the static IP address, mask address & default gateway according to the DHCP settings
				# self.TCP_ServerURL.set(TcpServerIpAddr)
				# self.TCP_ServerPort.set(TcpServerPort)
				# self.TCP_delayBeforeReconnectingSec.set(delayBeforeConnectingSec)
				IPv6 = False				# Remove in 2.8.1
				self.POE_IPv6.set(IPv6)
				IPv6LocalLinkOnly = False	# Remove in 2.8.1
				self.POE_IPv6LocalLinkOnly.set(IPv6LocalLinkOnly)
				# self.POE_TLS.set(TLS)
				self.notebook.select(self.tab_poe)
		else:
			self.combobox_BoardType.set("")
			return self.combobox_BoardType

		# common parameters to both NB-IoT & PoE
		if frame_backbone is not None:
			if BackboneProtocol >= iCOMOX_messages.cBACKBONE_PROTOCOL_COUNT:
				return frame_backbone["combobox_backbone"]
			frame_backbone["combobox_backbone"].current(BackboneProtocol)
			frame_backbone["notebook_backbone"].select(tab_id=BackboneProtocol)

			# TCP
			if BackboneProtocol == iCOMOX_messages.cBACKBONE_PROTOCOL_TCP:
				if TCP_ServerURL is None:
					return frame_backbone["entry_ServerURL"]

			frame_backbone["TCP_ServerURL"].set("" if TCP_ServerURL is None else TCP_ServerURL)
			frame_backbone["TCP_ServerPort"].set(1201 if TCP_ServerPort is None else TCP_ServerPort)
			frame_backbone["TCP_delayBeforeReconnectingSec"].set(5 if TCP_delayBeforeReconnectingSec is None else TCP_delayBeforeReconnectingSec)
			# self.NBIOT_connectAttemptsCount.set(0 if TCP_connectAttemptsCount is None else TCP_connectAttemptsCount)
			# self.TCP_IPv6.set(False if TCP_IPv6 is None else TCP_IPv6)
			# self.NBIOT_TLS.set(TLS)
			# MQTT
			frame_backbone["MQTT_ServerURL"].set("" if MQTT_ServerURL is None else MQTT_ServerURL)
			frame_backbone["MQTT_UserName"].set("" if MQTT_UserName is None else MQTT_UserName)
			frame_backbone["MQTT_Password"].set("" if MQTT_Password is None else MQTT_Password)
			frame_backbone["MQTT_ClientID"].set("" if MQTT_ClientID is None else MQTT_ClientID)
			frame_backbone["MQTT_ServerPort"].set(1883 if MQTT_ServerPort is None else MQTT_ServerPort)
			frame_backbone["MQTT_Flags_ProvideUserName"].set(True if MQTT_ProvideUserName is None else MQTT_ProvideUserName)
			frame_backbone["MQTT_Flags_ProvidePassword"].set(True if MQTT_ProvidePassword is None else MQTT_ProvidePassword)
			frame_backbone["MQTT_Keepalive"].set(400 if MQTT_Keepalive is None else MQTT_Keepalive)
			frame_backbone["combobox_MQTT_Cloud"].current(0 if MQTT_Cloud is None else MQTT_Cloud)
			if BackboneProtocol == iCOMOX_messages.cBACKBONE_PROTOCOL_MQTT:
				frame_backbone["notebook_MQTT_clouds"].select(tab_id=0 if MQTT_Cloud is None else MQTT_Cloud)
		return None, None

	def ReadEEPROM_callback(self, nextState, data):
		if self.state >= self.cSTATE_FINISH_OK:
			return
		self.state = nextState
		if self.state == self.cSTATE_INIT:
			self.address = -iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			self.StartCycle()
			self.after_id = self.after(ms=MemMap.cMEMMAP_USED_PAGES * 500, func=lambda : self.ReadEEPROM_callback(nextState=self.cSTATE_ERROR, data=None))
			self.state = self.cSTATE_OK
			common.app.StatusBar.set(text="Reading...")

		if self.state == self.cSTATE_OK:
			if data is not None:
				self.MemMap.memmap[self.address:self.address+iCOMOX_messages.M24C64_PAGE_BYTE_SIZE] = data[:iCOMOX_messages.M24C64_PAGE_BYTE_SIZE]
			self.address += iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			if self.address == MemMap.cMEMMAP_SIZE:
				# if finish successfully
				self.StopCycle()
				# self.MemMap.memmap[iCOMOX_messages.M24C64_PAGE_BYTE_SIZE*3:] = b"\xFF" * (len(self.MemMap.memmap)-iCOMOX_messages.M24C64_PAGE_BYTE_SIZE*3)
				try:
					widget, error_description = self.MemMapToWin()
					if widget is not None:
						common.app.StatusBar.set(text="Reading failed (invalid data)")
						self.state = self.cSTATE_FINISH_FAIL
						widget.focus_set()
						messagebox.showinfo("Error", error_description)
					else:
						common.app.StatusBar.set(text="Reading succeeded")
						self.state = self.cSTATE_FINISH_OK
				except:
					pass
				finally:
					pass
			else:
				# if another page should be read
				common.app.StatusBar.set(text="{}% read".format(100*self.address//MemMap.cMEMMAP_SIZE))
				common.app.iCOMOX_Data.send_msg(
					msg=iCOMOX_messages.OUT_MSG_ReadEEPROM(
						Count=iCOMOX_messages.M24C64_PAGE_BYTE_SIZE,
						Address=self.address))
				self.__progressbar_update__()
		else:
			# if error or timeout error was detected
			self.StopCycle()
			common.app.StatusBar.set(text="Reading failed")
			self.state = self.cSTATE_FINISH_FAIL

	def ReadEEPROM(self):
		self.state = self.cSTATE_INIT
		self.ReadEEPROM_callback(nextState=self.cSTATE_INIT, data=None)
		self.SetDefaultFocus()

	def ProgramEEPROM_callback(self, nextState):
		if self.state >= self.cSTATE_FINISH_OK:
			return
		self.state = nextState
		if self.state == self.cSTATE_INIT:
			if common_symbols.__PRODUCTION_SUPPORT__:
				self.address = -iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			else:
				self.address = (3-1)*iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			self.StartCycle()
			self.after_id = self.after(ms=MemMap.cMEMMAP_USED_PAGES * 500, func=lambda : self.ProgramEEPROM_callback(nextState=self.cSTATE_ERROR))
			self.state = self.cSTATE_OK
			common.app.StatusBar.set(text="Programming...")

		if self.state == self.cSTATE_OK:
			self.address += iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			if self.address == MemMap.cMEMMAP_SIZE:
				# if finish successfully
				self.StopCycle()
				common.app.StatusBar.set(text="Programming succeeded")
				self.state = self.cSTATE_FINISH_OK
			else:
				# if another page should be programmed
				common.app.StatusBar.set(text="{}% programmed".format(100*self.address//MemMap.cMEMMAP_SIZE))
				common.app.iCOMOX_Data.send_msg(
					msg=iCOMOX_messages.OUT_MSG_WriteEEPROM(
						Count=iCOMOX_messages.M24C64_PAGE_BYTE_SIZE,
						Address=self.address,
						Data=self.MemMap.memmap[self.address:self.address + iCOMOX_messages.M24C64_PAGE_BYTE_SIZE]))
				self.__progressbar_update__()
		else:
			# if error or timeout error was detected
			self.StopCycle()
			common.app.StatusBar.set(text="Programming failed")
			self.state = self.cSTATE_FINISH_FAIL

	def ProgramEEPROM(self):
		widget, error_description = self.ValidateWin(ProductionSupport=common_symbols.__PRODUCTION_SUPPORT__)
		if widget is not None:
			widget.focus_set()
			messagebox.showinfo("Error", error_description)
			return
		self.WinToMemMap(ProductionSupport=common_symbols.__PRODUCTION_SUPPORT__)
		self.state = self.cSTATE_INIT
		self.ProgramEEPROM_callback(nextState=self.cSTATE_INIT)
		self.SetDefaultFocus()

	def VerifyEEPROM_callback(self, nextState):
		if self.state >= self.cSTATE_FINISH_OK:
			return
		self.state = nextState
		if self.state == self.cSTATE_INIT:
			if common_symbols.__PRODUCTION_SUPPORT__:
				self.address = 0	#-iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			else:
				self.address = (3)*iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			self.StartCycle()
			self.after_id = self.after(ms=MemMap.cMEMMAP_USED_PAGES * 500, func=lambda : self.VerifyEEPROM_callback(nextState=self.cSTATE_ERROR))
			self.state = self.cSTATE_OK
			common.app.StatusBar.set(text="Verifying...")

		if self.state == self.cSTATE_OK:
			self.address += iCOMOX_messages.M24C64_PAGE_BYTE_SIZE
			if self.address == MemMap.cMEMMAP_SIZE:
				# if finish successfully
				self.StopCycle()
				common.app.StatusBar.set(text="Verifying succeeded")
				self.state = self.cSTATE_FINISH_OK
			else:
				# if another page should be programmed
				common.app.StatusBar.set(text="{}% verified".format(100*self.address//MemMap.cMEMMAP_SIZE))
				common.app.iCOMOX_Data.send_msg(
					msg=iCOMOX_messages.OUT_MSG_VerifyEEPROM(
						Count=iCOMOX_messages.M24C64_PAGE_BYTE_SIZE,
						Address=self.address,
						Data=self.MemMap.memmap[self.address:self.address + iCOMOX_messages.M24C64_PAGE_BYTE_SIZE]))
				self.__progressbar_update__()
		else:
			# if error or timeout error was detected
			self.StopCycle()
			common.app.StatusBar.set(text="Verifying failed")
			self.state = self.cSTATE_FINISH_FAIL

	def VerifyEEPROM(self):
		widget, error_description = self.ValidateWin(ProductionSupport=common_symbols.__PRODUCTION_SUPPORT__)
		if widget is not None:
			widget.focus_set()
			messagebox.showinfo("Error", error_description)
			return
		self.WinToMemMap(ProductionSupport=common_symbols.__PRODUCTION_SUPPORT__)
		self.state = self.cSTATE_INIT
		self.VerifyEEPROM_callback(nextState=self.cSTATE_INIT)
		self.SetDefaultFocus()


class Clients(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent

		self.SmipClientsTreeView = iCOMOX_ClientsTreeView.cClientsTreeView(parent=self.parent, clientsType=iCOMOX_list.cCLIENT_TYPE_SMIP, on_select_client_command=self.SelectSmipClient, \
		   namewidth_arr = ((50, "Type"), (220, "iCOMOX name"), (220, "Unique ID"), (150, "MAC address"), (60, "Mote ID")))
		self.SmipClientsTreeView.tree_clients.grid(row=0, column=0, pady=(10, 0))

		self.TcpIpClientsTreeView = iCOMOX_ClientsTreeView.cClientsTreeView(parent=self.parent, clientsType=iCOMOX_list.cCLIENT_TYPE_TCPIP, on_select_client_command=self.SelectTcpIpClient, \
			namewidth_arr=((50, "Type"), (220, "iCOMOX name"), (220, "Unique ID"), (150, "IP address"), (60, "Port")))
		self.TcpIpClientsTreeView.tree_clients.grid(row=1, column=0, pady=(0, 10))

		self.buttonDeselectItem = tk.Button(self.parent, command=self.DeselectClient, width=20, text="Deselect iComox", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
		self.buttonDeselectItem.grid(row=2, column=0, padx=10, sticky=tk.W)

		self.buttonDisconnectClient = tk.Button(self.parent, command=self.DisconnectSelectedClient, width=20, text="Disconnect iComox", font=("Verdana", 8), bg="gray80", fg="black", state="disabled")
		self.buttonDisconnectClient.grid(row=2, column=0, padx=10, sticky=tk.E)

	def SelectTcpIpClient(self, iComox):
		self.EnableButtons(enable=True)
		common.app.TopPane.update_iComox()

	def SelectSmipClient(self, iComox):
		self.EnableButtons(enable=True)
		common.app.TopPane.update_iComox()

	def DeselectClient(self):
		if self.SmipClientsTreeView.deselect_items():
			common.app.drawLedState(connection_state=iCOMOX_datahandling.CONNECTION_STATE_Dongle_Connected)
		elif self.TcpIpClientsTreeView.deselect_items():
			common.app.drawLedState(connection_state=iCOMOX_datahandling.CONNECTION_STATE_TcpIp_ServerConnected)
		self.EnableButtons(enable=False)
		common.app.TopPane.update_iComox()
		common.app.EnableButtons(liveDataTab=False, diagnosticTab=False, configurationTab=False, eepromTab=False, clientsTab=False)

	def DisconnectSelectedClient(self):
		# SMIP clients
		for item_id in self.SmipClientsTreeView.tree_clients.selection():
			client = self.SmipClientsTreeView.ItemID_to_iComox(ID_str=item_id)
			if client is not None:
				# self.SmipClientsTreeView.delete(iComox=client)
				common.app.iCOMOX_Data.Dongle_Comm.send_resetMote(macAddress=client.macAddress)
				self.SmipClientsTreeView.delete(iComox=client)
				common.iCOMOXs.list.remove(client)
				client.clear()
				# common.iCOMOXs.current = None
				common.app.TopPane.update_iComox()
				common.app.EnableButtons(liveDataTab=False, diagnosticTab=False, configurationTab=False, eepromTab=False, clientsTab=False)
				common.app.drawLedState(connection_state=iCOMOX_datahandling.CONNECTION_STATE_Dongle_Connected)

		# TCP/IP clients
		for item_id in self.TcpIpClientsTreeView.tree_clients.selection():
			client = self.TcpIpClientsTreeView.ItemID_to_iComox(ID_str=item_id)
			if client is not None:
				self.TcpIpClientsTreeView.delete(iComox=client)
				common.app.iCOMOX_Data.TcpIp_Comm.close_client(iComox=client)
				common.app.TopPane.update_iComox()
				common.app.EnableButtons(liveDataTab=False, diagnosticTab=False, configurationTab=False, eepromTab=False, clientsTab=False)
				common.app.drawLedState(connection_state=iCOMOX_datahandling.CONNECTION_STATE_TcpIp_ServerConnected)

	def deleteZombieClients(self):
		self.TcpIpClientsTreeView.deleteZombieClients()
		if common.iCOMOXs.current is None:
			common.app.drawLedState(connection_state=iCOMOX_datahandling.CONNECTION_STATE_TcpIp_ServerConnected)

	def insert(self, iComox, index=tk.END):
		if iComox.Type == iCOMOX_list.cCLIENT_TYPE_USB:
			pass
		elif iComox.Type == iCOMOX_list.cCLIENT_TYPE_SMIP:
			self.SmipClientsTreeView.insert(iComox=iComox, index=index)
		elif iComox.Type == iCOMOX_list.cCLIENT_TYPE_TCPIP:
			self.TcpIpClientsTreeView.insert(iComox=iComox, index=index)

	def EnableButtons(self, enable):
		state = "normal" if enable else "disabled"
		self.buttonDeselectItem.config(state=state)
		self.buttonDisconnectClient.config(state=state)

	def delete(self, iComox):
		if iComox.Type == iCOMOX_list.cCLIENT_TYPE_USB:
			return
		elif iComox.Type == iCOMOX_list.cCLIENT_TYPE_SMIP:
			self.SmipClientsTreeView.delete(iComox=iComox)
			if common.iCOMOXs.current is None:
				common.app.drawLedState(connection_state=iCOMOX_datahandling.CONNECTION_STATE_Dongle_Connected)
		elif iComox.Type == iCOMOX_list.cCLIENT_TYPE_TCPIP:
			self.TcpIpClientsTreeView.delete(iComox=iComox)
			if common.iCOMOXs.current is None:
				common.app.drawLedState(connection_state=iCOMOX_datahandling.CONNECTION_STATE_TcpIp_ServerConnected)
		common.app.TopPane.update_iComox()

	def delete_all(self):
		self.SmipClientsTreeView.delete_all()
		self.TcpIpClientsTreeView.delete_all()
		common.app.TopPane.update_iComox()


class DiagnosticMode(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent
		# implement stretchability
		# self.parent.grid_columnconfigure(0, weight=1)
		# self.parent.grid_rowconfigure(0, weight=1)
		# self.frame_diag = tk.Frame(self.parent)
		# self.frame_diag.grid(row=0, column=0, padx=0, pady=0, sticky="nswe")
		self.notebook = ttk.Notebook(self.parent)
		self.notebook.pack(expand=True, fill="both")# self.notebook.grid(row=0, column=0, padx=0, pady=0, sticky="nswe")

		# self.notebook.grid_rowconfigure(0, weight=1)
		# self.notebook.grid_columnconfigure(0, weight=1)
		self.tab_smip = tk.Frame(self.notebook)
		self.tab_nbiot = tk.Frame(self.notebook)
		self.tab_poe = tk.Frame(self.notebook)
		self.notebook.add(self.tab_smip, text="SMIP")
		self.notebook.add(self.tab_nbiot, text="NB-IOT")
		self.notebook.add(self.tab_poe, text="POE")
		self.notebook.select(tab_id=self.tab_nbiot)

		self.tab_nbiot.grid_columnconfigure(0, weight=1)
		self.tab_nbiot.grid_rowconfigure(1, weight=1)

		self.NBIOT_ATCommand = tk.StringVar(value="")
		self.NBIOT_Timeout = tk.IntVar(value=300)
		self.NBIOT_Response = tk.StringVar(value="")

		# self.NBIOT_Manual_AT_Commands 			= ["AT+COPN", "AT+CPOL?", "AT+CREG?", "AT+CSQ", "AT+QCCID", "AT+QCFG=?", "AT+QICSGP=1", "AT+QISTATE", "AT+QPING=1,<host>", "AT+QSPN", "AT+QCSQ", "AT+QIGETERROR", "AT+QINISTAT", "AT+QNWINFO", "AT+QSIMSTAT?", "AT+COPS=?", "AT+COPS?"]
		# self.NBIOT_Manual_AT_Commands_Timeouts 	= [5000, 		300, 		300, 		300, 		300,		300,		1000,			1000,		4000,				1000,		300,		1000,				300,		300,			300,		180000,		180000]
		self.NBIOT_Manual_AT_Commands 			= ["AT+CREG?", "AT+CSQ", "AT+QCCID", "AT+QSIMSTAT?"]
		self.NBIOT_Manual_AT_Commands_Timeouts 	= [300, 		300, 		300,		300]

		frame_command = tk.Frame(self.tab_nbiot, bd=1, relief=tk.RIDGE)
		frame_command.grid(row=0, column=0, sticky="we")
		if True:
			self.label_NBIOT_ATCommand = tk.Label(frame_command, justify="left", text="AT command:", width=12, font=("Verdana", 10), bd=0)
			self.label_NBIOT_ATCommand.grid(row=0, column=0, ipadx=2, ipady=10, padx=(0,5), sticky="w")
			self.combobox_NBIOT_ATCommand = ttk.Combobox(frame_command, textvariable=self.NBIOT_ATCommand, width=90, values=self.NBIOT_Manual_AT_Commands)
			self.combobox_NBIOT_ATCommand.grid(row=0, column=1, padx=(0, 5), sticky="w")
			self.label_NBIOT_Timeout = tk.Label(frame_command, justify="left", text="Timeout [msec]:", width=16, font=("Verdana", 10), bd=0)	# width=10,
			self.label_NBIOT_Timeout.grid(row=0, column=2, padx=(5, 0), ipadx=2, ipady=10, sticky="w")
			self.entry_NBIOT_Timeout = tk.Entry(frame_command, textvariable=self.NBIOT_Timeout, width=10)
			self.entry_NBIOT_Timeout.grid(row=0, column=3, padx=(0, 5), sticky="w")
			def on_combobox_NBIOT_ATCommand_Selected(eventObject):
				new_timeout = self.NBIOT_Manual_AT_Commands_Timeouts[self.combobox_NBIOT_ATCommand.current()]
				self.NBIOT_Timeout.set(new_timeout)
			self.combobox_NBIOT_ATCommand.bind("<<ComboboxSelected>>", on_combobox_NBIOT_ATCommand_Selected)

			self.button_NBIOT_Send_ATCommand = tk.Button(frame_command, command=self.NBIOT_Send_ATCommand, width=20, text="Send...", font=("Verdana", 8), bg="gray80", fg="black", state="normal")
			self.button_NBIOT_Send_ATCommand.grid(row=0, column=4, padx=5, sticky=tk.W)

			self.button_NBIOT_Reset_BG96 = tk.Button(frame_command, command=self.NBIOT_Reset_BG96, width=20, text="Reset modem", font=("Verdana", 8), bg="gray80", fg="black", state="normal")
			self.button_NBIOT_Reset_BG96.grid(row = 0, column=5, padx=(5,0), sticky=tk.W)

		self.button_NBIOT_TestConnectivity = tk.Button(self.tab_nbiot, command=self.NBIOT_TestConnectivity, width=20, text="Test connectivity...", font=("Verdana", 8), bg="gray80", fg="black", state="normal")
		self.button_NBIOT_TestConnectivity.grid(row = 0, column=1, padx=5, sticky=tk.W)


		self.text_NBIOT_Response = tk.Text(self.tab_nbiot, wrap=tk.CHAR)
		self.text_NBIOT_Response.grid(row=1, column=0, columnspan=7, padx=0, sticky="nsew")
		self.text_NBIOT_Response.config(borderwidth=3, relief="sunken")
		self.scrollbar_NBIOT_Response = tk.Scrollbar(self.tab_nbiot, command=self.text_NBIOT_Response.yview, orient=tk.VERTICAL)
		self.scrollbar_NBIOT_Response.grid(row=1, column=8, sticky='esn')
		self.text_NBIOT_Response.config(yscrollcommand=self.scrollbar_NBIOT_Response.set)

		self.Enable(False)

	def NBIOT_Send_ATCommand(self):
		AT_command = self.combobox_NBIOT_ATCommand.get() + "\r"
		common.app.iCOMOX_Data.send_msg(msg=iCOMOX_messages.OUT_MSG_Debug_Send_AT_Command(AT_command_string=AT_command, timeout_in_msec=self.NBIOT_Timeout.get()))

	def NBIOT_Reset_BG96(self):
		common.app.iCOMOX_Data.send_msg(msg=iCOMOX_messages.OUT_MSG_Debug_Reset_BG96())

	def NBIOT_TestConnectivity(self):
		common.app.iCOMOX_Data.send_msg(msg=iCOMOX_messages.OUT_MSG_Debug_TestConnectivity())

	def NBIOT_UpdateResponse(self, payload):
		common.app.DiagnosticMode.NBIOT_Response = payload.decode("utf-8", "replace")   #helpers.string_without(str=payload.decode("utf-8", "replace"), chars_to_remove="\0")
		last_char_visible = self.text_NBIOT_Response.bbox("end-1c")
		common.app.DiagnosticMode.text_NBIOT_Response.insert("end", common.app.DiagnosticMode.NBIOT_Response)
		if last_char_visible:
			self.text_NBIOT_Response.see("end")

	def Enable(self, enable):
		state = "normal" if enable else "disabled"
		self.label_NBIOT_ATCommand.configure(state=state)
		self.combobox_NBIOT_ATCommand.configure(state=state)
		self.label_NBIOT_Timeout.configure(state=state)
		self.entry_NBIOT_Timeout.configure(state=state)
		self.button_NBIOT_Send_ATCommand.configure(state=state)
		self.button_NBIOT_Reset_BG96.configure(state=state)
		self.button_NBIOT_TestConnectivity.configure(state=state)
		self.text_NBIOT_Response.configure(state=state)
		# self.notebook.tab(self.notebook.tabs()[1], state=state)

class StatusBar(tk.Frame):
	def __init__(self, parent, *args, **kwargs):
		tk.Frame.__init__(self, *args, **kwargs)
		self.parent = parent
		self.status = tk.Label(self.parent, bd=0, relief=tk.SUNKEN, anchor=tk.W)
		self.status.grid(row=1, column=1, sticky=tk.W)
		self.progressbar = ttk.Progressbar(self.parent, orient="horizontal", length=200, mode="determinate", value=0, maximum=1)
		self.progressbar.grid(row=1, column=2, sticky=tk.E)
		self.progressbar_hide()
		self.parent.grid_columnconfigure(1, weight=1)   # <- without it, the sticky option in grid() has no effect

	def set(self, text):
		text = datetime.datetime.now().strftime("%d/%m/%Y, %H:%M:%S - ") + text
		self.status.config(text=text)
		self.status.update_idletasks()

	def clear(self):
		self.status.config(text="")
		self.status.update_idletasks()

	def progressbar_show(self):
		self.progressbar.grid()

	def progressbar_hide(self):
		self.progressbar.grid_remove()

	def message_progress(self, packet_index, total_packets):
		self.progressbar_show()
		self.progressbar["value"] = packet_index
		self.progressbar["maximum"] = total_packets
		self.progressbar.update()


def about_window():
	messagebox.showinfo(
		title="About",
		message="iCOMOX Monitor PC Software\nVersion {}\nCopyright © 2020\nShiratech Solutions Ltd.\nhttps://www.shiratech-solutions.com/products/icomox/".format(common_symbols.__VERSION__))


def main():
	root = tk.Tk()      # Create empty main window
	root.withdraw()     # Hide the main window

	myFont = font.Font(name="Verdana 8", family="Verdana", size=8, weight=font.NORMAL) # Needed for the failure procedure to adjust the combobox dropdown width

	# Check if single instance of the application exists. If not then terminate
	if not single_app_instance.IsSingleAppInstance():
		single_app_instance.CloseSingleAppInstanceMarker()
		#messagebox.showinfo("Error", "Application instance already exists")
		helpers.OUT("Application instance already exists")
		sys.exit(0)

	# Fill the empty window with content
	common.app = mainApp(root)
	root.iconbitmap(default=helpers.resource_path("logo_icon.ico"))
	root.wm_title("iCOMOX Monitor version {}".format(common_symbols.__VERSION__))
	menubar = tk.Menu(root)
	filemenu = tk.Menu(menubar, tearoff=0)
	helpmenu = tk.Menu(menubar, tearoff=0)
	menubar.add_cascade(label="File", menu=filemenu)
	menubar.add_cascade(label="Help", menu=helpmenu)
	filemenu.add_command(label="Exit", command=root.destroy)
	helpmenu.add_command(label="About", command=about_window)
	root.config(menu=menubar)
	helpers.centralize_window(root=root)
	root.resizable(0, 0)    # it must be after centralize_main_window(). It locks the main window size

	# Show the main window again
	root.update()
	root.deiconify()

	common.app.drawConnectionState(True)
	common.app.drawLedState(True)
	common.app.EnableButtons(True, True, True, True, True)

	common.app.Configuration.EnableSensorsToSaveToExcel(True)
	common.app.Configuration.FileDir()
	#common.app.Configuration.StateStartStop()
	common.app.Configuration.StateBrowse()     #Enable the browse button

	#about_window()
	# def onClose():
	#     if app.iCOMOX_Data.iCOMOX_Comm is not None:
	#         app.iCOMOX_Data.iCOMOX_Comm.Terminate = True
	#         root.destroy()
	#
	# root.protocol("WM_DELETE_WINDOW", onClose)
	# root.bind('<Escape>', lambda e: root.destroy())

	try:
		helpers.OUT("APPLICATION STARTED")
		common.app.iCOMOX_Data.TcpIp_Comm = TCP_connectivity.class_iCOMOX_TcpServer(callback_process_message=common.app.iCOMOX_Data.on_process_messages, callback_state_changed=common.app.iCOMOX_Data.on_tcpip_state_changed)

		common.app.iCOMOX_Data.iCOMOX_Comm = iCOMOX_communication.class_iCOMOX_Communication(
			comm_port="",
			baudrate=125000,
			duration_of_break_us=1000,  #8*12,  # length of 12 bits, 8usec for each (for baudrate = 125Kbps) results in 96usec, we provide 1 msec as a safety margin
			callback_get_msg_size=messages_utils.on_get_in_message_size,
			callback_process_message=common.app.iCOMOX_Data.USB_on_process_messages,
			callback_open=common.app.iCOMOX_Data.USB_on_open,
			callback_close=common.app.iCOMOX_Data.USB_on_close,
			no_msg_read_timeout_sec=iCOMOX_messages.NO_REPORT_TIMEOUT_SEC,
			callback_no_msg_read_timeout=common.app.iCOMOX_Data.on_no_activity,
		)

		common.app.iCOMOX_Data.Dongle_Comm = iCOMOX_over_Dongle_Communication.class_iCOMOX_over_Dongle_Communication(
			comm_port="",
			callback_connection_state=common.app.iCOMOX_Data.on_dongle_connection_state_changed,
			callback_process_message=common.app.iCOMOX_Data.on_process_messages,
			callback_updated_iCOMOX_SMIP_list=common.app.iCOMOX_Data.on_dongle_updated_iCOMOX_list,
			callback_data_sent=None
		)

		common.app.buttonConnectCommand()
		common.app.Configuration.StartPressed()


		# import time
		# max_time = 10
		# start_time = time.time()
		# print(time.time() - start_time)
		# while (time.time() - start_time) < max_time:
		# 	print(time.time() - start_time)
		# 	common.app.Configuration.StopPressed()

		def ClearCommObjects():
			if common.app.iCOMOX_Data is not None:
				if common.app.iCOMOX_Data.TcpIp_Comm is not None:
					common.app.iCOMOX_Data.TcpIp_Comm.Terminated = True
				if common.app.iCOMOX_Data.Dongle_Comm is not None:
					common.app.iCOMOX_Data.Dongle_Comm.Terminate = True
				if common.app.iCOMOX_Data.iCOMOX_Comm is not None:
					common.app.iCOMOX_Data.iCOMOX_Comm.Terminate = True


		def onClose():
			common.app.terminated = True
			ClearCommObjects()
			root.destroy()

		root.protocol("WM_DELETE_WINDOW", onClose)
		root.after_idle(common.app.on_idle_task, root)  # provide root as an argument to common.app.on_idle_task
		root.mainloop()
		common.app.terminated = True
		# common.app.iCOMOX_Data.iCOMOX_Comm.thread.join()

	except Exception as ex:
		if hasattr(ex, 'message'):
			helpers.OUT(ex.message)
			common.app.StatusBar.set(text=ex.message)
			common.app.TopPane.set_led_status_color(1)  # set GUI LED color to red

	else:
		pass
	finally:
		ClearCommObjects()
		# common.app.destroy()
		# common.app.iCOMOX_Data.iCOMOX_Comm = None
		# common.app.StatusBar.set(text="iCOMOX disconnected")
		# common.app.TopPane.set_led_status_color(1)
		helpers.OUT("APPLICATION FINISHED")
		single_app_instance.CloseSingleAppInstanceMarker()

if __name__ == "__main__":
	main()
